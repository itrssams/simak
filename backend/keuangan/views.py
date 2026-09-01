from rest_framework import viewsets, status

from rest_framework.decorators import action, permission_classes, api_view

from rest_framework.response import Response

from rest_framework.permissions import BasePermission, IsAuthenticated, AllowAny, SAFE_METHODS

from rest_framework.viewsets import ModelViewSet

from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from rest_framework.pagination import PageNumberPagination

from rest_framework.exceptions import PermissionDenied, ValidationError

from django.db.models import Sum

from django.db.models.functions import Coalesce

from django.db import connection, transaction, models

from django.db.models.deletion import ProtectedError

from decimal import Decimal, InvalidOperation

from rest_framework.views import APIView

from django.http import HttpResponse

from django.shortcuts import get_object_or_404

from django.utils.html import escape

from django.db.models import Sum, Count, Q, F

from collections import defaultdict

import calendar

import re

from django.utils import timezone

from datetime import datetime, date, time, timedelta

import openpyxl

from openpyxl import Workbook

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from openpyxl.utils import get_column_letter

from fpdf import FPDF

import os

from django.conf import settings

from .models import (
    Akun, Transaksi, Jurnal, JurnalItem,
    Pelanggan, Pemasok,
    Faktur, FakturItem, PembayaranFaktur, AlokasiDana, AlokasiDanaPemakaian,
    IndukPembiayaan, PembiayaanIndukMapping,
    UtangSupplier, PembayaranUtang, DepositVendor,
    Tagihan, TagihanItem, PembayaranTagihan,
    RekeningBank, RiwayatSaldoRekening,
    
    PettyCash, LaporanPenggunaan, Reimbursement, SaldoPettyCash, RiwayatSaldoPettyCash, PengajuanPenambahanSaldo,
)

from .serializers import (
    AkunSerializer, TransaksiSerializer, TransaksiInputSerializer,
    JurnalSerializer, JurnalInputSerializer,
    PelangganSerializer, PemasokSerializer,
    FakturSerializer, FakturInputSerializer,
    PembayaranFakturSerializer, PembayaranFakturInputSerializer,
    AlokasiDanaSerializer,
    IndukPembiayaanSerializer, PembiayaanIndukMappingSerializer,
    UtangSupplierSerializer, PembayaranUtangSerializer, PembayaranUtangInputSerializer, DepositVendorSerializer,
    TagihanSerializer, TagihanInputSerializer,
    PembayaranTagihanSerializer, PembayaranTagihanInputSerializer,
    RekeningBankSerializer, RekeningBankInputSerializer,
    RiwayatSaldoRekeningSerializer, UpdateSaldoSerializer,
    
    PettyCashSerializer, PettyCashInputSerializer,
    LaporanPenggunaanSerializer, LaporanPenggunaanInputSerializer,
    ReimbursementSerializer, ReimbursementInputSerializer, SaldoPettyCashSerializer, RiwayatSaldoPettyCashSerializer,
    PengajuanPenambahanSaldoSerializer, PengajuanPenambahanSaldoInputSerializer,
)

from system.audit import can_view_audit

class OptionalPageNumberPagination(PageNumberPagination):
    page_size_query_param = 'page_size'
    max_page_size = 100

    def get_page_size(self, request):
        if 'page' not in request.query_params and self.page_size_query_param not in request.query_params:
            return None
        if self.page_size_query_param in request.query_params:
            return super().get_page_size(request)
        return 10

class OptionalPaginationMixin:
    pagination_class = OptionalPageNumberPagination

def is_direktur_or_wadir(user):
    return user.is_authenticated and (user.role in ('direktur', 'wakil_direktur') or user.is_superuser)

def is_manajer_or_above(user):
    return user.is_authenticated and (user.role in ('manajer', 'wakil_direktur', 'direktur') or user.is_superuser)

def is_kepala_seksi_or_above(user):
    return user.is_authenticated and (user.role in ('kepala_seksi', 'manajer', 'wakil_direktur', 'direktur') or user.is_superuser)

def is_keuangan(user):
    return user.is_authenticated and (getattr(user, 'is_keuangan', False) or user.is_superuser)

def can_access_catatan_utang_obat_bhp(user):
    return user.is_authenticated and (
        user.is_superuser
        or getattr(user, 'akses_catatan_utang', False)
    )

def is_petty_cash_cashier(user):
    return user.is_authenticated and (getattr(user, 'is_petty_cash_cashier', False) or user.is_superuser)

def is_manajer_keuangan(user):
    return user.is_authenticated and (
        user.is_superuser
        or (getattr(user, 'is_keuangan', False) and user.role in ('manajer', 'wakil_direktur', 'direktur'))
    )

def laporan_unit_label(user):
    if not user:
        return 'Tidak diketahui'
    if user.unit:
        return user.unit.nama
    if getattr(user, 'is_driver', False):
        return 'Driver'
    if getattr(user, 'is_it', False):
        return 'IT'
    role_labels = {
        'direktur': 'Direktur',
        'wakil_direktur': 'Wakil Direktur',
        'manajer': 'Manajer',
        'kepala_seksi': 'Kepala Seksi',
        'karyawan': 'Karyawan Tanpa Unit',
    }
    return role_labels.get(user.role, 'Tanpa Unit')

def user_display_name(user):
    if not user:
        return ''
    return user.get_full_name() or user.username

def generate_nomor_faktur(tanggal):
    from django.db import connection

    prefix = f"{tanggal:%y}"
    max_urut = 0

    # Cek invoice Django
    last_faktur = (
        Faktur.objects
        .filter(nomor_faktur__startswith=prefix)
        .order_by('-nomor_faktur')
        .values_list('nomor_faktur', flat=True)
        .first()
    )

    if last_faktur and str(last_faktur)[2:6].isdigit():
        max_urut = max(max_urut, int(str(last_faktur)[2:6]))

    # Cek invoice lama SIMRS
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT MAX(CAST(SUBSTR(no, 3, 4) AS UNSIGNED))
                FROM rssams.invoice
                WHERE LEFT(no, 2) = %s
            """, [prefix])

            row = cursor.fetchone()

            if row and row[0]:
                max_urut = max(max_urut, int(row[0]))
    except Exception:
        pass

    next_number = max_urut + 1
    nomor = f"{prefix}{next_number:04d}"

    while Faktur.objects.filter(nomor_faktur=nomor).exists():
        next_number += 1
        nomor = f"{prefix}{next_number:04d}"

    return nomor

class IsManajerOrAbovePermission(BasePermission):
    def has_permission(self, request, view):
        return is_manajer_or_above(request.user)

class IsDirekturOrWadirPermission(BasePermission):
    def has_permission(self, request, view):
        return is_direktur_or_wadir(request.user)

class IsKeuanganPermission(BasePermission):
    def has_permission(self, request, view):
        return is_keuangan(request.user)

class IsCatatanUtangObatBhpPermission(BasePermission):
    def has_permission(self, request, view):
        return can_access_catatan_utang_obat_bhp(request.user)

class IsPettyCashSaldoPermission(BasePermission):
    def has_permission(self, request, view):
        return is_manajer_or_above(request.user) or is_petty_cash_cashier(request.user)

class IsKeuanganOrManajerPermission(BasePermission):
    def has_permission(self, request, view):
        return is_keuangan(request.user) or is_manajer_or_above(request.user)


def _normalize_pembiayaan_name(value):
    return ' '.join(str(value or '').strip().lower().split())

def _normalize_logistik_name(value):
    text = str(value or '').strip()
    if not text:
        return ''
    return ' '.join(text.split())

class PembiayaanListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get list of pembiayaan (insurance providers) from rssams.pbiaya with Induk Pembiayaan info"""
        from django.db import connection
        include_inactive = str(request.query_params.get('include_inactive') or '').lower() in ('1', 'true', 'yes')
        search = (request.query_params.get('search') or '').strip()
        induk_filter = request.query_params.get('induk_id')
        unassigned_only = str(request.query_params.get('unassigned') or '').lower() in ('1', 'true', 'yes')

        try:
            with connection.cursor() as cursor:
                where = []
                values = []
                if not include_inactive:
                    where.append("status <> 0")
                if search:
                    where.append("(pembiayaan LIKE %s OR CAST(id_pembiayaan AS CHAR) LIKE %s OR alamat LIKE %s)")
                    needle = f"%{search}%"
                    values.extend([needle, needle, needle])
                where_sql = f"WHERE {' AND '.join(where)}" if where else ""
                cursor.execute(f"""
                    SELECT id_pembiayaan, pembiayaan as nama, alamat, status
                    FROM rssams.pbiaya
                    {where_sql}
                    ORDER BY status DESC, pembiayaan
                """, values)
                columns = [col[0] for col in cursor.description]
                pembiayaan = [dict(zip(columns, row)) for row in cursor.fetchall()]

            # Attach Induk Mapping info
            mappings = {
                str(m.id_pembiayaan): (m.induk_id, m.induk.nama)
                for m in PembiayaanIndukMapping.objects.select_related('induk').all()
            }
            results = []
            for item in pembiayaan:
                str_id = str(item['id_pembiayaan'])
                if str_id in mappings:
                    item['induk_id'] = mappings[str_id][0]
                    item['induk_nama'] = mappings[str_id][1]
                else:
                    item['induk_id'] = None
                    item['induk_nama'] = None

                if unassigned_only and item['induk_id'] is not None:
                    continue
                if induk_filter and str(item['induk_id']) != str(induk_filter):
                    continue
                results.append(item)

            return Response({
                'count': len(results),
                'results': results
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        if not is_keuangan(request.user):
            raise PermissionDenied('Hanya user keuangan yang dapat menambah pembiayaan.')

        nama = (request.data.get('nama') or request.data.get('pembiayaan') or '').strip()
        if not nama:
            return Response({'nama': 'Nama pembiayaan wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        alamat = (request.data.get('alamat') or '').strip()
        try:
            with transaction.atomic():
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT id_pembiayaan, pembiayaan
                        FROM rssams.pbiaya
                    """)
                    existing = next(
                        (
                            row for row in cursor.fetchall()
                            if _normalize_pembiayaan_name(row[1]) == _normalize_pembiayaan_name(nama)
                        ),
                        None,
                    )
                    if existing:
                        return Response({
                            'error': f"Pembiayaan '{existing[1]}' sudah ada dengan ID {existing[0]}. Pilih dari daftar pembiayaan.",
                            'id_pembiayaan': existing[0],
                            'nama': existing[1],
                        }, status=status.HTTP_400_BAD_REQUEST)

                    cursor.execute("""
                        INSERT INTO rssams.pbiaya
                            (pembiayaan, ket, status, apt, r_inap, kode, persen_apt, alamat)
                        VALUES
                            (%s, %s, 1, 'Y', 'Y', 1, 0.00, %s)
                    """, [nama, nama, alamat])
                    cursor.execute("SELECT LAST_INSERT_ID()")
                    id_pembiayaan = int(cursor.fetchone()[0])

            return Response({
                'id_pembiayaan': id_pembiayaan,
                'nama': nama,
                'alamat': alamat,
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class PembiayaanDetailView(APIView):
    permission_classes = [IsKeuanganPermission]

    def patch(self, request, id_pembiayaan):
        nama = (request.data.get('nama') or request.data.get('pembiayaan') or '').strip()
        alamat = (request.data.get('alamat') or '').strip()
        status_value = request.data.get('status')

        try:
            with transaction.atomic():
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT id_pembiayaan, pembiayaan, alamat, status
                        FROM rssams.pbiaya
                        WHERE id_pembiayaan = %s
                        LIMIT 1
                    """, [id_pembiayaan])
                    current = cursor.fetchone()
                    if not current:
                        return Response({'error': 'Pembiayaan tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)

                    next_nama = nama or current[1]
                    if not next_nama:
                        return Response({'nama': 'Nama pembiayaan wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

                    cursor.execute("""
                        SELECT id_pembiayaan, pembiayaan
                        FROM rssams.pbiaya
                        WHERE id_pembiayaan <> %s
                    """, [id_pembiayaan])
                    duplicate = next(
                        (
                            row for row in cursor.fetchall()
                            if _normalize_pembiayaan_name(row[1]) == _normalize_pembiayaan_name(next_nama)
                        ),
                        None,
                    )
                    if duplicate:
                        return Response({
                            'error': f"Pembiayaan '{duplicate[1]}' sudah ada dengan ID {duplicate[0]}.",
                        }, status=status.HTTP_400_BAD_REQUEST)

                    updates = ["pembiayaan = %s", "ket = %s", "alamat = %s"]
                    values = [next_nama, next_nama, alamat]
                    next_status = int(current[3] or 0)
                    if status_value is not None:
                        next_status = 1 if str(status_value) not in ('0', 'false', 'False') else 0
                        updates.append("status = %s")
                        values.append(next_status)
                    values.append(id_pembiayaan)
                    cursor.execute(f"""
                        UPDATE rssams.pbiaya
                        SET {', '.join(updates)}
                        WHERE id_pembiayaan = %s
                    """, values)

            return Response({
                'id_pembiayaan': int(id_pembiayaan),
                'nama': next_nama,
                'alamat': alamat,
                'status': next_status,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def delete(self, request, id_pembiayaan):
        try:
            with connection.cursor() as cursor:
                cursor.execute("UPDATE rssams.pbiaya SET status = 0 WHERE id_pembiayaan = %s", [id_pembiayaan])
                if cursor.rowcount == 0:
                    return Response({'error': 'Pembiayaan tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

KUNJUNGAN_TYPE_FILTERS = {
    'rawat_jalan': "substr(a.j_lay,18,1)='1'",
    'rawat_inap': "substr(a.j_lay,17,1)='1'",
    'ugd': "substr(a.j_lay,16,1)='1'",
    'vk': "substr(a.j_lay,19,1)='1'",
    'ok': "substr(a.j_lay,20,1)='1'",
}

KUNJUNGAN_TYPE_LABELS = {
    'rawat_jalan': 'Rawat Jalan',
    'rawat_inap': 'Rawat Inap',
    'ugd': 'UGD',
    'vk': 'VK',
    'ok': 'OK',
}

KUNJUNGAN_TOTAL_SQL = """
    COALESCE(a.adm,0)+COALESCE(a.jasa,0)+COALESCE(a.farmasi,0)+COALESCE(a.tindakan,0)+
    COALESCE(a.fisio,0)+COALESCE(a.lab,0)+COALESCE(a.lab_pa,0)+COALESCE(a.kamar,0)+
    COALESCE(a.rad,0)+COALESCE(a.bhp,0)+COALESCE(a.lainnya,0)+COALESCE(a.ambulan,0)+COALESCE(a.alat,0)
"""

def _dict_fetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def _decimal_from_row(row, key):
    return Decimal(str(row.get(key) or 0))

def _get_pembiayaan_name(id_pembiayaan):
    if not id_pembiayaan:
        return ''
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT pembiayaan
            FROM rssams.pbiaya
            WHERE id_pembiayaan = %s AND status <> 0
            LIMIT 1
        """, [id_pembiayaan])
        row = cursor.fetchone()
    return row[0] if row else ''

def _legacy_kunjungan_where(params):
    kunjungan_type = params.get('jenis') or 'semua'
    where = []
    values = []
    if kunjungan_type != 'semua':
        where.append(KUNJUNGAN_TYPE_FILTERS.get(kunjungan_type, KUNJUNGAN_TYPE_FILTERS['rawat_jalan']))

    search = (params.get('search') or '').strip()
    if search:
        where.append("(a.no LIKE %s OR a.noreg LIKE %s OR b.nama LIKE %s OR c.pembiayaan LIKE %s)")
        needle = f"%{search}%"
        values.extend([needle, needle, needle, needle])

    id_pembiayaan = (params.get('id_pembiayaan') or '').strip()
    if id_pembiayaan == 'non_bpjs':
        where.append("(c.pembiayaan IS NULL OR c.pembiayaan NOT LIKE %s)")
        values.append('%BPJS%')
    elif id_pembiayaan:
        where.append("a.id_pembiayaan = %s")
        values.append(id_pembiayaan)
    elif not search:
        # Hanya tampilkan kunjungan dari Asuransi (exclude Swadana & BPJS) bila tidak sedang mencari
        where.append("(a.id_pembiayaan != 1 AND (c.pembiayaan IS NULL OR (c.pembiayaan NOT LIKE %s AND LOWER(c.pembiayaan) NOT LIKE %s)))")
        values.extend(['%BPJS%', '%swadana%'])

    dari = (params.get('dari') or '').strip()
    if dari:
        where.append("a.tgl_masuk >= %s")
        values.append(f"{dari} 00:00:00" if len(dari) == 10 else dari)

    sampai = (params.get('sampai') or '').strip()
    if sampai:
        where.append("a.tgl_masuk <= %s")
        values.append(f"{sampai} 23:59:59" if len(sampai) == 10 else sampai)

    done = (params.get('done') or '').strip()
    if done == '1':
        where.append("a.cek = 1")
    elif done == '0':
        where.append("a.cek = 0")

    invoice_status = (params.get('invoice_status') or '').strip()
    if invoice_status == 'belum':
        where.append("(e.no_invoice IS NULL OR e.no_invoice = '')")
    elif invoice_status == 'sudah':
        where.append("(e.no_invoice IS NOT NULL AND e.no_invoice <> '')")

    return " AND ".join(where) if where else "1=1", values, kunjungan_type

def _detect_type_from_j_lay(j_lay):
    value = str(j_lay or '')
    if len(value) >= 20 and value[19:20] == '1':
        return 'OK'
    if len(value) >= 19 and value[18:19] == '1':
        return 'VK'
    if len(value) >= 18 and value[17:18] == '1':
        return 'Rawat Jalan'
    if len(value) >= 17 and value[16:17] == '1':
        return 'Rawat Inap'
    if len(value) >= 16 and value[15:16] == '1':
        return 'UGD'
    return 'Kunjungan'

class KunjunganInvoiceView(APIView):
    permission_classes = [IsKeuanganPermission]

    def _fetch_kunjungan_for_invoice(self, cursor, nomor_kunjungan):
        placeholders = ','.join(['%s'] * len(nomor_kunjungan))
        cursor.execute(f"""
            SELECT
                a.no, a.id_pembiayaan, c.pembiayaan AS nama_pembiayaan,
                a.cek, IFNULL(e.no_invoice, '') AS no_invoice,
                a.adm, a.jasa, a.farmasi, a.tindakan, a.fisio, a.lab,
                a.lab_pa, a.rad, a.kamar, a.bhp, a.lainnya, a.ambulan, a.alat,
                ({KUNJUNGAN_TOTAL_SQL}) AS total_biaya
            FROM rssams.kunjung a
            LEFT JOIN rssams.pbiaya c ON a.id_pembiayaan = c.id_pembiayaan
            INNER JOIN rssams.verif_kunjung e ON a.no = e.no
            WHERE a.no IN ({placeholders})
        """, nomor_kunjungan)
        return _dict_fetchall(cursor)

    def _sum_kunjungan_totals(self, rows):
        totals = defaultdict(lambda: Decimal('0'))
        for row in rows:
            totals['adm'] += _decimal_from_row(row, 'adm')
            totals['jasa'] += _decimal_from_row(row, 'jasa')
            totals['farmasi'] += _decimal_from_row(row, 'farmasi')
            totals['tindakan'] += _decimal_from_row(row, 'tindakan')
            totals['fisio'] += _decimal_from_row(row, 'fisio')
            totals['lab'] += _decimal_from_row(row, 'lab') + _decimal_from_row(row, 'lab_pa')
            totals['rad'] += _decimal_from_row(row, 'rad')
            totals['kamar'] += _decimal_from_row(row, 'kamar')
            totals['bhp'] += _decimal_from_row(row, 'bhp')
            totals['lainnya'] += _decimal_from_row(row, 'lainnya')
            totals['ambulan'] += _decimal_from_row(row, 'ambulan')
            totals['alat'] += _decimal_from_row(row, 'alat')
        return totals

    def _refresh_faktur_totals(self, faktur):
        rows = get_invoice_kunjungan_rows(faktur)
        if not rows:
            return
        totals = self._sum_kunjungan_totals(rows)
        for field, value in totals.items():
            setattr(faktur, field, value)
        total_real = sum(totals.values())
        faktur.total_real_rs = total_real
        if faktur.is_cob and faktur.tanggungan_bpjs:
            faktur.total_tagihan = max(Decimal('0'), Decimal(str(total_real)) - Decimal(str(faktur.tanggungan_bpjs)))
        else:
            faktur.total_tagihan = total_real
        faktur.save()

    def get(self, request):
        from django.db import connection

        detail_no = (request.query_params.get('no') or '').strip()
        invoice_search = (request.query_params.get('invoice_search') or '').strip()
        try:
            with connection.cursor() as cursor:
                if invoice_search:
                    qs = (
                        Faktur.objects
                        .select_related('pelanggan')
                        .filter(
                            Q(nomor_faktur__icontains=invoice_search) |
                            Q(nama_pembiayaan__icontains=invoice_search) |
                            Q(pelanggan__nama__icontains=invoice_search)
                        )
                        .exclude(status='batal')
                        .order_by('-nomor_faktur')[:10]
                    )
                    return Response(FakturSerializer(qs, many=True).data, status=status.HTTP_200_OK)

                if detail_no:
                    cursor.execute(f"""
                        SELECT
                            a.no, a.noreg, b.nama, b.sex, b.telp, DATE(a.tgl_masuk) AS tgl_masuk,
                            DATE(a.tgl_keluar) AS tgl_keluar, a.no_sep, a.no_jam, a.id_pembiayaan,
                            c.pembiayaan AS nama_pembiayaan, a.cek,
                            IFNULL(e.no_invoice, '') AS no_invoice,
                            a.adm, a.jasa, a.farmasi, a.tindakan, a.fisio, a.lab,
                            a.lab_pa, a.rad, a.kamar, a.bhp, a.lainnya, a.ambulan, a.alat,
                            a.dp3, a.jmlbyr, ({KUNJUNGAN_TOTAL_SQL}) AS total_biaya
                        FROM rssams.kunjung a
                        INNER JOIN rssams.regpasien b ON a.noreg = b.noreg
                        LEFT JOIN rssams.pbiaya c ON a.id_pembiayaan = c.id_pembiayaan
                        INNER JOIN rssams.verif_kunjung e ON a.no = e.no
                        WHERE a.no = %s
                        LIMIT 1
                    """, [detail_no])
                    rows = _dict_fetchall(cursor)
                    if not rows:
                        return Response({'error': 'Kunjungan tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
                    row = rows[0]
                    row['status_done'] = bool(row.get('cek'))
                    row['status_invoice'] = 'sudah' if row.get('no_invoice') else 'belum'
                    row['jenis_label'] = self._detect_type(row.get('no'))
                    return Response(row, status=status.HTTP_200_OK)

                where_sql, values, kunjungan_type = _legacy_kunjungan_where(request.query_params)
                page = max(int(request.query_params.get('page') or 1), 1)
                page_size = min(max(int(request.query_params.get('page_size') or 10), 1), 100)
                offset = (page - 1) * page_size

                search = (request.query_params.get('search') or '').strip()
                count_base_sql = f"""
                    FROM rssams.kunjung a
                    {'INNER JOIN rssams.regpasien b ON a.noreg = b.noreg' if search else ''}
                    LEFT JOIN rssams.pbiaya c ON a.id_pembiayaan = c.id_pembiayaan
                    INNER JOIN rssams.verif_kunjung e ON a.no = e.no
                    WHERE {where_sql}
                """
                cursor.execute(f"SELECT COUNT(*) AS total {count_base_sql}", values)
                total = cursor.fetchone()[0]

                list_base_sql = f"""
                    FROM rssams.kunjung a
                    INNER JOIN rssams.regpasien b ON a.noreg = b.noreg
                    LEFT JOIN rssams.pbiaya c ON a.id_pembiayaan = c.id_pembiayaan
                    INNER JOIN rssams.verif_kunjung e ON a.no = e.no
                    WHERE {where_sql}
                """
                cursor.execute(f"""
                    SELECT
                        a.no, a.noreg, b.nama, b.sex, DATE(a.tgl_masuk) AS tgl_masuk,
                        DATE(a.tgl_keluar) AS tgl_keluar, a.id_pembiayaan,
                        c.pembiayaan AS nama_pembiayaan, a.cek, a.j_lay,
                        IFNULL(e.no_invoice, '') AS no_invoice,
                        ({KUNJUNGAN_TOTAL_SQL}) AS total_biaya,
                        a.dp3, a.jmlbyr
                    {list_base_sql}
                    ORDER BY a.tgl_masuk DESC, a.no DESC
                    LIMIT %s OFFSET %s
                """, [*values, page_size, offset])
                rows = _dict_fetchall(cursor)
            for row in rows:
                row['jenis'] = kunjungan_type
                row['jenis_label'] = (
                    _detect_type_from_j_lay(row.get('j_lay'))
                    if kunjungan_type == 'semua'
                    else KUNJUNGAN_TYPE_LABELS.get(kunjungan_type, 'Rawat Jalan')
                )
                row['status_done'] = bool(row.get('cek'))
                row['status_invoice'] = 'sudah' if row.get('no_invoice') else 'belum'
            return Response({'count': total, 'results': rows}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        from django.db import connection

        action_name = (request.data.get('action') or '').strip()
        nomor_kunjungan = request.data.get('nomor_kunjungan') or request.data.get('selected') or []
        if isinstance(nomor_kunjungan, str):
            nomor_kunjungan = [nomor_kunjungan]
        nomor_kunjungan = [str(no).strip() for no in nomor_kunjungan if str(no).strip()]
        if not nomor_kunjungan:
            return Response({'error': 'Pilih minimal satu kunjungan.'}, status=status.HTTP_400_BAD_REQUEST)

        if action_name == 'append':
            invoice_number = str(request.data.get('nomor_faktur') or '').strip()
            if not invoice_number:
                return Response({'nomor_faktur': 'Nomor invoice wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                faktur = Faktur.objects.get(nomor_faktur=invoice_number)
            except Faktur.DoesNotExist:
                return Response({'nomor_faktur': 'Invoice tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
            if faktur.status == 'batal':
                return Response({'error': 'Kunjungan tidak bisa ditambahkan ke invoice batal.'}, status=status.HTTP_400_BAD_REQUEST)

            placeholders = ','.join(['%s'] * len(nomor_kunjungan))
            try:
                with transaction.atomic():
                    with connection.cursor() as cursor:
                        rows = self._fetch_kunjungan_for_invoice(cursor, nomor_kunjungan)
                        found = {str(row['no']) for row in rows}
                        missing = [no for no in nomor_kunjungan if no not in found]
                        if missing:
                            return Response({'error': f"Kunjungan tidak ditemukan: {', '.join(missing)}"}, status=status.HTTP_400_BAD_REQUEST)
                        not_done = [str(row['no']) for row in rows if not row.get('cek')]
                        if not_done:
                            return Response({'error': f"Transaksi belum done: {', '.join(not_done)}"}, status=status.HTTP_400_BAD_REQUEST)
                        invoiced = [str(row['no']) for row in rows if row.get('no_invoice')]
                        if invoiced:
                            return Response({'error': f"Kunjungan sudah masuk invoice: {', '.join(invoiced)}"}, status=status.HTTP_400_BAD_REQUEST)
                        no_charge = [str(row['no']) for row in rows if _decimal_from_row(row, 'total_biaya') <= 0]
                        if no_charge:
                            return Response({'error': f"Kunjungan belum memiliki biaya: {', '.join(no_charge)}"}, status=status.HTTP_400_BAD_REQUEST)
                        cursor.execute(
                            f"UPDATE rssams.verif_kunjung SET no_invoice=%s WHERE no IN ({placeholders})",
                            [faktur.nomor_faktur, *nomor_kunjungan],
                        )
                    self._refresh_faktur_totals(faktur)
                return Response(FakturSerializer(faktur).data, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        tanggal = request.data.get('tanggal') or timezone.localdate().isoformat()
        try:
            tanggal = datetime.strptime(str(tanggal), '%Y-%m-%d').date()
        except ValueError:
            return Response({'tanggal': 'Format tanggal harus YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        placeholders = ','.join(['%s'] * len(nomor_kunjungan))
        try:
            with connection.cursor() as cursor:
                rows = self._fetch_kunjungan_for_invoice(cursor, nomor_kunjungan)

            found = {str(row['no']) for row in rows}
            missing = [no for no in nomor_kunjungan if no not in found]
            if missing:
                return Response({'error': f"Kunjungan tidak ditemukan: {', '.join(missing)}"}, status=status.HTTP_400_BAD_REQUEST)
            not_done = [str(row['no']) for row in rows if not row.get('cek')]
            if not_done:
                return Response({'error': f"Transaksi belum done: {', '.join(not_done)}"}, status=status.HTTP_400_BAD_REQUEST)
            invoiced = [str(row['no']) for row in rows if row.get('no_invoice')]
            if invoiced:
                return Response({'error': f"Kunjungan sudah masuk invoice: {', '.join(invoiced)}"}, status=status.HTTP_400_BAD_REQUEST)

            totals = self._sum_kunjungan_totals(rows)

            id_pembiayaan = str(request.data.get('id_pembiayaan') or '').strip()
            if not id_pembiayaan:
                return Response({'id_pembiayaan': 'Pembiayaan invoice wajib dipilih.'}, status=status.HTTP_400_BAD_REQUEST)

            nama_pembiayaan = _get_pembiayaan_name(id_pembiayaan)
            if not nama_pembiayaan:
                return Response({'id_pembiayaan': 'Pembiayaan invoice tidak ditemukan atau tidak aktif.'}, status=status.HTTP_400_BAD_REQUEST)

            visit_pembiayaan = sorted({
                str(row.get('nama_pembiayaan') or row.get('id_pembiayaan') or 'Tanpa Pembiayaan')
                for row in rows
            })
            nomor_faktur = generate_nomor_faktur(tanggal)
            jenis = request.data.get('jenis') or 'Kunjungan Pasien'
            periode = request.data.get('periode') or tanggal.strftime('%B %Y')
            beban = request.data.get('beban') or nama_pembiayaan or 'PEMBIAYAAN'
            default_keterangan = f"Dibuat dari kunjungan: {', '.join(nomor_kunjungan)}"
            if len(visit_pembiayaan) > 1:
                default_keterangan += f". Pembiayaan asal kunjungan: {', '.join(visit_pembiayaan)}."
            keterangan = request.data.get('keterangan') or default_keterangan

            is_cob = bool(request.data.get('is_cob'))
            try:
                tanggungan_bpjs = Decimal(str(request.data.get('tanggungan_bpjs') or '0'))
            except Exception:
                tanggungan_bpjs = Decimal('0')

            total_real_rs = sum(totals.values())
            if is_cob and tanggungan_bpjs > 0:
                total_tagihan = max(Decimal('0'), total_real_rs - tanggungan_bpjs)
            else:
                total_tagihan = total_real_rs

            with transaction.atomic():
                faktur = Faktur.objects.create(
                    nomor_faktur=nomor_faktur,
                    tanggal=tanggal,
                    jatuh_tempo=tanggal,
                    id_pembiayaan=id_pembiayaan,
                    nama_pembiayaan=nama_pembiayaan,
                    jenis=jenis,
                    periode=periode,
                    beban=beban,
                    keterangan=keterangan,
                    is_cob=is_cob,
                    tanggungan_bpjs=tanggungan_bpjs,
                    total_real_rs=total_real_rs,
                    total_tagihan=total_tagihan,
                    created_by=request.user,
                    **totals,
                )
                with connection.cursor() as cursor:
                    cursor.execute(
                        f"UPDATE rssams.verif_kunjung SET no_invoice=%s WHERE no IN ({placeholders})",
                        [faktur.nomor_faktur, *nomor_kunjungan],
                    )
            return Response(FakturSerializer(faktur).data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)    

    def delete(self, request):
        from django.db import connection

        invoice_number = str(request.data.get('nomor_faktur') or request.query_params.get('nomor_faktur') or '').strip()
        no_kunjungan = str(request.data.get('no') or request.query_params.get('no') or '').strip()
        if not invoice_number or not no_kunjungan:
            return Response({'error': 'Nomor invoice dan nomor kunjungan wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            faktur = Faktur.objects.get(nomor_faktur=invoice_number)
        except Faktur.DoesNotExist:
            return Response({'error': 'Invoice tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
        if faktur.tgl_kirim:
            return Response({'error': 'Kunjungan tidak bisa dihapus dari invoice yang sudah dikirim.'}, status=status.HTTP_400_BAD_REQUEST)
        if faktur.pembayaran.exists():
            return Response({'error': 'Kunjungan tidak bisa dihapus dari invoice yang sudah memiliki pembayaran atau pengajuan pembayaran.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                with connection.cursor() as cursor:
                    cursor.execute(
                        "SELECT IFNULL(no_invoice, '') FROM rssams.verif_kunjung WHERE no=%s LIMIT 1",
                        [no_kunjungan],
                    )
                    row = cursor.fetchone()
                    if not row:
                        return Response({'error': 'Kunjungan tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
                    if str(row[0] or '') != faktur.nomor_faktur:
                        return Response({'error': 'Kunjungan tidak terdaftar pada invoice ini.'}, status=status.HTTP_400_BAD_REQUEST)
                    cursor.execute(
                        "UPDATE rssams.verif_kunjung SET no_invoice='' WHERE no=%s AND no_invoice=%s",
                        [no_kunjungan, faktur.nomor_faktur],
                    )
                    cursor.execute(
                        "SELECT COUNT(*) FROM rssams.verif_kunjung WHERE no_invoice=%s",
                        [faktur.nomor_faktur],
                    )
                    remaining = cursor.fetchone()[0]
                if remaining:
                    self._refresh_faktur_totals(faktur)
                else:
                    faktur.adm = faktur.jasa = faktur.farmasi = faktur.tindakan = Decimal('0')
                    faktur.fisio = faktur.lab = faktur.rad = faktur.kamar = Decimal('0')
                    faktur.bhp = faktur.lainnya = faktur.ambulan = faktur.alat = Decimal('0')
                    faktur.save()
            return Response(FakturSerializer(faktur).data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _detect_type(self, no):
        from django.db import connection
        with connection.cursor() as cursor:
            cursor.execute("SELECT j_lay FROM rssams.kunjung WHERE no=%s LIMIT 1", [no])
            row = cursor.fetchone()
        j_lay = row[0] if row else ''
        if len(j_lay) >= 17 and j_lay[16:17] == '1':
            return 'Rawat Inap'
        if len(j_lay) >= 15 and j_lay[15:16] == '1':
            return 'UGD'
        if len(j_lay) >= 18 and j_lay[18:19] == '1':
            return 'VK'
        if len(j_lay) >= 19 and j_lay[19:20] == '1':
            return 'OK'
        return 'Rawat Jalan'

class InvoiceDashboardView(APIView):
    permission_classes = [IsKeuanganOrManajerPermission]

    def get(self, request):
        today = timezone.localdate()
        active_qs = Faktur.objects.exclude(status='batal')
        dari = request.query_params.get('dari')
        sampai = request.query_params.get('sampai')
        if dari:
            active_qs = active_qs.filter(tanggal__gte=dari)
        if sampai:
            active_qs = active_qs.filter(tanggal__lte=sampai)
        grouped = list(active_qs.values('id_pembiayaan').annotate(
            tagihan_total=Coalesce(Sum('total_tagihan'), Decimal('0')),
            dibayar_total=Coalesce(Sum('total_dibayar'), Decimal('0')),
            invoice_count=Count('id'),
            belum_bayar_count=Count('id', filter=Q(status='belum_bayar')),
            sebagian_count=Count('id', filter=Q(status='bayar_sebagian')),
            lunas_count=Count('id', filter=Q(status='lunas')),
            overdue_count=Count('id', filter=Q(jatuh_tempo__lt=today) & Q(total_tagihan__gt=F('total_dibayar'))),
        ))
        pending_qs = PembayaranFaktur.objects.filter(status_verifikasi='menunggu').exclude(faktur__status='batal')
        if dari:
            pending_qs = pending_qs.filter(faktur__tanggal__gte=dari)
        if sampai:
            pending_qs = pending_qs.filter(faktur__tanggal__lte=sampai)
        pending_by_pembiayaan = {
            str(row['faktur__id_pembiayaan'] or ''): row['total']
            for row in pending_qs
            .values('faktur__id_pembiayaan')
            .annotate(total=Coalesce(Sum('jumlah'), Decimal('0')))
        }
        pembiayaan_names = {}
        ids = [str(item['id_pembiayaan']) for item in grouped if item['id_pembiayaan']]
        if ids:
            from django.db import connection
            placeholders = ','.join(['%s'] * len(ids))
            try:
                with connection.cursor() as cursor:
                    cursor.execute(
                        f"SELECT id_pembiayaan, pembiayaan FROM rssams.pbiaya WHERE id_pembiayaan IN ({placeholders})",
                        ids,
                    )
                    pembiayaan_names = {str(row[0]): row[1] for row in cursor.fetchall()}
            except Exception:
                pembiayaan_names = {}

        rows = []
        for row in grouped:
            total_tagihan = Decimal(row['tagihan_total'] or 0)
            total_dibayar = Decimal(row['dibayar_total'] or 0)
            sisa = total_tagihan - total_dibayar
            key = str(row['id_pembiayaan'] or '')
            rows.append({
                'id_pembiayaan': row['id_pembiayaan'] or '',
                'nama_pembiayaan': pembiayaan_names.get(key) or 'Tanpa Pembiayaan',
                'total_tagihan': total_tagihan,
                'total_dibayar': total_dibayar,
                'sisa_piutang': sisa,
                'pending_verifikasi': pending_by_pembiayaan.get(key, Decimal('0')),
                'invoice_count': row['invoice_count'],
                'belum_bayar_count': row['belum_bayar_count'],
                'sebagian_count': row['sebagian_count'],
                'lunas_count': row['lunas_count'],
                'overdue_count': row['overdue_count'],
                'collection_rate': float((total_dibayar / total_tagihan * 100) if total_tagihan else 0),
            })
        rows.sort(key=lambda item: item['sisa_piutang'], reverse=True)

        totals = active_qs.aggregate(
            tagihan_total=Coalesce(Sum('total_tagihan'), Decimal('0')),
            dibayar_total=Coalesce(Sum('total_dibayar'), Decimal('0')),
            invoice_count=Count('id'),
            belum_bayar_count=Count('id', filter=Q(status='belum_bayar')),
            sebagian_count=Count('id', filter=Q(status='bayar_sebagian')),
            lunas_count=Count('id', filter=Q(status='lunas')),
            overdue_count=Count('id', filter=Q(jatuh_tempo__lt=today) & Q(total_tagihan__gt=F('total_dibayar'))),
        )
        pending_total = pending_qs.aggregate(
            total=Coalesce(Sum('jumlah'), Decimal('0')),
            count=Count('id'),
        )

        aging = {
            'belum_jatuh_tempo': Decimal('0'),
            'hari_1_30': Decimal('0'),
            'hari_31_60': Decimal('0'),
            'hari_lebih_60': Decimal('0'),
        }
        for invoice in active_qs.exclude(status='lunas').values('jatuh_tempo', 'total_tagihan', 'total_dibayar'):
            sisa = Decimal(invoice['total_tagihan'] or 0) - Decimal(invoice['total_dibayar'] or 0)
            if sisa <= 0:
                continue
            jatuh_tempo = invoice['jatuh_tempo']
            if not jatuh_tempo or jatuh_tempo >= today:
                aging['belum_jatuh_tempo'] += sisa
                continue
            days = (today - jatuh_tempo).days
            if days <= 30:
                aging['hari_1_30'] += sisa
            elif days <= 60:
                aging['hari_31_60'] += sisa
            else:
                aging['hari_lebih_60'] += sisa

        total_tagihan = Decimal(totals['tagihan_total'] or 0)
        total_dibayar = Decimal(totals['dibayar_total'] or 0)
        sisa_piutang = total_tagihan - total_dibayar
        response = {
            'summary': {
                'total_tagihan': total_tagihan,
                'total_dibayar': total_dibayar,
                'sisa_piutang': sisa_piutang,
                'pending_verifikasi': pending_total['total'] or Decimal('0'),
                'pending_verifikasi_count': pending_total['count'],
                'invoice_count': totals['invoice_count'],
                'pembiayaan_count': len(rows),
                'belum_bayar_count': totals['belum_bayar_count'],
                'sebagian_count': totals['sebagian_count'],
                'lunas_count': totals['lunas_count'],
                'overdue_count': totals['overdue_count'],
                'collection_rate': float((total_dibayar / total_tagihan * 100) if total_tagihan else 0),
            },
            'aging': aging,
            'pembiayaan': rows,
            'top_piutang': rows[:8],
        }
        return Response(response, status=status.HTTP_200_OK)

class InvoiceVerificationView(APIView):
    permission_classes = [IsKeuanganOrManajerPermission]

    def get(self, request):
        search = (request.query_params.get('search') or '').strip()
        dari = (request.query_params.get('dari') or '').strip()
        sampai = (request.query_params.get('sampai') or '').strip()
        status_filter = (request.query_params.get('status') or 'menunggu').strip()

        qs = (
            PembayaranFaktur.objects
            .exclude(faktur__status='batal')
            .select_related('faktur', 'created_by', 'verified_by')
        )
        if status_filter == 'history':
            qs = qs.exclude(status_verifikasi='menunggu').order_by('-verified_at', '-created_at')
        elif status_filter in ('terverifikasi', 'dibatalkan', 'ditolak'):
            qs = qs.filter(status_verifikasi=status_filter).order_by('-verified_at', '-created_at')
        else:
            qs = qs.filter(status_verifikasi='menunggu').order_by('-created_at')

        if search:
            qs = qs.filter(
                Q(faktur__nomor_faktur__icontains=search) |
                Q(faktur__nama_pembiayaan__icontains=search) |
                Q(faktur__id_pembiayaan__icontains=search) |
                Q(keterangan__icontains=search) |
                Q(created_by__username__icontains=search) |
                Q(verified_by__username__icontains=search)
            )
        if dari:
            qs = qs.filter(tanggal__gte=dari)
        if sampai:
            qs = qs.filter(tanggal__lte=sampai)

        page = max(int(request.query_params.get('page') or 1), 1)
        page_size = min(max(int(request.query_params.get('page_size') or 10), 1), 100)
        total = qs.count()
        start = (page - 1) * page_size
        payments = qs[start:start + page_size]

        results = []
        for pay in payments:
            faktur = pay.faktur
            results.append({
                'id': pay.id,
                'tanggal': pay.tanggal,
                'jumlah': pay.jumlah,
                'metode': pay.metode,
                'keterangan': pay.keterangan,
                'created_at': pay.created_at,
                'created_by_name': getattr(pay.created_by, 'username', '') or '-',
                'verified_at': pay.verified_at,
                'verified_by_name': getattr(pay.verified_by, 'username', '') or '-',
                'status_verifikasi': pay.status_verifikasi,
                'status_verifikasi_label': pay.get_status_verifikasi_display(),
                'faktur': {
                    'id': faktur.id,
                    'nomor_faktur': faktur.nomor_faktur,
                    'tanggal': faktur.tanggal,
                    'id_pembiayaan': faktur.id_pembiayaan,
                    'nama_pembiayaan': faktur.nama_pembiayaan,
                    'total_tagihan': faktur.total_tagihan,
                    'total_dibayar': faktur.total_dibayar,
                    'sisa_tagihan': faktur.sisa_tagihan,
                    'status': faktur.status,
                    'status_label': faktur.get_status_display(),
                },
            })

        return Response({
            'count': total,
            'page': page,
            'page_size': page_size,
            'results': results,
        }, status=status.HTTP_200_OK)

class AkunViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset           = Akun.objects.filter(is_active=True)
    serializer_class   = AkunSerializer
    permission_classes = [IsManajerOrAbovePermission]

class PelangganViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset           = Pelanggan.objects.all()
    serializer_class   = PelangganSerializer
    permission_classes = [IsManajerOrAbovePermission]

    def get_queryset(self):
        qs     = super().get_queryset()
        search = self.request.query_params.get('search')
        tipe   = self.request.query_params.get('tipe')
        if search: qs = qs.filter(nama__icontains=search) | qs.filter(kode__icontains=search)
        if tipe:   qs = qs.filter(tipe=tipe)
        return qs

class PemasokViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset           = Pemasok.objects.all()
    serializer_class   = PemasokSerializer
    permission_classes = [IsManajerOrAbovePermission]

    def get_queryset(self):
        qs     = super().get_queryset()
        search = self.request.query_params.get('search')
        tipe   = self.request.query_params.get('tipe')
        if search: qs = qs.filter(nama__icontains=search) | qs.filter(kode__icontains=search)
        if tipe:   qs = qs.filter(tipe=tipe)
        return qs

class JurnalViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset           = Jurnal.objects.prefetch_related('items__akun').select_related('created_by').all()
    permission_classes = [IsManajerOrAbovePermission]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return JurnalInputSerializer
        return JurnalSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'], url_path='posting')
    def posting(self, request, pk=None):
        jurnal = self.get_object()
        if not jurnal.is_balanced:
            return Response({'error': f'Jurnal tidak seimbang. Debit: {jurnal.total_debit}, Kredit: {jurnal.total_kredit}'}, status=status.HTTP_400_BAD_REQUEST)
        jurnal.status = 'posted'
        jurnal.save()
        return Response({'message': 'Jurnal berhasil diposting.'})

    @action(detail=True, methods=['post'], url_path='unpost')
    def unpost(self, request, pk=None):
        jurnal = self.get_object()
        jurnal.status = 'draft'
        jurnal.save()
        return Response({'message': 'Jurnal dikembalikan ke draft.'})

class TransaksiViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset           = Transaksi.objects.select_related('akun', 'created_by').all()
    permission_classes = [IsManajerOrAbovePermission]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return TransaksiInputSerializer
        return TransaksiSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        faktur = self.get_object()
        if faktur.pembayaran.exists():
            return Response({'error': 'Invoice yang sudah memiliki pembayaran tidak bisa dihapus.'}, status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)

    def get_queryset(self):
        qs     = super().get_queryset()
        dari   = self.request.query_params.get('dari')
        sampai = self.request.query_params.get('sampai')
        if dari:   qs = qs.filter(tanggal__gte=dari)
        if sampai: qs = qs.filter(tanggal__lte=sampai)
        return qs

    @action(detail=False, methods=['get'], url_path='laporan-arus-kas')
    def laporan_arus_kas(self, request):
        tanggal_dari   = request.query_params.get('dari')
        tanggal_sampai = request.query_params.get('sampai')

        if not tanggal_dari or not tanggal_sampai:
            return Response({'error': 'Parameter dari dan sampai wajib diisi'}, status=status.HTTP_400_BAD_REQUEST)

        qs = Transaksi.objects.filter(tanggal__gte=tanggal_dari, tanggal__lte=tanggal_sampai)

        def get_total(queryset, jenis):
            return queryset.filter(jenis=jenis).aggregate(total=Coalesce(Sum('jumlah'), Decimal('0')))['total']

        def get_total_sub(queryset, sub_kategori):
            masuk  = queryset.filter(sub_kategori=sub_kategori, jenis='masuk').aggregate(total=Coalesce(Sum('jumlah'), Decimal('0')))['total']
            keluar = queryset.filter(sub_kategori=sub_kategori, jenis='keluar').aggregate(total=Coalesce(Sum('jumlah'), Decimal('0')))['total']
            return masuk, keluar

        qs_sebelum = Transaksi.objects.filter(tanggal__lt=tanggal_dari)
        kas_awal   = get_total(qs_sebelum, 'masuk') - get_total(qs_sebelum, 'keluar')

        qs_operasi = qs.filter(kategori_arus='operasi')
        tm_pelanggan_masuk, tm_pelanggan_keluar = get_total_sub(qs_operasi, 'tagihan_muka_pelanggan')
        kas_operasi_masuk, kas_operasi_keluar   = get_total_sub(qs_operasi, 'kas_masuk_operasi')
        tm_pemasok_masuk, tm_pemasok_keluar     = get_total_sub(qs_operasi, 'tagihan_muka_pemasok')
        kas_bayar_masuk, kas_bayar_keluar       = get_total_sub(qs_operasi, 'kas_keluar_operasi')
        total_operasi_masuk  = tm_pelanggan_masuk + kas_operasi_masuk + tm_pemasok_masuk + kas_bayar_masuk
        total_operasi_keluar = tm_pelanggan_keluar + kas_operasi_keluar + tm_pemasok_keluar + kas_bayar_keluar
        total_operasi        = total_operasi_masuk - total_operasi_keluar

        qs_investasi    = qs.filter(kategori_arus='investasi')
        inv_masuk       = get_total(qs_investasi, 'masuk')
        inv_keluar      = get_total(qs_investasi, 'keluar')
        total_investasi = inv_masuk - inv_keluar

        qs_keuangan    = qs.filter(kategori_arus='keuangan')
        keu_masuk      = get_total(qs_keuangan, 'masuk')
        keu_keluar     = get_total(qs_keuangan, 'keluar')
        total_keuangan = keu_masuk - keu_keluar

        qs_tidak       = qs.filter(kategori_arus='tidak_diklasifikasi')
        tidak_per_akun = []
        akun_ids       = qs_tidak.values_list('akun_id', flat=True).distinct()
        for akun_id in akun_ids:
            akun_obj = Akun.objects.get(id=akun_id)
            qs_akun  = qs_tidak.filter(akun_id=akun_id)
            a_masuk  = get_total(qs_akun, 'masuk')
            a_keluar = get_total(qs_akun, 'keluar')
            tidak_per_akun.append({'kode_akun': akun_obj.kode_akun, 'nama_akun': akun_obj.nama_akun, 'kas_masuk': a_masuk, 'kas_keluar': a_keluar, 'total': a_masuk - a_keluar})

        total_tidak_masuk  = get_total(qs_tidak, 'masuk')
        total_tidak_keluar = get_total(qs_tidak, 'keluar')
        total_tidak        = total_tidak_masuk - total_tidak_keluar

        kas_akhir_total    = kas_awal + total_operasi + total_investasi + total_keuangan + total_tidak
        akun_kas           = Akun.objects.filter(is_kas_setara=True)
        kas_akhir_per_akun = []
        for akun in akun_kas:
            qs_akun  = qs.filter(akun=akun)
            a_masuk  = get_total(qs_akun, 'masuk')
            a_keluar = get_total(qs_akun, 'keluar')
            kas_akhir_per_akun.append({'kode_akun': akun.kode_akun, 'nama_akun': akun.nama_akun, 'total': a_masuk - a_keluar})

        return Response({
            'periode': {'dari': tanggal_dari, 'sampai': tanggal_sampai},
            'kas_awal': kas_awal,
            'operasi': {
                'tagihan_muka_pelanggan': {'masuk': tm_pelanggan_masuk, 'keluar': tm_pelanggan_keluar},
                'kas_masuk_operasi': {'masuk': kas_operasi_masuk, 'keluar': kas_operasi_keluar},
                'tagihan_muka_pemasok': {'masuk': tm_pemasok_masuk, 'keluar': tm_pemasok_keluar},
                'kas_keluar_operasi': {'masuk': kas_bayar_masuk, 'keluar': kas_bayar_keluar},
                'total_masuk': total_operasi_masuk, 'total_keluar': total_operasi_keluar, 'total': total_operasi,
            },
            'investasi': {'kas_masuk': inv_masuk, 'kas_keluar': inv_keluar, 'total': total_investasi},
            'keuangan': {'kas_masuk': keu_masuk, 'kas_keluar': keu_keluar, 'total': total_keuangan},
            'tidak_diklasifikasi': {'per_akun': tidak_per_akun, 'total_masuk': total_tidak_masuk, 'total_keluar': total_tidak_keluar, 'total': total_tidak},
            'kas_akhir': {'per_akun': kas_akhir_per_akun, 'total': kas_akhir_total},
        })

def _get_available_alokasi_for_faktur(faktur):
    """
    Mengambil list AlokasiDana yang dapat digunakan untuk membayar faktur ini.
    Bisa dari Alokasi spesifik (id_pembiayaan) atau dari Alokasi Induk Pembiayaan.
    """
    id_pbiaya = str(faktur.id_pembiayaan or '').strip()
    mapping = None
    if id_pbiaya:
        mapping = PembiayaanIndukMapping.objects.filter(id_pembiayaan=id_pbiaya).select_related('induk').first()
    
    if mapping and mapping.induk_id:
        return list(
            AlokasiDana.objects
            .filter(Q(induk_pembiayaan=mapping.induk) | Q(id_pembiayaan=id_pbiaya), sisa_alokasi__gt=0)
            .order_by('tanggal_penerimaan', 'created_at', 'id')
        )
    return list(
        AlokasiDana.objects
        .filter(id_pembiayaan=id_pbiaya, sisa_alokasi__gt=0)
        .order_by('tanggal_penerimaan', 'created_at', 'id')
    )

class FakturViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset           = Faktur.objects.select_related('pelanggan', 'created_by').prefetch_related('items', 'pembayaran__akun').all()
    permission_classes = [IsKeuanganOrManajerPermission]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return FakturInputSerializer
        return FakturSerializer

    def perform_create(self, serializer):
        tanggal = serializer.validated_data.get('tanggal') or timezone.localdate()

        faktur = serializer.save(
            created_by=self.request.user,
            nomor_faktur=generate_nomor_faktur(tanggal)
        )
        total_real = (
            faktur.adm + faktur.jasa + faktur.farmasi + faktur.tindakan +
            faktur.fisio + faktur.lab + faktur.rad + faktur.kamar +
            faktur.bhp + faktur.lainnya + faktur.ambulan + faktur.alat
        )
        faktur.total_real_rs = total_real
        if faktur.is_cob and faktur.tanggungan_bpjs:
            faktur.total_tagihan = max(Decimal('0'), Decimal(str(total_real)) - Decimal(str(faktur.tanggungan_bpjs)))
        else:
            faktur.total_tagihan = total_real
        faktur.save()

    def perform_update(self, serializer):
        faktur = serializer.save()
        total_real = (
            faktur.adm + faktur.jasa + faktur.farmasi + faktur.tindakan +
            faktur.fisio + faktur.lab + faktur.rad + faktur.kamar +
            faktur.bhp + faktur.lainnya + faktur.ambulan + faktur.alat
        )
        faktur.total_real_rs = total_real
        if faktur.is_cob and faktur.tanggungan_bpjs:
            faktur.total_tagihan = max(Decimal('0'), Decimal(str(total_real)) - Decimal(str(faktur.tanggungan_bpjs)))
        else:
            faktur.total_tagihan = total_real
        faktur.save()
        
    def generate_nomor_faktur(self, tanggal):
        prefix = f"{tanggal:%y}"

        max_urut = 0

        last_faktur = (
            Faktur.objects
            .filter(nomor_faktur__startswith=prefix)
            .order_by('-nomor_faktur')
            .values_list('nomor_faktur', flat=True)
            .first()
        )

        if last_faktur and str(last_faktur)[2:6].isdigit():
            max_urut = int(str(last_faktur)[2:6])

        nomor = f"{prefix}{max_urut + 1:04d}"

        while Faktur.objects.filter(nomor_faktur=nomor).exists():
            max_urut += 1
            nomor = f"{prefix}{max_urut + 1:04d}"

        return nomor

    def destroy(self, request, *args, **kwargs):
        faktur = self.get_object()
        if faktur.pembayaran.exists():
            return Response({'error': 'Invoice yang sudah memiliki pembayaran atau pengajuan pembayaran tidak bisa dihapus.'}, status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)

    def get_queryset(self):
        qs        = super().get_queryset()
        pelanggan = self.request.query_params.get('pelanggan')
        search    = self.request.query_params.get('search')
        st        = self.request.query_params.get('status')
        id_pbiaya = self.request.query_params.get('id_pembiayaan')
        dari      = self.request.query_params.get('dari')
        sampai    = self.request.query_params.get('sampai')
        aging     = self.request.query_params.get('aging')
        if search:
            qs = qs.filter(
                Q(nomor_faktur__icontains=search) |
                Q(nama_pembiayaan__icontains=search) |
                Q(id_pembiayaan__icontains=search) |
                Q(pelanggan__nama__icontains=search)
            )
        if pelanggan: qs = qs.filter(pelanggan_id=pelanggan)
        if id_pbiaya: qs = qs.filter(id_pembiayaan=id_pbiaya)
        if st:        qs = qs.filter(status=st)
        if dari:      qs = qs.filter(tanggal__gte=dari)
        if sampai:    qs = qs.filter(tanggal__lte=sampai)
        if aging:
            today = timezone.localdate()
            qs = qs.exclude(status='batal').filter(total_tagihan__gt=F('total_dibayar'))
            if aging == 'not_due':
                qs = qs.filter(Q(jatuh_tempo__isnull=True) | Q(jatuh_tempo__gte=today))
            elif aging == '1_30':
                qs = qs.filter(jatuh_tempo__lt=today, jatuh_tempo__gte=today - timedelta(days=30))
            elif aging == '31_60':
                qs = qs.filter(jatuh_tempo__lt=today - timedelta(days=30), jatuh_tempo__gte=today - timedelta(days=60))
            elif aging == 'over_60':
                qs = qs.filter(jatuh_tempo__lt=today - timedelta(days=60))
        return qs.order_by('-nomor_faktur')

    @action(detail=True, methods=['post'], url_path='bayar')
    def bayar(self, request, pk=None):
        faktur = self.get_object()
        if faktur.status in ['lunas', 'batal']:
            return Response({'error': 'Faktur sudah lunas atau dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = PembayaranFakturInputSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        jumlah = serializer.validated_data['jumlah']
        if jumlah > faktur.sisa_tagihan:
            return Response({'error': f'Jumlah bayar melebihi sisa tagihan ({faktur.sisa_tagihan}).'}, status=status.HTTP_400_BAD_REQUEST)
        pending_total = faktur.pembayaran.filter(status_verifikasi='menunggu').aggregate(
            total=Sum('jumlah')
        )['total'] or Decimal('0')
        if pending_total + jumlah > faktur.sisa_tagihan:
            return Response({'error': f'Total pembayaran menunggu verifikasi melebihi sisa tagihan ({faktur.sisa_tagihan}).'}, status=status.HTTP_400_BAD_REQUEST)
        alokasi_list = _get_available_alokasi_for_faktur(faktur)
        saldo_wallet = sum((alokasi.sisa_alokasi for alokasi in alokasi_list), Decimal('0'))
        if pending_total + jumlah > saldo_wallet:
            return Response({'error': f'Jumlah bayar melebihi saldo pembiayaan ({saldo_wallet}).'}, status=status.HTTP_400_BAD_REQUEST)

        pembayaran = PembayaranFaktur.objects.create(
            faktur=faktur, created_by=request.user,
            tanggal=serializer.validated_data['tanggal'],
            jumlah=jumlah, metode=serializer.validated_data['metode'],
            keterangan=serializer.validated_data.get('keterangan', ''),
            akun=serializer.validated_data.get('akun'),
            status_verifikasi='menunggu',
        )
        faktur.refresh_from_db()
        return Response(PembayaranFakturSerializer(pembayaran).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path=r'pembayaran/(?P<pembayaran_id>[^/.]+)/verifikasi')
    def verifikasi_pembayaran(self, request, pk=None, pembayaran_id=None):
        if not is_manajer_keuangan(request.user):
            return Response({'error': 'Hanya manajer keuangan yang bisa verifikasi pembayaran.'}, status=status.HTTP_403_FORBIDDEN)
        faktur = self.get_object()
        try:
            pembayaran = faktur.pembayaran.get(pk=pembayaran_id)
        except PembayaranFaktur.DoesNotExist:
            return Response({'error': 'Pembayaran tidak ditemukan pada invoice ini.'}, status=status.HTTP_404_NOT_FOUND)
        if pembayaran.status_verifikasi == 'terverifikasi':
            return Response({'error': 'Pembayaran sudah terverifikasi.'}, status=status.HTTP_400_BAD_REQUEST)
        if pembayaran.status_verifikasi == 'ditolak':
            return Response({'error': 'Pembayaran sudah ditolak.'}, status=status.HTTP_400_BAD_REQUEST)
        if faktur.status in ['lunas', 'batal']:
            return Response({'error': 'Invoice sudah lunas atau dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)
        if pembayaran.jumlah > faktur.sisa_tagihan:
            return Response({'error': f'Jumlah pembayaran melebihi sisa tagihan saat ini ({faktur.sisa_tagihan}).'}, status=status.HTTP_400_BAD_REQUEST)

        alokasi_list = _get_available_alokasi_for_faktur(faktur)
        saldo_wallet = sum((alokasi.sisa_alokasi for alokasi in alokasi_list), Decimal('0'))
        if pembayaran.jumlah > saldo_wallet:
            return Response({'error': f'Jumlah pembayaran melebihi saldo pembiayaan ({saldo_wallet}).'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            sisa_potong = pembayaran.jumlah
            for alokasi in alokasi_list:
                if sisa_potong <= 0:
                    break
                nominal_potong = min(alokasi.sisa_alokasi, sisa_potong)
                AlokasiDanaPemakaian.objects.create(
                    alokasi_dana=alokasi,
                    pembayaran=pembayaran,
                    jumlah=nominal_potong,
                )
                alokasi.save()
                sisa_potong -= nominal_potong
            pembayaran.status_verifikasi = 'terverifikasi'
            pembayaran.verified_by = request.user
            pembayaran.verified_at = timezone.now()
            pembayaran.save()
        faktur.refresh_from_db()
        return Response(FakturSerializer(faktur).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path=r'pembayaran/(?P<pembayaran_id>[^/.]+)/batal-verifikasi')
    def batal_verifikasi_pembayaran(self, request, pk=None, pembayaran_id=None):
        if not is_manajer_keuangan(request.user):
            return Response({'error': 'Hanya manajer keuangan ke atas yang bisa membatalkan verifikasi pembayaran.'}, status=status.HTTP_403_FORBIDDEN)
        faktur = self.get_object()
        try:
            pembayaran = faktur.pembayaran.get(pk=pembayaran_id)
        except PembayaranFaktur.DoesNotExist:
            return Response({'error': 'Pembayaran tidak ditemukan pada invoice ini.'}, status=status.HTTP_404_NOT_FOUND)
        if pembayaran.status_verifikasi != 'terverifikasi':
            return Response({'error': 'Hanya pembayaran berstatus terverifikasi yang dapat dibatalkan verifikasinya.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            pemakaian_qs = AlokasiDanaPemakaian.objects.filter(pembayaran=pembayaran)
            affected_alokasi = list({p.alokasi_dana for p in pemakaian_qs})
            pemakaian_qs.delete()
            for alokasi in affected_alokasi:
                alokasi.save()

            pembayaran.status_verifikasi = 'menunggu'
            pembayaran.verified_by = None
            pembayaran.verified_at = None
            pembayaran.save()

        faktur.refresh_from_db()
        return Response({
            'message': f'Verifikasi pembayaran invoice {faktur.nomor_faktur} berhasil dibatalkan.',
            'faktur': FakturSerializer(faktur).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['delete'], url_path=r'pembayaran/(?P<pembayaran_id>[^/.]+)')
    def hapus_pembayaran(self, request, pk=None, pembayaran_id=None):
        faktur = self.get_object()
        try:
            pembayaran = faktur.pembayaran.get(pk=pembayaran_id)
        except PembayaranFaktur.DoesNotExist:
            return Response({'error': 'Pembayaran tidak ditemukan pada invoice ini.'}, status=status.HTTP_404_NOT_FOUND)
        pembayaran.delete()
        faktur.refresh_from_db()
        return Response(FakturSerializer(faktur).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path=r'pembayaran/(?P<pembayaran_id>[^/.]+)/batal')
    def batal_pembayaran(self, request, pk=None, pembayaran_id=None):
        faktur = self.get_object()
        try:
            pembayaran = faktur.pembayaran.get(pk=pembayaran_id)
        except PembayaranFaktur.DoesNotExist:
            return Response({'error': 'Pembayaran tidak ditemukan pada invoice ini.'}, status=status.HTTP_404_NOT_FOUND)
        if pembayaran.status_verifikasi == 'dibatalkan':
            return Response({'error': 'Pembayaran sudah dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            alokasi_terpakai = list(
                AlokasiDana.objects.filter(pemakaian_alokasi__pembayaran=pembayaran).distinct()
            )
            pembayaran.pemakaian_alokasi.all().delete()
            pembayaran.status_verifikasi = 'dibatalkan'
            pembayaran.verified_by = request.user
            pembayaran.verified_at = timezone.now()
            pembayaran.save()
            for alokasi in alokasi_terpakai:
                alokasi.save()
        faktur.refresh_from_db()
        return Response(FakturSerializer(faktur).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='kirim')
    def kirim(self, request, pk=None):
        faktur = self.get_object()
        tanggal_kirim = request.data.get('tgl_kirim') or request.data.get('tanggal_kirim')
        if not tanggal_kirim:
            return Response({'tgl_kirim': 'Tanggal kirim invoice wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            tanggal_kirim = datetime.strptime(str(tanggal_kirim), '%Y-%m-%d').date()
        except ValueError:
            return Response({'tgl_kirim': 'Format tanggal kirim harus YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
        faktur.tgl_kirim = tanggal_kirim
        faktur.jatuh_tempo = tanggal_kirim + timedelta(days=45)
        faktur.save()
        return Response(FakturSerializer(faktur).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='batal')
    def batal(self, request, pk=None):
        faktur = self.get_object()
        if faktur.tgl_kirim:
            return Response({'error': 'Invoice yang sudah dikirim tidak bisa dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)
        if faktur.status == 'lunas':
            return Response({'error': 'Faktur yang sudah lunas tidak bisa dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)
        if faktur.pembayaran.exists():
            return Response({'error': 'Invoice yang sudah memiliki pembayaran atau pengajuan pembayaran tidak bisa dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            faktur.status = 'batal'
            faktur.save()
            with connection.cursor() as cursor:
                cursor.execute(
                    "UPDATE rssams.verif_kunjung SET no_invoice='' WHERE no_invoice=%s",
                    [faktur.nomor_faktur],
                )
        return Response(FakturSerializer(faktur).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='kembalikan')
    def kembalikan(self, request, pk=None):
        """
        Batalkan invoice yang sudah dikirim ke asuransi karena dikembalikan/ditolak.
        Berbeda dari /batal/ biasa — endpoint ini khusus untuk invoice yang tgl_kirim sudah terisi.
        Wajib menyertakan alasan_batal di request body.
        Kunjungan yang terikat akan otomatis dilepas sehingga bisa di-assign ke invoice baru.
        """
        faktur = self.get_object()

        if faktur.status == 'batal':
            return Response({'error': 'Invoice ini sudah dibatalkan sebelumnya.'}, status=status.HTTP_400_BAD_REQUEST)
        if faktur.status == 'lunas':
            return Response({'error': 'Invoice yang sudah lunas tidak bisa dikembalikan.'}, status=status.HTTP_400_BAD_REQUEST)
        if not faktur.tgl_kirim:
            return Response(
                {'error': 'Invoice belum dikirim. Gunakan tombol Batalkan biasa untuk membatalkan invoice ini.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        alasan = str(request.data.get('alasan_batal') or '').strip()
        if not alasan:
            return Response({'alasan_batal': 'Alasan pengembalian wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            faktur.status = 'batal'
            faktur.alasan_batal = alasan
            faktur.dibatalkan_oleh = request.user
            faktur.dibatalkan_at = timezone.now()
            faktur.save()
            with connection.cursor() as cursor:
                cursor.execute(
                    "UPDATE rssams.verif_kunjung SET no_invoice='' WHERE no_invoice=%s",
                    [faktur.nomor_faktur],
                )

        return Response(FakturSerializer(faktur, context={'view': self}).data, status=status.HTTP_200_OK)

_ROMAN_MONTHS = {
    1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V', 6: 'VI',
    7: 'VII', 8: 'VIII', 9: 'IX', 10: 'X', 11: 'XI', 12: 'XII',
}

_ID_MONTHS = {
    1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
    5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
    9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember',
}

def _invoice_money(value, decimals=2):
    amount = Decimal(value or 0)
    return f"{amount:,.{decimals}f}"

def _invoice_date(value):
    if not value:
        return ''
    return f"{value.day:02d} {_ID_MONTHS[value.month]} {value.year}"

def _legacy_invoice_number(faktur):
    tanggal = faktur.tanggal or timezone.localdate()
    return f"{faktur.nomor_faktur}/Keu-02/RS-SAMS/{_ROMAN_MONTHS[tanggal.month]}/{tanggal.year}"

def _invoice_total_tagihan(faktur):
    if faktur.is_cob and Decimal(faktur.tanggungan_bpjs or 0) > 0:
        total_real = Decimal(faktur.total_real_rs or 0) or Decimal(faktur.total_tagihan or 0)
        return max(Decimal('0'), total_real - Decimal(str(faktur.tanggungan_bpjs)))
    return Decimal(faktur.total_tagihan or 0)

def _invoice_print_amounts(faktur):
    rows = get_invoice_kunjungan_rows(faktur)
    if rows:
        jumlah_bayar = sum((Decimal(row.get('jmlbyr') or 0) for row in rows), Decimal('0'))
        total_real = sum((Decimal(row.get('ttl') or 0) for row in rows), Decimal('0'))
        if faktur.is_cob and Decimal(faktur.tanggungan_bpjs or 0) > 0:
            total = max(Decimal('0'), total_real - Decimal(str(faktur.tanggungan_bpjs)))
        else:
            total = total_real
        return total, jumlah_bayar
    return _invoice_total_tagihan(faktur), Decimal('0')

def _legacy_words(number):
    units = ['', 'satu', 'dua', 'tiga', 'empat', 'lima', 'enam', 'tujuh', 'delapan', 'sembilan', 'sepuluh', 'sebelas']
    number = int(Decimal(number or 0).quantize(Decimal('1')))
    if number < 12:
        return units[number]
    if number < 20:
        return f"{_legacy_words(number - 10)} belas"
    if number < 100:
        return f"{_legacy_words(number // 10)} puluh {_legacy_words(number % 10)}".strip()
    if number < 200:
        return f"seratus {_legacy_words(number - 100)}".strip()
    if number < 1000:
        return f"{_legacy_words(number // 100)} ratus {_legacy_words(number % 100)}".strip()
    if number < 2000:
        return f"seribu {_legacy_words(number - 1000)}".strip()
    if number < 1000000:
        return f"{_legacy_words(number // 1000)} ribu {_legacy_words(number % 1000)}".strip()
    if number < 1000000000:
        return f"{_legacy_words(number // 1000000)} juta {_legacy_words(number % 1000000)}".strip()
    return f"{_legacy_words(number // 1000000000)} milyar {_legacy_words(number % 1000000000)}".strip()

def _invoice_print_rows(faktur, ppn=False):
    farmasi = Decimal(faktur.farmasi or 0)
    ppn_obat = Decimal(faktur.ppn_farmasi or 0)
    rows = [('-  BIAYA JASA', faktur.jasa), ('-  BIAYA TINDAKAN', faktur.tindakan)]
    if ppn:
        rows.extend([
            ('-  O b a t', max(farmasi - ppn_obat, Decimal('0'))),
            ('-  PPN Obat', ppn_obat),
            ('-  Embalasi', Decimal('0')),
        ])
    else:
        rows.append(('-  BIAYA FARMASI', max(farmasi - ppn_obat, Decimal('0'))))
    rows.extend([
        ('-  BIAYA LABORATORIUM', faktur.lab),
        ('-  BIAYA RADIOLOGI', faktur.rad),
        ('-  BIAYA KAMAR', faktur.kamar),
        ('-  BIAYA BAHAN HABIS PAKAI (BHP)', faktur.bhp),
        ('-  BIAYA ALAT', faktur.alat),
        ('-  BIAYA AMBULAN', faktur.ambulan),
        ('-  BIAYA FISIOTERAPI', faktur.fisio),
        ('-  BIAYA ADMINISTRASI DAN LAINNYA', Decimal(faktur.lainnya or 0) + Decimal(faktur.adm or 0)),
    ])
    if not ppn and ppn_obat > 0:
        rows.append(('-  PPN OBAT', ppn_obat))
    return rows

def _invoice_pembiayaan_name(faktur):
    # Debug: log what we're getting
    print(f"DEBUG: pelanggan={faktur.pelanggan}, nama_pembiayaan={faktur.nama_pembiayaan}")
    if faktur.pelanggan:
        print(f"DEBUG: returning pelanggan.nama={faktur.pelanggan.nama}")
        return faktur.pelanggan.nama
    result = faktur.nama_pembiayaan or '-'
    print(f"DEBUG: returning nama_pembiayaan={result}")
    return result

def get_invoice_logo_path():
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo-1.jpg')
    return logo_path if os.path.exists(logo_path) else None

def get_invoice_farmasi_ppn_totals(faktur):
    from django.db import connection

    totals = {
        'obat': Decimal('0'),
        'ppn_obat': Decimal('0'),
        'embalasi': Decimal('0'),
    }

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT no
                FROM rssams.verif_kunjung
                WHERE no_invoice = %s
                """,
                [faktur.nomor_faktur]
            )
            kunjungan_rows = cursor.fetchall()

            for (no_kunj,) in kunjungan_rows:
                cursor.execute(
                    """
                    SELECT
                        COALESCE(SUM((modal * persen) * (qty - retur)), 0) AS ttlobat,
                        COALESCE(SUM(ppn * (qty - retur)), 0) AS ttlppnobat,
                        COALESCE(SUM(embalace), 0) AS ttlemba
                    FROM rssams.item_tran_apt
                    WHERE no_kunj = %s
                    """,
                    [no_kunj]
                )
                row = cursor.fetchone()
                if row:
                    totals['obat'] += Decimal(row[0] or 0)
                    totals['ppn_obat'] += Decimal(row[1] or 0)
                    totals['embalasi'] += Decimal(row[2] or 0)

    except Exception:
        pass

    return totals

def get_invoice_kunjungan_rows(faktur):
    from django.db import connection

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    a.no,
                    a.noreg,
                    b.nama,
                    a.adm,
                    a.jasa,
                    a.farmasi,
                    a.tindakan,
                    a.fisio,
                    a.lab,
                    a.lab_pa,
                    a.kamar,
                    a.rad,
                    a.bhp,
                    a.lainnya,
                    a.ambulan,
                    a.alat,
                    a.jmlbyr,
                    (
                        COALESCE(a.adm,0) +
                        COALESCE(a.jasa,0) +
                        COALESCE(a.farmasi,0) +
                        COALESCE(a.tindakan,0) +
                        COALESCE(a.fisio,0) +
                        COALESCE(a.lab,0) +
                        COALESCE(a.kamar,0) +
                        COALESCE(a.rad,0) +
                        COALESCE(a.bhp,0) +
                        COALESCE(a.lainnya,0) +
                        COALESCE(a.ambulan,0) +
                        COALESCE(a.alat,0) -
                        COALESCE(a.jmlbyr,0)
                    ) AS ttl,
                    DATE(a.tgl_masuk) AS tgl_masuk
                FROM rssams.kunjung a
                INNER JOIN rssams.regpasien b ON a.noreg = b.noreg
                INNER JOIN rssams.verif_kunjung c ON a.no = c.no
                WHERE c.no_invoice = %s
                ORDER BY a.no
                """,
                [faktur.nomor_faktur]
            )
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]
    except Exception:
        return []

def get_kunjungan_farmasi_ppn_detail(no_kunj):
    from django.db import connection

    totals = {
        'obat': Decimal('0'),
        'ppn_obat': Decimal('0'),
        'embalasi': Decimal('0'),
    }

    if not no_kunj:
        return totals

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    COALESCE(ROUND(SUM((modal * persen) * (qty - retur)), 2), 0) AS ttlobat,
                    COALESCE(ROUND(SUM(ppn * (qty - retur)), 2), 0) AS ttlppnobat,
                    COALESCE(SUM(embalace), 0) AS ttlemba
                FROM rssams.item_tran_apt
                WHERE no_kunj = %s
                """,
                [no_kunj]
            )
            row = cursor.fetchone()

        if row:
            totals['obat'] = Decimal(row[0] or 0)
            totals['ppn_obat'] = Decimal(row[1] or 0)
            totals['embalasi'] = Decimal(row[2] or 0)

    except Exception:
        pass

    return totals

class InvoicePDF(FPDF):
    def header(self):
        logo_path = get_invoice_logo_path()
        if logo_path:
            self.image(logo_path, x=172, y=5, h=21)
        self.ln(20)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

def render_invoice_pdf_response(faktur, mode='invoice'):
    pdf = InvoicePDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    print("WIDTH =", pdf.w)
    print("HEIGHT =", pdf.h)
    
    pembiayaan_detail = get_pembiayaan_detail(faktur.id_pembiayaan)
    nama_pbiaya = pembiayaan_detail.get('pembiayaan') or _invoice_pembiayaan_name(faktur) or 'PEMBIAYAAN'
    alamat = pembiayaan_detail.get('alamat') or ''

    tanggal = faktur.tanggal or timezone.localdate()
    tgl = f"{tanggal.day:02d}-{tanggal.month:02d}-{tanggal.year}"
    no_invoice = _legacy_invoice_number(faktur)

    ttl_tagihan, jml_bayar = _invoice_print_amounts(faktur)

    if faktur.xround == 'Y':
        angka = ttl_tagihan.to_integral_value(rounding='ROUND_CEILING')
    else:
        angka = ttl_tagihan

    terbilang = _legacy_words(angka).title()
    terbilang = terbilang.replace('Koma Nol Nol', '').replace('Koma', '').strip()

    pdf.set_font('Times', '', 16)
    pdf.set_x(20)
    pdf.cell(180, 10, "INVOICE", 0, 0, 'C')
    pdf.ln(20)

    pdf.set_font('Times', '', 10)
    pdf.text(20, 48, "Kepada Yth.")
    pdf.text(20, 53, "Bagian Keuangan")
    pdf.text(120, 48, "Nomor")
    pdf.text(135, 48, f": {no_invoice}")
    pdf.text(120, 53, "Tanggal")
    pdf.text(135, 53, f": {tgl}")
    pdf.text(20, 58, str(nama_pbiaya))
    if alamat:
        pdf.text(20, 63, str(alamat))

    pdf.line(20, 70, 200, 70)
    pdf.line(20, 71, 200, 71)

    pdf.text(20, 80, str(faktur.jenis or ''))
    pdf.text(20, 85, f"BEBAN {faktur.beban or ''}")
    pdf.text(20, 90, f"PERIODE : {faktur.periode or ''}")

    def row(y, label, value, x_label=20, is_negative=False):
        pdf.set_font('Times', '', 10)
        pdf.set_xy(x_label, y)
        pdf.cell(90, 5, label, 0, 0, 'L')
        pdf.set_xy(147, y)
        pdf.cell(20, 5, "Rp.", 0, 0, 'R')
        pdf.set_xy(170, y)
        val_str = f"{Decimal(value or 0):,.2f}"
        if is_negative:
            val_str = f"({val_str})"
        pdf.cell(30, 5, val_str, 0, 0, 'R')

    y = 96
    row(y, "-  BIAYA JASA", faktur.jasa)
    y += 5
    row(y, "-  BIAYA TINDAKAN", faktur.tindakan)
    y += 5

    if mode == 'invoice_ppn':
        farmasi_ppn = get_invoice_farmasi_ppn_totals(faktur)

        pdf.set_xy(20, y)
        pdf.cell(90, 5, "-  BIAYA FARMASI", 0, 0, 'L')

        y += 5
        row(y, "- O b a t", farmasi_ppn['obat'], x_label=23)

        y += 5
        row(y, "- PPN Obat", farmasi_ppn['ppn_obat'], x_label=23)

        y += 5
        row(y, "- Embalasi", farmasi_ppn['embalasi'], x_label=23)
    else:
        row(y, "-  BIAYA FARMASI", Decimal(faktur.farmasi or 0) - Decimal(faktur.ppn_farmasi or 0))

    y += 5
    row(y, "-  BIAYA LABORATORIUM", faktur.lab)
    y += 5
    row(y, "-  BIAYA RADIOLOGI", faktur.rad)
    y += 5
    row(y, "-  BIAYA KAMAR", faktur.kamar)
    y += 5
    row(y, "-  BIAYA BAHAN HABIS PAKAI (BHP)", faktur.bhp)
    y += 5
    row(y, "-  BIAYA ALAT", faktur.alat)
    y += 5
    row(y, "-  BIAYA AMBULAN", faktur.ambulan)
    y += 5
    row(y, "-  BIAYA FISIOTERAPI", faktur.fisio)
    y += 5
    row(y, "-  BIAYA ADMINISTRASI DAN LAINNYA", Decimal(faktur.lainnya or 0) + Decimal(faktur.adm or 0))

    if Decimal(faktur.ppn_farmasi or 0) > 0:
        y += 5
        row(y, "-  PPN OBAT", faktur.ppn_farmasi)

    if faktur.is_cob and Decimal(faktur.tanggungan_bpjs or 0) > 0:
        y += 5
        row(y, "-  DITANGGUNG BPJS (INA-CBGs)", faktur.tanggungan_bpjs, is_negative=True)

    y += 5
    row(y, "-  JUMLAH YANG SUDAH DIBAYAR", jml_bayar, is_negative=True)

    y += 5
    pdf.set_font('Times', 'B', 12)
    pdf.set_xy(147, y)
    pdf.cell(20, 5, "Total", 0, 0, 'R')
    pdf.set_xy(170, y)
    pdf.cell(30, 5, f"{ttl_tagihan:,.2f}", 0, 0, 'R')
    pdf.line(142, y, 200, y)

    if faktur.xround == 'Y':
        y += 5
        pembulatan = ttl_tagihan.to_integral_value(rounding='ROUND_CEILING')
        pdf.set_font('Times', 'B', 12)
        pdf.set_xy(147, y)
        pdf.cell(20, 5, "Pembulatan", 0, 0, 'R')
        pdf.set_xy(170, y)
        pdf.cell(30, 5, f"{pembulatan:,.2f}", 0, 0, 'R')
        pdf.line(142, y, 200, y)
        pdf.line(142, y + 5, 200, y + 5)

    if mode == 'invoice_ppn':
        note_y = 185
        words_y = 190
    else:
        note_y = 175
        words_y = 180

    pdf.set_font('Times', 'I', 11)
    pdf.text(20, note_y, "#REKAPITULASI & BACKUP TERLAMPIR")
    pdf.set_xy(19, words_y)
    pdf.multi_cell(180, 5, f"#{terbilang} Rupiah#", 0, 'J')

    pdf.line(20, 200, 200, 200)
    pdf.line(20, 201, 200, 201)

    pdf.set_font('Times', '', 10)
    pdf.text(20, 209, "Jangka waktu pelunasan maksimal 14 hari setelah")
    pdf.text(20, 213, "diterimanya tagihan ini, ke rekening Kami")
    pdf.set_font('Times', 'B', 12)
    pdf.text(20, 219, "No. 777.998.9978")
    pdf.set_font('Times', '', 10)
    pdf.text(20, 223, "Bank Syariah Indonesia (BSI)")
    pdf.set_font('Times', 'B', 10)
    pdf.text(20, 227, "a.n. RS Siaga Al Munawwarah")
    pdf.set_font('Times', '', 10)
    pdf.text(20, 250, "cc.")
    pdf.text(25, 250, "- Akuntansi RS-SAMS")
    pdf.text(25, 254, "- Arsip")
    pdf.text(148, 209, "RS. SIAGA AL MUNAWWARAH")
    pdf.text(164, 213, "SAMARINDA")
    pdf.text(150, 246, "Nevi Nevada, S.Ip., M.M.")

    pdf_bytes = bytes(pdf.output(dest='S'))

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Invoice_{faktur.nomor_faktur}.pdf"'
    return response

def render_rincian_pdf_response(faktur, mode='rincian'):
    pdf = FPDF('P', 'mm', (330, 216))
    pdf.set_auto_page_break(False)
    pdf.add_page()
    print("WIDTH =", pdf.w)
    print("HEIGHT =", pdf.h)

    tanggal = faktur.tanggal or timezone.localdate()
    tgl = f"{tanggal.day:02d}-{tanggal.month:02d}-{tanggal.year}"
    no_invoice = _legacy_invoice_number(faktur)

    rows = get_invoice_kunjungan_rows(faktur)

    pdf.set_font('Times', '', 11)
    pdf.set_x(10)
    pdf.cell(57, 5, 'LAMPIRAN', 0, 1, 'L')

    pdf.set_x(10)
    pdf.cell(195, 5, 'Nomor Invoice', 0, 0, 'L')
    pdf.set_x(35)
    pdf.cell(57, 5, f': {no_invoice}', 0, 1, 'L')

    pdf.set_x(10)
    pdf.cell(195, 5, 'Tanggal', 0, 0, 'L')
    pdf.set_x(35)
    pdf.cell(57, 5, f': {tgl}', 0, 1, 'L')

    pdf.set_font('Times', '', 9)
    pdf.set_x(10)
    pdf.cell(310, 5, '-' * 300, 0, 1, 'L')

    pdf.set_x(10)
    pdf.cell(10, 5, 'NO.', 0, 0, 'L')
    pdf.set_x(20)
    pdf.cell(10, 5, 'NAMA PASIEN', 0, 0, 'L')
    pdf.set_x(75)
    pdf.cell(20, 5, 'TANGGAL', 0, 0, 'R')
    pdf.set_x(95)
    pdf.cell(20, 5, 'JASA', 0, 0, 'R')
    pdf.set_x(115)
    pdf.cell(20, 5, 'TINDAKAN', 0, 0, 'R')
    pdf.set_x(135)
    pdf.cell(20, 5, 'FARMASI', 0, 0, 'R')
    pdf.set_x(158)
    pdf.cell(15, 5, 'LAB', 0, 0, 'R')
    pdf.set_x(175)
    pdf.cell(15, 5, 'RAD', 0, 0, 'R')
    pdf.set_x(187)
    pdf.cell(20, 5, 'KAMAR', 0, 0, 'R')
    pdf.set_x(205)
    pdf.cell(20, 5, 'BHP', 0, 0, 'R')
    pdf.set_x(222)
    pdf.cell(20, 5, 'ALAT', 0, 0, 'R')
    pdf.set_x(243)
    pdf.cell(15, 5, 'AMB', 0, 0, 'R')
    pdf.set_x(255)
    pdf.cell(20, 5, 'ADM &', 0, 0, 'R')
    pdf.set_x(277)
    pdf.cell(20, 5, 'JML BAYAR', 0, 0, 'R')
    pdf.set_x(300)
    pdf.cell(20, 5, 'TOTAL', 0, 0, 'R')
    pdf.ln(4)
    pdf.set_x(255)
    pdf.cell(20, 5, 'LAINNYA', 0, 0, 'R')
    pdf.ln(4)

    pdf.set_x(10)
    pdf.cell(310, 5, '-' * 300, 0, 1, 'L')

    totals = {
        'jasa': Decimal('0'),
        'tindakan': Decimal('0'),
        'farmasi': Decimal('0'),
        'lab': Decimal('0'),
        'rad': Decimal('0'),
        'kamar': Decimal('0'),
        'bhp': Decimal('0'),
        'alat': Decimal('0'),
        'ambulan': Decimal('0'),
        'fisio': Decimal('0'),
        'lainnya': Decimal('0'),
        'jmlbyr': Decimal('0'),
        'ttl': Decimal('0'),
    }

    def d(row, key):
        return Decimal(row.get(key) or 0)

    def fmt(value, dec=0):
        return f"{Decimal(value or 0):,.{dec}f}"

    for idx, row in enumerate(rows, start=1):
        nama = str(row.get('nama') or '')
        noreg = str(row.get('noreg') or '')
        if len(nama) > 23:
            nama_tampil = f"{nama[:23]}... ({noreg})"
        else:
            nama_tampil = f"{nama} ({noreg})"

        tgl_masuk = row.get('tgl_masuk')
        if hasattr(tgl_masuk, 'strftime'):
            tgl_masuk_text = tgl_masuk.strftime('%d-%m-%Y')
        else:
            tgl_masuk_text = str(tgl_masuk or '')

        lainnya_adm = d(row, 'lainnya') + d(row, 'adm')

        pdf.set_x(10)
        pdf.cell(10, 5, f'{idx}.', 0, 0, 'L')
        pdf.set_x(20)
        pdf.cell(10, 5, nama_tampil, 0, 0, 'L')
        pdf.set_x(75)
        pdf.cell(20, 5, tgl_masuk_text, 0, 0, 'R')
        pdf.set_x(95)
        pdf.cell(20, 5, fmt(d(row, 'jasa'), 0), 0, 0, 'R')
        pdf.set_x(115)
        pdf.cell(20, 5, fmt(d(row, 'tindakan'), 0), 0, 0, 'R')
        pdf.set_x(135)
        pdf.cell(20, 5, fmt(d(row, 'farmasi'), 2), 0, 0, 'R')
        pdf.set_x(158)
        pdf.cell(15, 5, fmt(d(row, 'lab'), 0), 0, 0, 'R')
        pdf.set_x(175)
        pdf.cell(15, 5, fmt(d(row, 'rad'), 0), 0, 0, 'R')
        pdf.set_x(187)
        pdf.cell(20, 5, fmt(d(row, 'kamar'), 0), 0, 0, 'R')
        pdf.set_x(205)
        pdf.cell(20, 5, fmt(d(row, 'bhp'), 0), 0, 0, 'R')
        pdf.set_x(222)
        pdf.cell(20, 5, fmt(d(row, 'alat'), 0), 0, 0, 'R')
        pdf.set_x(238)
        pdf.cell(20, 5, fmt(d(row, 'ambulan'), 0), 0, 0, 'R')
        pdf.set_x(255)
        pdf.cell(20, 5, fmt(lainnya_adm, 0), 0, 0, 'R')
        pdf.set_x(277)
        pdf.cell(20, 5, fmt(d(row, 'jmlbyr'), 2), 0, 0, 'R')
        pdf.set_x(300)
        pdf.cell(20, 5, fmt(d(row, 'ttl'), 2), 0, 0, 'R')
        pdf.ln(5)

        totals['jasa'] += d(row, 'jasa')
        totals['tindakan'] += d(row, 'tindakan')
        totals['farmasi'] += d(row, 'farmasi')
        totals['lab'] += d(row, 'lab')
        totals['rad'] += d(row, 'rad')
        totals['kamar'] += d(row, 'kamar')
        totals['bhp'] += d(row, 'bhp')
        totals['alat'] += d(row, 'alat')
        totals['ambulan'] += d(row, 'ambulan')
        totals['fisio'] += d(row, 'fisio')
        totals['lainnya'] += lainnya_adm
        totals['jmlbyr'] += d(row, 'jmlbyr')
        totals['ttl'] += d(row, 'ttl')

    pdf.set_x(10)
    pdf.cell(310, 5, '-' * 300, 0, 1, 'L')

    pdf.set_x(75)
    pdf.cell(20, 5, '', 0, 0, 'R')
    pdf.set_x(95)
    pdf.cell(20, 5, fmt(totals['jasa'], 0), 0, 0, 'R')
    pdf.set_x(115)
    pdf.cell(20, 5, fmt(totals['tindakan'], 0), 0, 0, 'R')
    pdf.set_x(135)
    pdf.cell(20, 5, fmt(totals['farmasi'], 2), 0, 0, 'R')
    pdf.set_x(158)
    pdf.cell(15, 5, fmt(totals['lab'], 0), 0, 0, 'R')
    pdf.set_x(175)
    pdf.cell(15, 5, fmt(totals['rad'], 0), 0, 0, 'R')
    pdf.set_x(187)
    pdf.cell(20, 5, fmt(totals['kamar'], 0), 0, 0, 'R')
    pdf.set_x(205)
    pdf.cell(20, 5, fmt(totals['bhp'], 0), 0, 0, 'R')
    pdf.set_x(222)
    pdf.cell(20, 5, fmt(totals['alat'], 0), 0, 0, 'R')
    pdf.set_x(238)
    pdf.cell(20, 5, fmt(totals['ambulan'], 0), 0, 0, 'R')
    pdf.set_x(255)
    pdf.cell(20, 5, fmt(totals['lainnya'], 0), 0, 0, 'R')
    pdf.set_x(277)
    pdf.cell(20, 5, fmt(totals['jmlbyr'], 2), 0, 0, 'R')
    pdf.set_x(300)
    pdf.cell(20, 5, fmt(totals['ttl'], 2), 0, 0, 'R')

    pdf.ln(4)
    pdf.set_x(10)
    pdf.cell(310, 5, '-' * 300, 0, 1, 'L')

    if faktur.is_cob and Decimal(faktur.tanggungan_bpjs or 0) > 0:
        total_real = totals['ttl']
        tanggungan = Decimal(faktur.tanggungan_bpjs)
        net_tagihan = max(Decimal('0'), total_real - tanggungan)

        pdf.set_font('Times', 'B', 9)
        pdf.set_x(230)
        pdf.cell(50, 5, 'TOTAL BIAYA RIIL RS:', 0, 0, 'R')
        pdf.set_x(280)
        pdf.cell(40, 5, f"Rp. {fmt(total_real, 2)}", 0, 1, 'R')

        pdf.set_x(230)
        pdf.cell(50, 5, 'DITANGGUNG BPJS (INA-CBGs):', 0, 0, 'R')
        pdf.set_x(280)
        pdf.cell(40, 5, f"- Rp. {fmt(tanggungan, 2)}", 0, 1, 'R')

        pdf.set_x(230)
        pdf.cell(50, 5, 'NET TAGIHAN ASURANSI:', 0, 0, 'R')
        pdf.set_x(280)
        pdf.cell(40, 5, f"Rp. {fmt(net_tagihan, 2)}", 0, 1, 'R')

        pdf.set_x(10)
        pdf.cell(310, 5, '-' * 300, 0, 1, 'L')

    pdf_bytes = bytes(pdf.output(dest='S'))

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Rincian_Invoice_{faktur.nomor_faktur}.pdf"'
    return response

def render_rincian_ppn_pdf_response(faktur, mode='rincian_ppn'):
    pdf = FPDF('P', 'mm', (330, 216))
    pdf.set_auto_page_break(False)
    pdf.add_page()

    tanggal = faktur.tanggal or timezone.localdate()
    tgl = f"{tanggal.day:02d}-{tanggal.month:02d}-{tanggal.year}"
    no_invoice = _legacy_invoice_number(faktur)

    rows = get_invoice_kunjungan_rows(faktur)

    pdf.set_font('Times', '', 11)
    pdf.set_x(10)
    pdf.cell(57, 5, 'LAMPIRAN', 0, 1, 'L')

    pdf.set_x(10)
    pdf.cell(195, 5, 'Nomor Invoice', 0, 0, 'L')
    pdf.set_x(35)
    pdf.cell(57, 5, f': {no_invoice}', 0, 1, 'L')

    pdf.set_x(10)
    pdf.cell(195, 5, 'Tanggal', 0, 0, 'L')
    pdf.set_x(35)
    pdf.cell(57, 5, f': {tgl}', 0, 1, 'L')

    pdf.set_font('Times', '', 9)

    pdf.set_x(10)
    pdf.cell(300, 5, '-' * 292, 0, 1, 'L')

    pdf.set_x(10)
    pdf.cell(10, 5, 'NO.', 0, 0, 'L')
    pdf.set_x(20)
    pdf.cell(10, 5, 'NAMA PASIEN', 0, 0, 'L')
    pdf.set_x(75)
    pdf.cell(25, 5, 'TANGGAL', 0, 0, 'R')
    pdf.set_x(95)
    pdf.cell(20, 5, 'JASA', 0, 0, 'R')
    pdf.set_x(115)
    pdf.cell(20, 5, 'TINDAKAN', 0, 0, 'R')
    pdf.set_x(135)
    pdf.cell(20, 5, 'OBAT', 0, 0, 'R')
    pdf.set_x(163)
    pdf.cell(9, 5, 'PPN', 0, 0, 'R')
    pdf.set_x(168)
    pdf.cell(15, 5, 'EMBA', 0, 0, 'R')
    pdf.set_x(185)
    pdf.cell(15, 5, 'LAB', 0, 0, 'R')
    pdf.set_x(197)
    pdf.cell(20, 5, 'RAD', 0, 0, 'R')
    pdf.set_x(215)
    pdf.cell(20, 5, 'BHP', 0, 0, 'R')
    pdf.set_x(230)
    pdf.cell(20, 5, 'ALAT', 0, 0, 'R')
    pdf.set_x(249)
    pdf.cell(15, 5, 'AMB', 0, 0, 'R')
    pdf.set_x(260)
    pdf.cell(20, 5, 'ADM &', 0, 0, 'R')
    pdf.set_x(277)
    pdf.cell(20, 5, 'JUMLAH', 0, 0, 'R')
    pdf.set_x(300)
    pdf.cell(20, 5, 'TOTAL', 0, 0, 'R')
    pdf.ln(4)

    pdf.set_x(152)
    pdf.cell(20, 5, 'OBAT', 0, 0, 'R')
    pdf.set_x(260)
    pdf.cell(20, 5, 'LAINNYA', 0, 0, 'R')
    pdf.set_x(277)
    pdf.cell(20, 5, 'BAYAR', 0, 0, 'R')
    pdf.ln(4)

    pdf.set_x(10)
    pdf.cell(300, 5, '-' * 292, 0, 1, 'L')

    totals = {
        'jasa': Decimal('0'),
        'tindakan': Decimal('0'),
        'obat': Decimal('0'),
        'ppn_obat': Decimal('0'),
        'embalasi': Decimal('0'),
        'lab': Decimal('0'),
        'rad': Decimal('0'),
        'bhp': Decimal('0'),
        'alat': Decimal('0'),
        'ambulan': Decimal('0'),
        'lainnya': Decimal('0'),
        'jmlbyr': Decimal('0'),
        'ttl': Decimal('0'),
    }

    def d(row, key):
        return Decimal(row.get(key) or 0)

    def fmt(value, dec=0):
        return f"{Decimal(value or 0):,.{dec}f}"

    for idx, row in enumerate(rows, start=1):
        nama = str(row.get('nama') or '')
        noreg = str(row.get('noreg') or '')

        if len(nama) > 23:
            nama_tampil = f"{nama[:23]}... ({noreg})"
        else:
            nama_tampil = f"{nama} ({noreg})"

        tgl_masuk = row.get('tgl_masuk')
        if hasattr(tgl_masuk, 'strftime'):
            tgl_masuk_text = tgl_masuk.strftime('%d-%m-%Y')
        else:
            tgl_masuk_text = str(tgl_masuk or '')

        farmasi_ppn = get_kunjungan_farmasi_ppn_detail(row.get('no'))

        obat = farmasi_ppn['obat']
        ppn_obat = farmasi_ppn['ppn_obat']
        embalasi = farmasi_ppn['embalasi']
        lainnya_adm = d(row, 'lainnya') + d(row, 'adm')

        pdf.set_x(10)
        pdf.cell(10, 5, f'{idx}.', 0, 0, 'L')
        pdf.set_x(20)
        pdf.cell(10, 5, nama_tampil, 0, 0, 'L')
        pdf.set_x(75)
        pdf.cell(25, 5, tgl_masuk_text, 0, 0, 'R')
        pdf.set_x(95)
        pdf.cell(20, 5, fmt(d(row, 'jasa'), 0), 0, 0, 'R')
        pdf.set_x(115)
        pdf.cell(20, 5, fmt(d(row, 'tindakan'), 0), 0, 0, 'R')
        pdf.set_x(135)
        pdf.cell(20, 5, fmt(obat, 2), 0, 0, 'R')
        pdf.set_x(152)
        pdf.cell(20, 5, fmt(ppn_obat, 2), 0, 0, 'R')
        pdf.set_x(168)
        pdf.cell(15, 5, fmt(embalasi, 0), 0, 0, 'R')
        pdf.set_x(185)
        pdf.cell(15, 5, fmt(d(row, 'lab'), 0), 0, 0, 'R')
        pdf.set_x(197)
        pdf.cell(20, 5, fmt(d(row, 'rad'), 0), 0, 0, 'R')
        pdf.set_x(215)
        pdf.cell(20, 5, fmt(d(row, 'bhp'), 0), 0, 0, 'R')
        pdf.set_x(230)
        pdf.cell(20, 5, fmt(d(row, 'alat'), 0), 0, 0, 'R')
        pdf.set_x(244)
        pdf.cell(20, 5, fmt(d(row, 'ambulan'), 0), 0, 0, 'R')
        pdf.set_x(260)
        pdf.cell(20, 5, fmt(lainnya_adm, 0), 0, 0, 'R')
        pdf.set_x(277)
        pdf.cell(20, 5, fmt(d(row, 'jmlbyr'), 2), 0, 0, 'R')
        pdf.set_x(300)
        pdf.cell(20, 5, fmt(d(row, 'ttl'), 2), 0, 0, 'R')
        pdf.ln(5)

        totals['jasa'] += d(row, 'jasa')
        totals['tindakan'] += d(row, 'tindakan')
        totals['obat'] += obat
        totals['ppn_obat'] += ppn_obat
        totals['embalasi'] += embalasi
        totals['lab'] += d(row, 'lab')
        totals['rad'] += d(row, 'rad')
        totals['bhp'] += d(row, 'bhp')
        totals['alat'] += d(row, 'alat')
        totals['ambulan'] += d(row, 'ambulan')
        totals['lainnya'] += lainnya_adm
        totals['jmlbyr'] += d(row, 'jmlbyr')
        totals['ttl'] += d(row, 'ttl')

    pdf.set_x(10)
    pdf.cell(300, 5, '-' * 292, 0, 1, 'L')

    pdf.set_x(75)
    pdf.cell(20, 5, '', 0, 0, 'R')
    pdf.set_x(95)
    pdf.cell(20, 5, fmt(totals['jasa'], 0), 0, 0, 'R')
    pdf.set_x(115)
    pdf.cell(20, 5, fmt(totals['tindakan'], 0), 0, 0, 'R')
    pdf.set_x(135)
    pdf.cell(20, 5, fmt(totals['obat'], 2), 0, 0, 'R')
    pdf.set_x(152)
    pdf.cell(20, 5, fmt(totals['ppn_obat'], 2), 0, 0, 'R')
    pdf.set_x(168)
    pdf.cell(15, 5, fmt(totals['embalasi'], 0), 0, 0, 'R')
    pdf.set_x(185)
    pdf.cell(15, 5, fmt(totals['lab'], 0), 0, 0, 'R')
    pdf.set_x(197)
    pdf.cell(20, 5, fmt(totals['rad'], 0), 0, 0, 'R')
    pdf.set_x(215)
    pdf.cell(20, 5, fmt(totals['bhp'], 0), 0, 0, 'R')
    pdf.set_x(230)
    pdf.cell(20, 5, fmt(totals['alat'], 0), 0, 0, 'R')
    pdf.set_x(244)
    pdf.cell(20, 5, fmt(totals['ambulan'], 0), 0, 0, 'R')
    pdf.set_x(260)
    pdf.cell(20, 5, fmt(totals['lainnya'], 0), 0, 0, 'R')
    pdf.set_x(277)
    pdf.cell(20, 5, fmt(totals['jmlbyr'], 2), 0, 0, 'R')
    pdf.set_x(300)
    pdf.cell(20, 5, fmt(totals['ttl'], 2), 0, 0, 'R')

    pdf.ln(4)
    pdf.set_x(10)
    pdf.cell(300, 5, '-' * 292, 0, 1, 'L')

    pdf_bytes = bytes(pdf.output(dest='S'))

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Rincian_Invoice_PPN_{faktur.nomor_faktur}.pdf"'
    return response

def _kwitansi_rupiah(value):
    amount = Decimal(value or 0).quantize(Decimal('1'))
    return f"Rp {int(amount):,}".replace(',', '.') + ",-"

def _pdf_wrap_line(pdf, text, max_width):
    words = str(text or '').split()
    if not words:
        return ['']
    lines = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if pdf.get_string_width(candidate) <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines

def render_kwitansi_pdf_response(faktur):
    pdf = FPDF('P', 'mm', 'A4')
    pdf.set_auto_page_break(False)
    pdf.add_page()

    margin_x = 8
    tanggal_obj = faktur.tanggal or timezone.localdate()
    tanggal_str = f"{tanggal_obj.day:02d} {_ID_MONTHS[tanggal_obj.month]} {tanggal_obj.year}"
    no_kwitansi = _legacy_invoice_number(faktur)

    pembiayaan_detail = get_pembiayaan_detail(faktur.id_pembiayaan)
    nama_pbiaya = (
        pembiayaan_detail.get('pembiayaan')
        or _invoice_pembiayaan_name(faktur)
        or 'PEMBIAYAAN'
    )

    ttl_tagihan, jml_bayar = _invoice_print_amounts(faktur)
    if faktur.xround == 'Y':
        total_tagihan = ttl_tagihan.to_integral_value(rounding='ROUND_CEILING')
    else:
        total_tagihan = ttl_tagihan.quantize(Decimal('1'))
    terbilang = f"{_legacy_words(total_tagihan).title()} Rupiah"
    terbilang_words = terbilang.split()
    terbilang_line_1 = terbilang
    terbilang_line_2 = ''
    if len(terbilang) > 58:
        split_at = max(1, len(terbilang_words) // 2)
        terbilang_line_1 = ' '.join(terbilang_words[:split_at])
        terbilang_line_2 = ' '.join(terbilang_words[split_at:])
    amount_plain = f"{int(total_tagihan):,}".replace(',', '.')
    jenis_lines = [
        line.strip().upper()
        for line in str(faktur.jenis or '').replace('\r\n', '\n').replace('\r', '\n').split('\n')
        if line.strip()
    ]
    periode_text = str(faktur.periode or '').strip().upper()
    if not jenis_lines:
        jenis_lines = ['']
    pembayaran_lines = []
    for index, line in enumerate(jenis_lines):
        prefix = 'BIAYA ' if index == 0 else ''
        pembayaran_lines.append(f"{prefix}{line}".strip())
    if periode_text:
        periode_suffix = f"PERIODE {periode_text}"
        if pembayaran_lines[-1]:
            pembayaran_lines[-1] = f"{pembayaran_lines[-1]} {periode_suffix}"
        else:
            pembayaran_lines[-1] = periode_suffix
    pembayaran_lines = pembayaran_lines[:4]

    pdf.set_draw_color(0, 0, 0)
    pdf.set_line_width(0.25)
    pdf.rect(0.8, 0.8, 208.4, 146.4)

    logo_path = get_invoice_logo_path()
    if logo_path:
        pdf.image(logo_path, x=168, y=3, h=21)

    pdf.set_font('Times', 'B', 18)
    pdf.set_xy(38, 8)
    pdf.cell(58, 8, 'KWITANSI', 0, 1, 'C')
    pdf.line(41, 16, 94, 16)
    pdf.set_font('Arial', '', 9)
    pdf.set_xy(40, 18)
    pdf.cell(80, 5, f"NO.  {no_kwitansi}", 0, 1, 'L')

    pdf.set_xy(18, 32)
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(39, 4, 'SUDAH TERIMA DARI', 0, 0, 'L')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(92, 4, str(nama_pbiaya), 0, 1, 'L')
    pdf.set_x(18)
    pdf.set_font('Arial', '', 9)
    pdf.cell(39, 4, 'RECEIVED FROM', 0, 1, 'L')

    box_x = 16
    box_y = 45
    box_w = 178
    box_h = 33
    pdf.rect(box_x, box_y, box_w, box_h)

    pdf.set_xy(box_x + 2, box_y + 4)
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(40, 4, 'UANG SEJUMLAH', 0, 1, 'L')
    pdf.set_x(box_x + 2)
    pdf.set_font('Arial', '', 9)
    pdf.cell(40, 4, 'THE SUM OF', 0, 1, 'L')

    words_x = box_x + 45
    words_y = box_y + 4
    words_w = box_w - 48
    pdf.set_fill_color(205, 205, 205)
    pdf.rect(words_x, words_y, words_w, 9, 'DF')
    pdf.set_xy(words_x + 2, words_y + 1.7)
    pdf.set_font('Times', 'I', 9.5)
    pdf.cell(words_w - 4, 5, terbilang_line_1, 0, 1, 'C')

    pdf.set_fill_color(205, 205, 205)
    pdf.rect(box_x + 4, box_y + 17, box_w - 8, 9, 'DF')
    pdf.set_xy(box_x + 12, box_y + 18.7)
    pdf.set_font('Times', 'I', 9.5)
    pdf.cell(box_w - 20, 5, terbilang_line_2, 0, 1, 'L')

    pay_y = 80
    pay_h = 29
    pdf.rect(box_x, pay_y, box_w, pay_h)
    pdf.set_xy(box_x + 2, pay_y + 4)
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(43, 4, 'UNTUK PEMBAYARAN', 0, 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(3, 4, ':', 0, 0, 'C')
    pdf.set_font('Arial', '', 8.6)
    text_x = box_x + 48
    text_y = pay_y + 4.1
    text_w = box_w - 51
    payment_line_gap = 6.3
    wrapped_payment_lines = []
    for line in pembayaran_lines:
        wrapped_payment_lines.extend(_pdf_wrap_line(pdf, line, text_w))
    wrapped_payment_lines = wrapped_payment_lines[:4]
    for index, line in enumerate(wrapped_payment_lines):
        pdf.set_xy(text_x, text_y + (index * payment_line_gap))
        pdf.cell(text_w, 4.2, line, 0, 1, 'L')
    pdf.set_xy(box_x + 2, pay_y + 8)
    pdf.set_x(box_x + 2)
    pdf.set_font('Arial', '', 9)
    pdf.cell(43, 4, 'IN PAYMENT OF', 0, 1, 'L')

    for y in (pay_y + 8.9, pay_y + 15.2, pay_y + 21.5, pay_y + 27.8):
        pdf.line(box_x + 48, y, box_x + box_w - 3, y)

    sign_x = 119
    sign_w = 72
    pdf.set_xy(sign_x, 111)
    pdf.set_font('Arial', '', 9)
    pdf.cell(sign_w, 5, f"Samarinda, {tanggal_str}", 0, 1, 'C')

    pdf.set_fill_color(205, 205, 205)
    amount_x = 18
    amount_y = 122
    pdf.rect(amount_x + 7, amount_y, 74, 8, 'DF')
    pdf.set_xy(amount_x + 7, amount_y + 1.3)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(74, 5, f"Rp {amount_plain},-", 0, 1, 'C')

    pdf.set_xy(sign_x, 142)
    pdf.set_font('Arial', '', 9)
    pdf.cell(sign_w, 4, 'Nevi Nevada, S.Ip., M.M.', 0, 1, 'C')

    pdf_bytes = bytes(pdf.output(dest='S'))
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Kwitansi_{faktur.nomor_faktur}.pdf"'
    return response

def _render_legacy_invoice(faktur, mode):
    """Replicate exact FPDF output from app_siaga print_invoice.php"""
    total, jml_bayar = _invoice_print_amounts(faktur)
    
    # Build cost rows dengan labels yang persis sama seperti PHP
    cost_rows_html = []
    
    # BIAYA JASA
    if faktur.jasa and faktur.jasa != 0:
        cost_rows_html.append((f"-  BIAYA JASA", faktur.jasa))
    
    # BIAYA TINDAKAN
    if faktur.tindakan and faktur.tindakan != 0:
        cost_rows_html.append((f"-  BIAYA TINDAKAN", faktur.tindakan))
    
    # BIAYA FARMASI (farmasi - ppn_farmasi jika ada)
    if faktur.farmasi and faktur.farmasi != 0:
        farmasi_display = faktur.farmasi - (faktur.ppn_farmasi or 0)
        if farmasi_display != 0:
            cost_rows_html.append((f"-  BIAYA FARMASI", farmasi_display))
    
    # BIAYA LABORATORIUM
    if faktur.lab and faktur.lab != 0:
        cost_rows_html.append((f"-  BIAYA LABORATORIUM", faktur.lab))
    
    # BIAYA RADIOLOGI
    if faktur.rad and faktur.rad != 0:
        cost_rows_html.append((f"-  BIAYA RADIOLOGI", faktur.rad))
    
    # BIAYA KAMAR
    if faktur.kamar and faktur.kamar != 0:
        cost_rows_html.append((f"-  BIAYA KAMAR", faktur.kamar))
    
    # BIAYA BAHAN HABIS PAKAI (BHP)
    if faktur.bhp and faktur.bhp != 0:
        cost_rows_html.append((f"-  BIAYA BAHAN HABIS PAKAI (BHP)", faktur.bhp))
    
    # BIAYA ALAT
    if faktur.alat and faktur.alat != 0:
        cost_rows_html.append((f"-  BIAYA ALAT", faktur.alat))
    
    # BIAYA AMBULAN
    if faktur.ambulan and faktur.ambulan != 0:
        cost_rows_html.append((f"-  BIAYA AMBULAN", faktur.ambulan))
    
    # BIAYA FISIOTERAPI
    if faktur.fisio and faktur.fisio != 0:
        cost_rows_html.append((f"-  BIAYA FISIOTERAPI", faktur.fisio))
    
    # BIAYA ADMINISTRASI DAN LAINNYA
    if (faktur.lainnya or faktur.adm) and (faktur.lainnya or 0) + (faktur.adm or 0) != 0:
        cost_rows_html.append((f"-  BIAYA ADMINISTRASI DAN LAINNYA", (faktur.lainnya or 0) + (faktur.adm or 0)))
    
    # PPN OBAT (jika ppn=False dan ppn_farmasi > 0)
    if mode != 'invoice_ppn' and faktur.ppn_farmasi and faktur.ppn_farmasi > 0:
        cost_rows_html.append((f"-  PPN OBAT", faktur.ppn_farmasi))

    if faktur.is_cob and Decimal(faktur.tanggungan_bpjs or 0) > 0:
        cost_rows_html.append((f"-  DITANGGUNG BPJS (INA-CBGs)", -Decimal(faktur.tanggungan_bpjs)))

    cost_rows_html.append(("-  JUMLAH PEMBAYARAN", jml_bayar))
    
    # Build HTML untuk cost items
    cost_items_html = []
    for label, value in cost_rows_html:
        cost_items_html.append(
            f"<div style=\"display: flex; margin-bottom: 1px;\">"
            f"  <div style=\"flex: 1; font-size: 11px;\">{escape(label)}</div>"
            f"  <div style=\"width: 50px; text-align: right; font-size: 11px;\">Rp.</div>"
            f"  <div style=\"width: 95px; text-align: right; font-size: 11px;\">{_invoice_money(value, 2)}</div>"
            f"</div>"
        )
    cost_items_str = ''.join(cost_items_html)
    
    # Total row (bold)
    total_html = (
        f"<div style=\"display: flex; margin-bottom: 1px; border-top: 1px solid #000; padding-top: 1px; font-weight: bold; font-size: 12px;\">"
        f"  <div style=\"flex: 1;\">Total</div>"
        f"  <div style=\"width: 50px; text-align: right;\">Rp.</div>"
        f"  <div style=\"width: 95px; text-align: right;\">{_invoice_money(total, 2)}</div>"
        f"</div>"
    )
    
    tanggal_obj = faktur.tanggal or timezone.localdate()
    tanggal_str = f"{tanggal_obj.day:02d} {_ID_MONTHS[tanggal_obj.month]} {tanggal_obj.year}"
    
    pembiayaan_detail = get_pembiayaan_detail(faktur.id_pembiayaan)

    nama_pbiaya = (
        pembiayaan_detail.get('pembiayaan')
        or _invoice_pembiayaan_name(faktur)
        or 'PEMBIAYAAN'
    )

    alamat = pembiayaan_detail.get('alamat') or ''

    alamat_html = ''
    if alamat.strip():
        alamat_html = f"<div style=\"font-size: 11px;\">{escape(alamat)}</div>"
    
    # Terbilang conversion
    terbilang = _legacy_words(total).upper()
    
    return f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Invoice {escape(faktur.nomor_faktur)}</title>
    <style>
        @page {{ size: A4 portrait; margin: 13mm 15mm 10mm 15mm; }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Times New Roman', Times, serif; font-size: 11px; color: #000; line-height: 1.3; }}
        .container {{ width: 100%; position: relative; }}
        .header {{ position: relative; height: 18mm; margin-bottom: 0; }}
        .logo {{ position: absolute; right: 0; top: -5px; height: 65px; text-align: right; }}
        .logo img {{ height: 65px; width: auto; max-height: 65px; object-fit: contain; }}
        h1 {{ text-align: center; font-size: 16px; font-weight: bold; text-decoration: underline; margin: 2mm 0; letter-spacing: 0; }}
        .content {{ margin-top: 1mm; }}
        .to-section {{ margin-bottom: 1.5mm; line-height: 1.5; font-size: 11px; }}
        .meta-section {{ display: flex; justify-content: space-between; margin-bottom: 1mm; font-size: 11px; line-height: 1.5; }}
        .meta-left {{ width: 45%; }}
        .meta-right {{ width: 45%; text-align: left; }}
        .divider {{ border-top: 1px solid #000; border-bottom: 1px solid #000; height: 2px; margin: 1mm 0; }}
        .desc {{ margin-bottom: 2mm; font-size: 11px; line-height: 1.5; }}
        .costs {{ margin-bottom: 2mm; font-size: 11px; }}
        .note {{ margin: 1.5mm 0 1mm 0; font-style: italic; font-size: 11px; }}
        .words {{ margin-bottom: 2mm; font-style: italic; font-size: 11px; }}
        .divider-double {{ border-top: 3px double #000; margin: 1.5mm 0; }}
        .bank {{ margin-bottom: 2mm; font-size: 10px; line-height: 1.5; }}
        .footer {{ display: flex; justify-content: space-between; margin-top: 2mm; font-size: 11px; line-height: 1.5; }}
        .cc {{ }}
        .signature {{ text-align: right; }}
        .sig-space {{ height: 18mm; }}
        .sig-name {{ font-weight: bold; text-decoration: underline; }}
        .print-btn {{ position: fixed; top: 10px; right: 10px; padding: 8px 12px; background: #0f766e; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 11px; z-index: 1000; }}
        @media print {{ .print-btn {{ display: none; }} }}
    </style>
</head>
<body onload="setTimeout(() => window.print(), 250)">
    <button class="print-btn" onclick="window.print()">Print</button>
    
    <div class="container">
        <div class="header">
            <div class="logo">
                <img src="/static/images/logo-1.jpg" alt="Logo">
            </div>
        </div>
        
        <h1>INVOICE</h1>
        
        <div class="content">
            <div class="to-section">
                <div>Kepada Yth.</div>
                <div>Bagian Keuangan</div>
                <div style="margin-top: 1mm;"><b>{escape(nama_pbiaya)}</b></div>
                {alamat_html}
            </div>
            
            <div class="meta-section">
                <div class="meta-left"></div>
                <div class="meta-right">
                    <div style="display: flex;"><div style="width: 50px;">Nomor</div><div>: {escape(_legacy_invoice_number(faktur))}</div></div>
                    <div style="display: flex;"><div style="width: 50px;">Tanggal</div><div>: {tanggal_str}</div></div>
                </div>
            </div>
            
            <div class="divider"></div>
            
            <div class="desc">
                <div>{escape(faktur.jenis or 'RAWAT INAP')}</div>
                <div>BEBAN {escape(faktur.beban or 'RUMAH SAKIT')}</div>
                <div>PERIODE : {escape(faktur.periode or '')}</div>
            </div>
            
            <div class="costs">
                {cost_items_str}
                {total_html}
            </div>
            
            <div class="note">#REKAPITULASI &amp; BACKUP TERLAMPIR</div>
            
            <div class="words">#{terbilang} RUPIAH#</div>
            
            <div class="divider-double"></div>
            
            <div class="bank">
                <div>Jangka waktu pelunasan maksimal 14 hari setelah</div>
                <div>diterimanya tagihan ini, ke rekening Kami</div>
                <div style="font-weight: bold; margin-top: 0.5mm;">No. 777.998.9978</div>
                <div>Bank Syariah Indonesia (BSI)</div>
                <div style="font-weight: bold;">a.n. RS Siaga Al Munawwarah</div>
            </div>
            
            <div class="footer">
                <div class="cc">
                    <div>cc.</div>
                    <div>- Akuntansi RS-SAMS</div>
                    <div>- Arsip</div>
                </div>
                <div class="signature">
                    <div>RS. SIAGA AL MUNAWWARAH</div>
                    <div>SAMARINDA</div>
                    <div class="sig-space"></div>
                    <div class="sig-name">Nevi Nevada, S.Ip., M.M.</div>
                </div>
            </div>
        </div>
    </div>
</body>
</html>"""

def _render_legacy_rincian(faktur, mode):
    """Replicate exact FPDF rincian output from app_siaga"""
    rows = _invoice_print_rows(faktur, ppn=mode == 'rincian_ppn')
    
    # Filter out zero values
    filtered_rows = [(label, value) for label, value in rows if value and value != 0]
    
    # Build table rows
    table_rows = []
    for idx, (label, value) in enumerate(filtered_rows, start=1):
        label_text = label.replace('-  ', '').strip()
        table_rows.append(
            f"<tr>"
            f"  <td style=\"width: 45px; text-align: center;\">{idx}</td>"
            f"  <td>{escape(label_text)}</td>"
            f"  <td style=\"width: 160px; text-align: right;\">{_invoice_money(value, 2)}</td>"
            f"</tr>"
        )
    
    table_rows_str = '\n'.join(table_rows)
    total = Decimal(faktur.total_tagihan or 0)
    
    tanggal_obj = faktur.tanggal or timezone.localdate()
    tanggal_str = f"{tanggal_obj.day:02d} {_ID_MONTHS[tanggal_obj.month]} {tanggal_obj.year}"
    
    return f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Rincian Invoice {escape(faktur.nomor_faktur)}</title>
    <style>
        @page {{ 
            size: A4 landscape; 
            margin: 10mm;
        }}
        * {{ 
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{ 
            font-family: 'Times New Roman', Times, serif;
            font-size: 11px;
            color: #000;
        }}
        .container {{
            width: 100%;
        }}
        h1 {{
            text-align: center;
            font-size: 15px;
            font-weight: bold;
            text-decoration: underline;
            margin-bottom: 3px;
        }}
        .sub {{
            text-align: center;
            margin-bottom: 8px;
            font-size: 11px;
            line-height: 1.4;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 11px;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 5px 6px;
            text-align: left;
        }}
        th {{
            background: #f2f2f2;
            text-align: center;
            font-weight: bold;
        }}
        .num {{
            text-align: right;
        }}
        .summary-row td {{
            font-weight: bold;
            border-top: 2px solid #000;
        }}
        .summary-row .num {{
            text-align: right;
        }}
        .print-btn {{
            position: fixed;
            top: 10px;
            right: 10px;
            padding: 8px 12px;
            background: #0f766e;
            color: white;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-size: 10px;
            z-index: 1000;
        }}
        @media print {{
            .print-btn {{ display: none; }}
        }}
    </style>
</head>
<body onload="setTimeout(() => window.print(), 250)">
    <button class="print-btn" onclick="window.print()">Print</button>
    
    <div class="container">
        <h1>RINCIAN INVOICE</h1>
        <div class="sub">
            <div>No. {escape(_legacy_invoice_number(faktur))}</div>
            <div>{escape(_invoice_pembiayaan_name(faktur))} - {escape(faktur.periode or '')}</div>
            <div>Tanggal: {tanggal_str}</div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>NO</th>
                    <th>URAIAN</th>
                    <th>JUMLAH</th>
                </tr>
            </thead>
            <tbody>
                {table_rows_str}
                <tr class="summary-row">
                    <td colspan="2">TOTAL</td>
                    <td class="num">{_invoice_money(total, 2)}</td>
                </tr>
            </tbody>
        </table>
    </div>
</body>
</html>"""

def faktur_legacy_print_view(request, pk):
    faktur = get_object_or_404(Faktur.objects.select_related('pelanggan'), pk=pk)
    mode = request.GET.get('mode', 'invoice')

    if mode in ('invoice', 'invoice_ppn'):
        return render_invoice_pdf_response(faktur, mode)

    if mode == 'rincian':
        return render_rincian_pdf_response(faktur, mode)
    
    if mode == 'rincian_ppn':
        return render_rincian_ppn_pdf_response(faktur, mode)

    if mode == 'kwitansi':
        return render_kwitansi_pdf_response(faktur)
        
    else:
        html = _render_legacy_invoice(faktur, mode)

    return HttpResponse(html)

def _receipt_display_date(value):
    value = (value or '').strip()
    if not value:
        return ''
    try:
        parsed = datetime.strptime(value, '%Y-%m-%d').date()
        return _invoice_date(parsed)
    except ValueError:
        return value

def _render_tanda_terima_invoice(fakturs, company_name, tanggal):
    logo_url = "/logo.png"
    company_name = company_name or '-'
    tanggal_label = _receipt_display_date(tanggal)
    total = sum((Decimal(item.total_tagihan or 0) for item in fakturs), Decimal('0'))

    invoice_rows = []
    for index, faktur in enumerate(fakturs, start=1):
        invoice_rows.append(
            '<tr>'
            f'<td class="row-no">{index}.</td>'
            f'<td class="invoice-no">{escape(_legacy_invoice_number(faktur))}</td>'
            f'<td class="currency">Rp.</td>'
            f'<td class="amount">{_invoice_money(faktur.total_tagihan, 2)}</td>'
            '</tr>'
        )
    rows_html = ''.join(invoice_rows)

    def receipt_html(label):
        return f"""
        <section class="receipt">
            <div class="copy-label">{escape(label)}</div>
            <header class="letterhead">
                <div class="hospital">
                    <h1>RS SIAGA AL MUNAWWARAH SAMARINDA</h1>
                    <p>Jl. Ramania No. 3 Sidodadi - Samarinda Ulu</p>
                    <p>Kota Samarinda - 75123, Fax. (0541) 7272700, Tlp. (0541) 739772 / 7272667</p>
                </div>
                <img src="{escape(logo_url)}" alt="Logo RS SIAGA AL MUNAWWARAH SAMARINDA">
            </header>
            <div class="rule"></div>

            <h2>TANDA TERIMA INVOICE</h2>

            <div class="company-line">
                <span>NAMA PERUSAHAAN :</span>
                <strong>{escape(company_name)}</strong>
            </div>

            <div class="invoice-block">
                <div class="invoice-lines">
                    <table class="invoice-table">
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>
                </div>
                <table class="invoice-table invoice-total">
                    <tbody>
                        <tr class="total-row">
                            <td class="row-no"></td>
                            <td class="invoice-no">TOTAL</td>
                            <td class="currency">Rp.</td>
                            <td class="amount">{_invoice_money(total, 2)}</td>
                        </tr>
                    </tbody>
                </table>
            </div>

            <div class="documents">
                <h3>DOKUMEN TERDIRI DARI :</h3>
                <div>
                    <ol>
                        <li>INVOICE</li>
                        <li>KWITANSI</li>
                        <li>PO</li>
                    </ol>
                    <ol start="4">
                        <li>NOTA KREDIT</li>
                        <li>BAPB</li>
                        <li>FP</li>
                    </ol>
                </div>
            </div>

            <div class="date-line">SAMARINDA, <span>{escape(tanggal_label)}</span></div>

            <div class="signatures">
                <div>
                    <strong>DISERAHKAN</strong>
                    <i></i>
                    <span></span>
                    <b>NAMA TERANG</b>
                </div>
                <div>
                    <strong>DITERIMA</strong>
                    <i></i>
                    <span></span>
                    <b>NPK.</b>
                </div>
            </div>
        </section>
    """
    receipt_one = receipt_html("Lembar 1: Arsip")
    receipt_two = receipt_html("Lembar 2: Asuransi/Perusahaan")

    html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Tanda Terima Invoice</title>
    <style>
        @page {{ size: A4 landscape; margin: 7mm; }}
        * {{ box-sizing: border-box; }}
        body {{
            margin: 0;
            color: #111;
            background: #f3f4f6;
            font-family: "Times New Roman", Times, serif;
        }}
        .print-btn {{
            position: fixed;
            top: 10px;
            right: 10px;
            z-index: 10;
            border: 0;
            border-radius: 5px;
            background: #0f766e;
            color: #fff;
            padding: 8px 12px;
            font: 700 12px Arial, sans-serif;
            cursor: pointer;
        }}
        .sheet {{
            width: 100%;
            max-width: 283mm;
            height: 196mm;
            margin: 0 auto;
            background: #fff;
            display: grid;
            grid-template-columns: 1fr 1fr;
        }}
        .receipt {{
            position: relative;
            width: 100%;
            height: 196mm;
            padding: 5mm 6mm 5mm;
            overflow: hidden;
            display: grid;
            grid-template-rows: 23mm 2mm 15mm 9mm minmax(0, 1fr) 28mm 9mm 35mm;
            row-gap: 0;
        }}
        .copy-label {{
            position: absolute;
            bottom: 1.5mm;
            left: 50%;
            transform: translateX(-50%);
            font: 700 7.5pt Arial, sans-serif;
            color: #111;
            white-space: nowrap;
        }}
        .receipt + .receipt {{
            border-top: 0;
            border-left: 1px dashed #777;
        }}
        .letterhead {{
            height: 23mm;
            display: grid;
            grid-template-columns: 1fr 20mm;
            align-items: center;
            gap: 4mm;
            text-align: center;
        }}
        .hospital h1 {{
            margin: 0 0 1mm;
            font-size: 12.5pt;
            line-height: 1.1;
            font-weight: 900;
        }}
        .hospital p {{
            margin: 0;
            font-size: 8.5pt;
            line-height: 1.25;
            font-weight: 700;
        }}
        .letterhead img {{
            max-width: 19mm;
            max-height: 19mm;
            object-fit: contain;
        }}
        .rule {{
            border-top: 2px solid #111;
            border-bottom: 1px solid #111;
            height: 2mm;
        }}
        h2 {{
            margin: 0;
            text-align: center;
            font-size: 17pt;
            line-height: 1;
            font-weight: 900;
            text-decoration: underline;
            align-self: end;
            padding-bottom: 2mm;
        }}
        .company-line {{
            display: flex;
            align-items: baseline;
            gap: 3mm;
            margin: 0;
            font-size: 12pt;
            font-weight: 900;
            align-self: center;
        }}
        .company-line span {{
            white-space: nowrap;
        }}
        .invoice-block {{
            min-height: 0;
            display: grid;
            grid-template-rows: minmax(0, 1fr) 8mm;
            align-self: stretch;
        }}
        .invoice-lines {{
            min-height: 0;
            overflow: hidden;
        }}
        .invoice-table {{
            width: calc(100% - 3mm);
            border-collapse: collapse;
            margin-left: 3mm;
            font-size: 10.5pt;
            font-weight: 900;
        }}
        .invoice-table td {{
            padding: 1mm 0;
            vertical-align: top;
        }}
        .row-no {{ width: 7mm; }}
        .invoice-no {{ width: auto; }}
        .currency {{ width: 9mm; }}
        .amount {{
            width: 33mm;
            text-align: right;
            padding-right: 0 !important;
            white-space: nowrap;
        }}
        .invoice-total {{
            align-self: end;
        }}
        .total-row td {{
            padding-top: 1mm;
            font-size: 11.5pt;
        }}
        .total-row td:nth-child(2) {{
            text-align: center;
        }}
        .documents {{
            margin-top: 0;
            font-size: 10.5pt;
            font-weight: 900;
            align-self: center;
        }}
        .documents h3 {{
            margin: 0 0 2mm;
            font-size: 11.5pt;
        }}
        .documents > div {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 6mm;
            padding-left: 7mm;
        }}
        .documents ol {{
            margin: 0;
            padding-left: 7mm;
        }}
        .date-line {{
            margin-top: 0;
            text-align: right;
            padding-right: 4mm;
            font-size: 10.5pt;
            font-weight: 900;
            align-self: center;
        }}
        .date-line span {{
            display: inline;
            text-align: center;
        }}
        .signatures {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 12mm;
            margin-top: 0;
            font-size: 10.5pt;
            font-weight: 900;
            align-self: stretch;
            padding-bottom: 7mm;
        }}
        .signatures div {{
            display: grid;
            grid-template-rows: auto 1fr auto auto;
            gap: 2mm;
            align-content: stretch;
        }}
        .signatures i {{
            display: block;
            min-height: 0;
        }}
        .signatures span {{
            display: block;
            border-bottom: 1px solid #111;
            height: 0;
            width: 50mm;
        }}
        .signatures b {{
            font-size: 10.5pt;
        }}
        @media print {{
            body {{ background: #fff; }}
            .print-btn {{ display: none; }}
            .sheet {{ width: 100%; height: 196mm; }}
        }}
    </style>
</head>
<body onload="setTimeout(() => window.print(), 250)">
    <button class="print-btn" onclick="window.print()">Print</button>
    <main class="sheet">
        {receipt_one}
        {receipt_two}
    </main>
</body>
</html>
"""
    return html

def faktur_tanda_terima_print_view(request):
    raw_ids = request.GET.get('ids', '')
    ids = [item.strip() for item in raw_ids.split(',') if item.strip()]
    if not ids:
        return HttpResponse('<h3>Pilih minimal satu invoice.</h3>', status=400)

    fakturs = list(
        Faktur.objects
        .select_related('pelanggan')
        .filter(pk__in=ids)
        .order_by('nomor_faktur')
    )
    if not fakturs:
        return HttpResponse('<h3>Invoice tidak ditemukan.</h3>', status=404)

    html = _render_tanda_terima_invoice(
        fakturs,
        request.GET.get('perusahaan', ''),
        request.GET.get('tanggal', ''),
    )
    return HttpResponse(html)

def build_pembiayaan_name_map(ids):
    from django.db import connection

    ids = [str(item) for item in ids if item]
    if not ids:
        return {}

    placeholders = ','.join(['%s'] * len(ids))

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT id_pembiayaan, pembiayaan
                FROM rssams.pbiaya
                WHERE id_pembiayaan IN ({placeholders})
                """,
                ids
            )
            return {str(row[0]): row[1] for row in cursor.fetchall()}
    except Exception:
        return {}

def get_pembiayaan_detail(id_pembiayaan):
    from django.db import connection

    if not id_pembiayaan:
        return {}

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id_pembiayaan, pembiayaan, alamat
                FROM rssams.pbiaya
                WHERE id_pembiayaan = %s
                LIMIT 1
                """,
                [id_pembiayaan]
            )
            row = cursor.fetchone()

        if not row:
            return {}

        return {
            'id_pembiayaan': str(row[0] or ''),
            'pembiayaan': row[1] or '',
            'alamat': row[2] or '',
        }
    except Exception:
        return {}

def faktur_rekap_print_view(request):
    """Print rekapitulasi invoice berdasarkan date range dan pembiayaan (optional)"""
    dari = request.GET.get('dari', '')
    sampai = request.GET.get('sampai', '')
    id_pembiayaan = request.GET.get('id_pembiayaan', '').strip()
    
    if not dari or not sampai:
        return HttpResponse('<h3>Parameter tanggal (dari/sampai) wajib diisi</h3>', status=400)
    
    # Query faktur dengan filter date range (mengabaikan invoice status batal)
    fakturs_qs = Faktur.objects.filter(
        tanggal__gte=dari,
        tanggal__lte=sampai
    ).exclude(status='batal')

    if id_pembiayaan:
        fakturs_qs = fakturs_qs.filter(Q(id_pembiayaan=str(id_pembiayaan)) | Q(id_pembiayaan=id_pembiayaan))

    fakturs = fakturs_qs.select_related('pelanggan').order_by('id_pembiayaan', 'tanggal')
    
    pembiayaan_map = build_pembiayaan_name_map(
        fakturs.values_list('id_pembiayaan', flat=True).distinct()
    )

    header_title = 'REKAPITULASI INVOICE'
    if id_pembiayaan:
        p_key = int(id_pembiayaan) if id_pembiayaan.isdigit() else id_pembiayaan
        p_name = pembiayaan_map.get(p_key) or pembiayaan_map.get(str(p_key))
        if p_name:
            header_title = f'REKAPITULASI INVOICE - {escape(p_name.upper())}'
    
    # Build HTML
    html_parts = [
        '<!DOCTYPE html>',
        '<html>',
        '<head>',
        '<meta charset="utf-8">',
        '<title>Rekapitulasi Invoice</title>',
        '<style>',
        'body { font-family: Arial, sans-serif; font-size: 10pt; }',
        '#wrapper { width: 2200px; margin: 0 auto; padding: 15px; }',
        '#isi { border: none; padding: 15px; }',
        'h3 { font-size: 14pt; margin-bottom: 20px; }',
        'table { border-collapse: collapse; margin: 10px 0; width: 100%; }',
        'th { background-color: #333; color: white; padding: 8px; border: 1px solid #999; text-align: center; }',
        'td { padding: 6px; border: 1px solid #999; }',
        '.text-right { text-align: right; }',
        '.text-center { text-align: center; }',
        '.mono { font-family: monospace; }',
        '</style>',
        '</head>',
        '<body>',
        '<body>',
        '<div id="wrapper">',
        '<div id="isi">',
        f'<h3>{header_title}<br>Periode: {dari} s/d {sampai}</h3>',
        '<table>',
        '<thead>',
        '<tr>',
        '<th>No</th>',
        '<th>No Invoice</th>',
        '<th>Tanggal</th>',
        '<th>Pembiayaan</th>',
        '<th class="text-right">Adm</th>',
        '<th class="text-right">Jasa</th>',
        '<th class="text-right">Farmasi</th>',
        '<th class="text-right">Tindakan</th>',
        '<th class="text-right">Fisio</th>',
        '<th class="text-right">Lab</th>',
        '<th class="text-right">Rad</th>',
        '<th class="text-right">Kamar</th>',
        '<th class="text-right">BHP</th>',
        '<th class="text-right">Ambulan</th>',
        '<th class="text-right">Alat</th>',
        '<th class="text-right">Lain-lain</th>',
        '<th class="text-right">Total Pendapatan</th>',
        '<th class="text-right">Piutang P3</th>',
        '<th class="text-right">Dibayar Pasien</th>',
        '<th class="text-right">Jml Bayar</th>',
        '<th class="text-right">Sisa Tagihan</th>',
        '<th>Jatuh Tempo</th>',
        '<th>Tgl Kirim</th>',
        '<th>Status</th>',
        '</tr>',
        '</thead>',
        '<tbody>',
    ]
    
    # Query dp3 & jmlbyr per no_invoice
    nomor_fakturs = [f.nomor_faktur for f in fakturs if f.nomor_faktur]
    kunjung_map = {}
    if nomor_fakturs:
        try:
            with connection.cursor() as cursor:
                placeholders = ','.join(['%s'] * len(nomor_fakturs))
                cursor.execute(f"""
                    SELECT 
                        c.no_invoice,
                        COALESCE(SUM(COALESCE(a.dp3, 0)), 0) AS total_dp3,
                        COALESCE(SUM(COALESCE(a.jmlbyr, 0)), 0) AS total_jmlbyr
                    FROM rssams.kunjung a
                    INNER JOIN rssams.verif_kunjung c ON a.no = c.no
                    WHERE c.no_invoice IN ({placeholders})
                    GROUP BY c.no_invoice
                """, nomor_fakturs)
                for r_row in cursor.fetchall():
                    kunjung_map[str(r_row[0])] = {
                        'dp3': Decimal(str(r_row[1] or 0)),
                        'jmlbyr': Decimal(str(r_row[2] or 0)),
                    }
        except Exception:
            pass

    total_pendapatan = Decimal('0.00')
    total_dp3 = Decimal('0.00')
    total_dibayar_pasien = Decimal('0.00')
    total_dibayar = Decimal('0.00')
    total_tagihan = Decimal('0.00')
    no = 1
    
    for f in fakturs:
        total_biaya = (
            Decimal(f.adm or 0) +
            Decimal(f.jasa or 0) +
            Decimal(f.farmasi or 0) +
            Decimal(f.tindakan or 0) +
            Decimal(f.fisio or 0) +
            Decimal(f.lab or 0) +
            Decimal(f.rad or 0) +
            Decimal(f.kamar or 0) +
            Decimal(f.bhp or 0) +
            Decimal(f.lainnya or 0) +
            Decimal(f.ambulan or 0) +
            Decimal(f.alat or 0)
        )

        k_info = kunjung_map.get(str(f.nomor_faktur or ''), {'dp3': Decimal('0'), 'jmlbyr': Decimal('0')})
        dp3_val = k_info['dp3']
        dibayar_pasien_val = k_info['jmlbyr']

        jml_bayar = Decimal(f.total_dibayar or 0)
        ttl = total_biaya - jml_bayar
        
        total_pendapatan += total_biaya
        total_dp3 += dp3_val
        total_dibayar_pasien += dibayar_pasien_val
        total_dibayar += jml_bayar
        total_tagihan += ttl
        
        status_label = dict(Faktur._meta.get_field('status').choices).get(f.status, f.status)
        stored_name = (f.nama_pembiayaan or '').strip()
        is_unknown = stored_name.lower() in ('', 'unknown', '-')

        pembiayaan_name = (
            f.pelanggan.nama
            if f.pelanggan
            else pembiayaan_map.get(str(f.id_pembiayaan or ''), stored_name or '-')
            if is_unknown
            else stored_name
        )
        
        html_parts.append(
            f'<tr>'
            f'<td class="text-center">{no}</td>'
            f'<td class="text-center mono">{escape(f.nomor_faktur or "")}</td>'
            f'<td class="text-center">{(f.tanggal.strftime("%d-%m-%Y") if f.tanggal else "-")}</td>'
            f'<td>{escape(pembiayaan_name)}</td>'
            f'<td class="text-right">{float(f.adm or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.jasa or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.farmasi or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.tindakan or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.fisio or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.lab or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.rad or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.kamar or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.bhp or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.ambulan or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.alat or 0):,.2f}</td>'
            f'<td class="text-right">{float(f.lainnya or 0):,.2f}</td>'
            f'<td class="text-right">{float(total_biaya):,.2f}</td>'
            f'<td class="text-right">{float(dp3_val):,.2f}</td>'
            f'<td class="text-right">{float(dibayar_pasien_val):,.2f}</td>'
            f'<td class="text-right">{float(jml_bayar):,.2f}</td>'
            f'<td class="text-right">{float(ttl):,.2f}</td>'
            f'<td class="text-center">{(f.jatuh_tempo.strftime("%d-%m-%Y") if f.jatuh_tempo else "-")}</td>'
            f'<td class="text-center">{(f.tgl_kirim.strftime("%d-%m-%Y") if f.tgl_kirim else "-")}</td>'
            f'<td class="text-center">{escape(status_label)}</td>'
            f'</tr>'
        )
        no += 1
    
    # Total row
    html_parts.append(
        f'<tr style="font-weight: bold; background-color: #f0f0f0;">'
        f'<td colspan="16" class="text-right">TOTAL</td>'
        f'<td class="text-right">{float(total_pendapatan):,.2f}</td>'
        f'<td class="text-right">{float(total_dp3):,.2f}</td>'
        f'<td class="text-right">{float(total_dibayar_pasien):,.2f}</td>'
        f'<td class="text-right">{float(total_dibayar):,.2f}</td>'
        f'<td class="text-right">{float(total_tagihan):,.2f}</td>'
        f'<td colspan="3"></td>'
        f'</tr>'
    )
    
    html_parts.extend([
        '</tbody>',
        '</table>',
        '</div>',
        '</div>',
        '</body>',
        '</html>',
    ])
    
    html = '\n'.join(html_parts)
    return HttpResponse(html)

def faktur_rekap_excel_view(request):
    dari = request.GET.get('dari', '')
    sampai = request.GET.get('sampai', '')
    id_pembiayaan = request.GET.get('id_pembiayaan', '').strip()

    if not dari or not sampai:
        return HttpResponse('Parameter tanggal dari/sampai wajib diisi.', status=400)

    fakturs_qs = (
        Faktur.objects
        .filter(tanggal__gte=dari, tanggal__lte=sampai)
        .exclude(status='batal')
    )

    if id_pembiayaan:
        fakturs_qs = fakturs_qs.filter(Q(id_pembiayaan=str(id_pembiayaan)) | Q(id_pembiayaan=id_pembiayaan))

    fakturs = (
        fakturs_qs
        .select_related('pelanggan')
        .prefetch_related('pembayaran')
        .order_by('id_pembiayaan', 'tanggal')
    )
    
    pembiayaan_map = build_pembiayaan_name_map(
        fakturs.values_list('id_pembiayaan', flat=True).distinct()
    )

    # Query dp3 & jmlbyr per no_invoice
    nomor_fakturs = [f.nomor_faktur for f in fakturs if f.nomor_faktur]
    kunjung_map = {}
    if nomor_fakturs:
        try:
            with connection.cursor() as cursor:
                placeholders = ','.join(['%s'] * len(nomor_fakturs))
                cursor.execute(f"""
                    SELECT 
                        c.no_invoice,
                        COALESCE(SUM(COALESCE(a.dp3, 0)), 0) AS total_dp3,
                        COALESCE(SUM(COALESCE(a.jmlbyr, 0)), 0) AS total_jmlbyr
                    FROM rssams.kunjung a
                    INNER JOIN rssams.verif_kunjung c ON a.no = c.no
                    WHERE c.no_invoice IN ({placeholders})
                    GROUP BY c.no_invoice
                """, nomor_fakturs)
                for r_row in cursor.fetchall():
                    kunjung_map[str(r_row[0])] = {
                        'dp3': Decimal(str(r_row[1] or 0)),
                        'jmlbyr': Decimal(str(r_row[2] or 0)),
                    }
        except Exception:
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Invoice"

    # Template kolom pembayaran dibatasi tepat sampai 4 pasang (Bayar 1 s/d Bayar 4)
    max_pay = 4

    headers = [
        'NO', 'NO INVOICE', 'TANGGAL FAKTUR', 'PENANGGUNG',
        'ADM', 'JASA', 'FARMASI', 'TINDAKAN', 'FISIO', 'LAB',
        'RAD', 'KAMAR', 'BHP', 'AMBULAN', 'SEWA ALAT', 'LAIN2',
        'TOTAL PENDAPATAN', 'PIUTANG P3', 'DIBAYAR PASIEN', 'JML BAYAR', 'SISA TAGIHAN', 'TGL J.TEMPO', 'TGL KIRIM', 'STATUS'
    ]

    for i in range(1, max_pay + 1):
        headers.append(f'TGL BAYAR {i}')
        headers.append(f'JML BAYAR {i}')

    end_col_letter = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{end_col_letter}1')
    ws['A1'] = 'REKAP INVOICE'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1E293B')
    
    ws['A2'] = f'Tanggal : {dari} s/d {sampai}'
    ws['A2'].font = Font(name='Calibri', size=11, italic=True, color='64748B')

    ws.append([])
    ws.append(headers)

    # ── Styling Definitions ───────────────────────────────────
    thin_black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    total_black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    # Warna Header Kelompok Kolom
    fill_info = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')      # Dark Slate
    fill_costs = PatternFill(start_color='334155', end_color='334155', fill_type='solid')     # Slate
    fill_tot_pend = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')  # Blue
    fill_dp3 = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')       # Amber
    fill_pasien = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')    # Purple
    fill_tot_byr = PatternFill(start_color='047857', end_color='047857', fill_type='solid')   # Emerald
    fill_tot_tag = PatternFill(start_color='4338CA', end_color='4338CA', fill_type='solid')   # Indigo
    fill_meta = PatternFill(start_color='475569', end_color='475569', fill_type='solid')      # Dark Gray
    fill_pay_tgl_hdr = PatternFill(start_color='0E7490', end_color='0E7490', fill_type='solid') # Dark Cyan Hdr
    fill_pay_jml_hdr = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid') # Dark Teal Hdr

    # Warna Sel Data Pembayaran (Persis seperti screenshot contoh)
    fill_tgl_bayar = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # Kuning
    fill_jml_bayar = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid') # Biru Muda

    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    # Apply Header Styles (Row 4)
    for col_idx, cell in enumerate(ws[4], start=1):
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_black_border

        if 1 <= col_idx <= 4:
            cell.fill = fill_info
        elif 5 <= col_idx <= 16:
            cell.fill = fill_costs
        elif col_idx == 17:
            cell.fill = fill_tot_pend
        elif col_idx == 18:
            cell.fill = fill_dp3
        elif col_idx == 19:
            cell.fill = fill_pasien
        elif col_idx == 20:
            cell.fill = fill_tot_byr
        elif col_idx == 21:
            cell.fill = fill_tot_tag
        elif 22 <= col_idx <= 24:
            cell.fill = fill_meta
        else: # Column >= 25 (TGL BAYAR / JML BAYAR pairs)
            cell.fill = fill_pay_tgl_hdr if (col_idx - 24) % 2 != 0 else fill_pay_jml_hdr

    ws.row_dimensions[4].height = 28

    total_pendapatan = Decimal('0.00')
    total_dp3 = Decimal('0.00')
    total_dibayar_pasien = Decimal('0.00')
    total_dibayar = Decimal('0.00')
    total_tagihan = Decimal('0.00')

    for idx, f in enumerate(fakturs, start=1):
        total_biaya = (
            Decimal(f.adm or 0) +
            Decimal(f.jasa or 0) +
            Decimal(f.farmasi or 0) +
            Decimal(f.tindakan or 0) +
            Decimal(f.fisio or 0) +
            Decimal(f.lab or 0) +
            Decimal(f.rad or 0) +
            Decimal(f.kamar or 0) +
            Decimal(f.bhp or 0) +
            Decimal(f.lainnya or 0) +
            Decimal(f.ambulan or 0) +
            Decimal(f.alat or 0)
        )

        k_info = kunjung_map.get(str(f.nomor_faktur or ''), {'dp3': Decimal('0'), 'jmlbyr': Decimal('0')})
        dp3_val = k_info['dp3']
        dibayar_pasien_val = k_info['jmlbyr']

        jml_bayar = Decimal(f.total_dibayar or 0)
        ttl = total_biaya - jml_bayar
        
        total_pendapatan += total_biaya
        total_dp3 += dp3_val
        total_dibayar_pasien += dibayar_pasien_val
        total_dibayar += jml_bayar
        total_tagihan += ttl

        status_label = dict(Faktur._meta.get_field('status').choices).get(f.status, f.status)
        stored_name = (f.nama_pembiayaan or '').strip()
        is_unknown = stored_name.lower() in ('', 'unknown', '-')

        penanggung = (
            f.pelanggan.nama
            if f.pelanggan
            else pembiayaan_map.get(str(f.id_pembiayaan or ''), stored_name or '-')
            if is_unknown
            else stored_name
        )

        row_data = [
            idx,
            f.nomor_faktur or '',
            f.tanggal.strftime('%d-%m-%Y') if f.tanggal else '',
            penanggung,
            float(f.adm or 0),
            float(f.jasa or 0),
            float(f.farmasi or 0),
            float(f.tindakan or 0),
            float(f.fisio or 0),
            float(f.lab or 0),
            float(f.rad or 0),
            float(f.kamar or 0),
            float(f.bhp or 0),
            float(f.ambulan or 0),
            float(f.alat or 0),
            float(f.lainnya or 0),
            float(total_biaya),
            float(dp3_val),
            float(dibayar_pasien_val),
            float(jml_bayar),
            float(ttl),
            f.jatuh_tempo.strftime('%d-%m-%Y') if f.jatuh_tempo else '',
            f.tgl_kirim.strftime('%d-%m-%Y') if f.tgl_kirim else '',
            status_label,
        ]

        pembayaran_list = list(f.pembayaran.all())
        for i in range(max_pay):
            if i < len(pembayaran_list):
                p = pembayaran_list[i]
                tgl_p = p.tanggal.strftime('%d-%m-%Y') if p.tanggal else ''
                jml_p = float(p.jumlah or 0)
            else:
                tgl_p = ''
                jml_p = 0.0
            row_data.append(tgl_p)
            row_data.append(jml_p)

        ws.append(row_data)
        curr_row = ws.max_row

        # Apply borders, alignment, and fills to data row
        for c_idx, cell in enumerate(ws[curr_row], start=1):
            cell.border = thin_black_border

            if c_idx > 24:
                # Kolom pembayaran: Kuning untuk TGL BAYAR N, Biru Muda untuk JML BAYAR N
                if (c_idx - 24) % 2 != 0:
                    cell.fill = fill_tgl_bayar
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.fill = fill_jml_bayar
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                # Alignment kolom utama 1-24
                if c_idx in (1, 2, 3, 22, 23, 24):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx == 4:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')

    total_row = ws.max_row + 1

    ws.cell(row=total_row, column=1, value='TOTAL')
    ws.merge_cells(
        start_row=total_row,
        start_column=1,
        end_row=total_row,
        end_column=16
    )

    ws.cell(row=total_row, column=17, value=float(total_pendapatan))
    ws.cell(row=total_row, column=18, value=float(total_dp3))
    ws.cell(row=total_row, column=19, value=float(total_dibayar_pasien))
    ws.cell(row=total_row, column=20, value=float(total_dibayar))
    ws.cell(row=total_row, column=21, value=float(total_tagihan))

    # Total per kolom JML BAYAR N
    for i in range(max_pay):
        col_idx = 24 + (i * 2) + 2
        sum_pay_i = sum(
            float(list(f.pembayaran.all())[i].jumlah or 0)
            for f in fakturs if len(f.pembayaran.all()) > i
        )
        ws.cell(row=total_row, column=col_idx, value=sum_pay_i)

    fill_total = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    font_total = Font(name='Calibri', size=11, bold=True, color='0F172A')

    for c_idx, cell in enumerate(ws[total_row], start=1):
        cell.border = total_black_border
        cell.font = font_total
        if c_idx > 24:
            if (c_idx - 24) % 2 != 0:
                cell.fill = fill_tgl_bayar
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.fill = fill_jml_bayar
                cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.fill = fill_total
            if c_idx <= 16:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # Apply currency format
    for r in range(5, ws.max_row + 1):
        for col_idx in range(5, len(headers) + 1):
            if (5 <= col_idx <= 21) or (col_idx > 24 and (col_idx - 24) % 2 == 0):
                ws.cell(row=r, column=col_idx).number_format = '#,##0.00'

    # Auto-adjust column widths so numbers/dates never display as '####'
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in (1, 2, 3):
                continue
            val = cell.value
            if val is not None:
                if isinstance(val, (int, float, Decimal)):
                    val_str = f"{val:,.2f}"
                else:
                    val_str = str(val)
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Rekap_Invoice_{dari}_sampai_{sampai}.xlsx"'
    )

    wb.save(response)
    return response

def _utang_order_clause(value, allowed):
    return allowed.get(value) or next(iter(allowed.values()))

def _build_pending_where(params):
    """WHERE builder untuk tabel farmasi (tran_beli_brg_farmasi)."""
    where = [
        'u.id IS NULL',
        "(t.id NOT IN (SELECT nomor_spb FROM keuangan_utang_supplier WHERE nomor_spb != ''))",
        "(t.no_faktur IS NULL OR t.no_faktur = '' OR t.no_faktur NOT IN (SELECT nomor_faktur FROM keuangan_utang_supplier WHERE nomor_faktur != ''))"
    ]
    values = []
    search = (params.get('search') or '').strip()
    vendor_id = (params.get('vendor_id') or '').strip()
    kategori = (params.get('kategori') or '').strip()
    dari = (params.get('dari') or '').strip()
    sampai = (params.get('sampai') or '').strip()

    if search:
        where.append('(t.no_spb LIKE %s OR t.no_faktur LIKE %s OR r.nama LIKE %s OR t.id LIKE %s)')
        needle = f'%{search}%'
        values.extend([needle, needle, needle, needle])
    if vendor_id:
        where.append('t.id_rekanan = %s')
        values.append(vendor_id)
    if dari:
        where.append('t.tgl_faktur >= %s')
        values.append(dari)
    if sampai:
        where.append('t.tgl_faktur <= %s')
        values.append(sampai)
    return ' AND '.join(where), values

def _build_pending_where_logistik(params):
    """WHERE builder untuk tabel logistik (tran_beli_brg_log)."""
    where = [
        't.done = \'Y\'',
        'u.id IS NULL',
        "COALESCE(t.rekanan, '') != 'STOCK OPNAME'",
        "COALESCE(t.no_spk, '') NOT LIKE 'OPNAME-%%'",
        "(t.id NOT IN (SELECT nomor_spb FROM keuangan_utang_supplier WHERE nomor_spb != ''))",
        "(s.no_spb IS NULL OR s.no_spb = '' OR s.no_spb NOT IN (SELECT nomor_spb FROM keuangan_utang_supplier WHERE nomor_spb != ''))",
        "(t.no_spk IS NULL OR t.no_spk = '' OR t.no_spk NOT IN (SELECT nomor_faktur FROM keuangan_utang_supplier WHERE nomor_faktur != ''))"
    ]
    values = []
    search = (params.get('search') or '').strip()
    vendor_id = (params.get('vendor_id') or '').strip()
    dari = (params.get('dari') or '').strip()
    sampai = (params.get('sampai') or '').strip()

    if search:
        where.append('(t.id LIKE %s OR t.no_spk LIKE %s OR t.rekanan LIKE %s OR r.nama LIKE %s)')
        needle = f'%{search}%'
        values.extend([needle, needle, needle, needle])
    if vendor_id:
        where.append('r.id_rekanan = %s')
        values.append(vendor_id)
    if dari:
        where.append('t.tgl_spk >= %s')
        values.append(dari)
    if sampai:
        where.append('t.tgl_spk <= %s')
        values.append(sampai)
    return ' AND '.join(where), values

def _build_pending_where_keuangan(params):
    """WHERE builder untuk pengajuan penambahan saldo / pengisian kembali (petty_cash_pengajuan_saldo)."""
    where = [
        "t.status = 'disetujui'",
        'u.id IS NULL',
        "(t.no_pengajuan COLLATE utf8mb4_general_ci NOT IN (SELECT nomor_spb FROM keuangan_utang_supplier WHERE nomor_spb != ''))",
        "(t.no_pengajuan COLLATE utf8mb4_general_ci NOT IN (SELECT nomor_faktur FROM keuangan_utang_supplier WHERE nomor_faktur != ''))"
    ]
    values = []
    search = (params.get('search') or '').strip()
    dari = (params.get('dari') or '').strip()
    sampai = (params.get('sampai') or '').strip()

    if search:
        where.append('(t.no_pengajuan LIKE %s OR t.alasan LIKE %s OR usr.first_name LIKE %s OR usr.username LIKE %s)')
        needle = f'%{search}%'
        values.extend([needle, needle, needle, needle])
    if dari:
        where.append('t.tanggal >= %s')
        values.append(dari)
    if sampai:
        where.append('t.tanggal <= %s')
        values.append(sampai)
    return ' AND '.join(where), values

def _pending_base_sql():
    """FROM clause untuk farmasi — JOIN ke utang_supplier by app_siaga_faktur_id."""
    return """
        FROM rssams.tran_beli_brg_farmasi t
        LEFT JOIN rssams.rekanan r ON r.id_rekanan = t.id_rekanan
        LEFT JOIN keuangan_utang_supplier u ON u.app_siaga_faktur_id = t.id
    """

def _pending_base_sql_logistik():
    """FROM clause untuk logistik — JOIN ke utang_supplier by app_siaga_faktur_id."""
    return """
        FROM rssams.tran_beli_brg_log t
        LEFT JOIN rssams.logistik_spb s ON s.id = t.id_spb
        LEFT JOIN rssams.rekanan r ON UPPER(TRIM(r.nama)) = UPPER(TRIM(t.rekanan)) AND r.del = 'N'
        LEFT JOIN keuangan_utang_supplier u ON u.app_siaga_faktur_id = CONCAT('LOG-', t.id)
    """

def _pending_base_sql_keuangan():
    """FROM clause untuk keuangan (pengisian kembali saldo) — JOIN ke utang_supplier."""
    return """
        FROM petty_cash_pengajuan_saldo t
        LEFT JOIN users_user usr ON usr.id = t.created_by_id
        LEFT JOIN keuangan_utang_supplier u ON (u.app_siaga_faktur_id = CONCAT('KEU-', t.id) COLLATE utf8mb4_general_ci OR u.nomor_spb = t.no_pengajuan COLLATE utf8mb4_general_ci)
    """

def _fetch_app_siaga_faktur(app_siaga_faktur_id):
    """Fetch 1 baris faktur farmasi dari tran_beli_brg_farmasi."""
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                t.id AS app_siaga_faktur_id,
                t.id AS nomor_spb,
                t.tgl_spb AS tanggal_spb,
                t.no_faktur AS nomor_faktur,
                t.id_rekanan AS vendor_id,
                COALESCE(r.nama, '') AS vendor_nama,
                t.tgl_faktur AS tanggal_faktur,
                t.tgl_jtempo AS tanggal_jatuh_tempo,
                t.gtotal AS nominal
            FROM rssams.tran_beli_brg_farmasi t
            LEFT JOIN rssams.rekanan r ON r.id_rekanan = t.id_rekanan
            WHERE t.id = %s
            LIMIT 1
            """,
            [app_siaga_faktur_id],
        )
        row = cursor.fetchone()
        if not row:
            return None
        columns = [col[0] for col in cursor.description]
        return dict(zip(columns, row))

def _fetch_logistik_pembelian(pembelian_id):
    """Fetch 1 baris dari tran_beli_brg_log + best-effort JOIN rekanan by-nama."""
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                CAST(t.id AS CHAR) AS app_siaga_faktur_id,
                CAST(COALESCE(NULLIF(s.no_spb, ''), NULLIF(t.id_spb, ''), t.id) AS CHAR) AS nomor_spb,
                t.tgl_spk AS tanggal_spb,
                CAST(COALESCE(NULLIF(NULLIF(t.no_spk, ''), '-'), '-') AS CHAR) AS nomor_faktur,
                r.id_rekanan AS vendor_id_hint,
                t.rekanan AS rekanan_text,
                COALESCE(r.nama, t.rekanan) AS vendor_nama_hint,
                t.tgl_spk AS tanggal_faktur,
                NULL AS tanggal_jatuh_tempo,
                t.nilai AS nominal,
                t.metode_pembayaran
            FROM rssams.tran_beli_brg_log t
            LEFT JOIN rssams.logistik_spb s ON s.id = t.id_spb
            LEFT JOIN rssams.rekanan r ON UPPER(TRIM(CONVERT(r.nama USING utf8mb4))) = UPPER(TRIM(CONVERT(t.rekanan USING utf8mb4))) AND r.del = 'N'
            WHERE t.id = %s AND t.done = 'Y'
            LIMIT 1
            """,
            [pembelian_id],
        )
        row = cursor.fetchone()
        if not row:
            return None
        columns = [col[0] for col in cursor.description]
        return dict(zip(columns, row))

def parse_lenient_date(val):
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    val_s = str(val).strip()
    if not val_s:
        return None
    # Try different separators
    parts = None
    for sep in ('/', '-', '.'):
        if sep in val_s:
            parts = val_s.split(sep)
            break
    if not parts or len(parts) < 3:
        return None
    try:
        # Check order: is it YYYY-MM-DD or DD/MM/YYYY?
        if len(parts[0]) == 4:
            year, month, day = int(parts[0]), int(parts[1]), int(parts[2][:2])
        elif len(parts[2][:4]) == 4:
            day, month, year = int(parts[0]), int(parts[1]), int(parts[2][:4])
        else:
            return None
        # Capping month
        month = max(1, min(12, month))
        # Capping day based on month and year
        if month in (4, 6, 9, 11):
            day = min(day, 30)
        elif month == 2:
            is_leap = year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)
            day = min(day, 29 if is_leap else 28)
        else:
            day = min(day, 31)
        day = max(1, day)
        return date(year, month, day)
    except Exception:
        return None

def _clean_vendor_name(name):
    if not name:
        return ''
    s = str(name).upper()
    s = re.sub(r'\b(PT|CV|PD|UD|TB|NV)\b', '', s)
    s = re.sub(r'[^A-Z0-9]', '', s)
    return s

class UtangSupplierViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    serializer_class = UtangSupplierSerializer

    def get_queryset(self):
        qs = UtangSupplier.objects.all()
        params = self.request.query_params

        search = (params.get('search') or '').strip()
        vendor_id = (params.get('vendor_id') or '').strip()
        kategori = (params.get('kategori') or '').strip()
        st = (params.get('status') or '').strip()
        sumber = (params.get('sumber') or '').strip()
        dari = (params.get('dari') or '').strip()
        sampai = (params.get('sampai') or '').strip()
        ordering = (params.get('ordering') or '').strip()

        if search:
            qs = qs.filter(
                models.Q(nomor_faktur__icontains=search) |
                models.Q(nomor_spb__icontains=search) |
                models.Q(vendor_nama__icontains=search) |
                models.Q(keterangan_titip__icontains=search) |
                models.Q(kategori__icontains=search) |
                models.Q(app_siaga_faktur_id__icontains=search)
            )

        if vendor_id:
            qs = qs.filter(vendor_id=vendor_id)

        if kategori:
            qs = qs.filter(kategori__icontains=kategori)

        if st == 'aktif':
            qs = qs.exclude(status__in=[UtangSupplier.STATUS_LUNAS, UtangSupplier.STATUS_DIBATALKAN])
        elif st and st not in ['semua', 'all']:
            qs = qs.filter(status=st)

        if sumber and sumber not in ['semua', 'all']:
            qs = qs.filter(sumber=sumber)

        if dari:
            qs = qs.filter(tanggal_titip__gte=dari)

        if sampai:
            qs = qs.filter(tanggal_titip__lte=sampai)

        if ordering:
            allowed_ordering = [
                'tanggal_titip', '-tanggal_titip',
                'tanggal_faktur', '-tanggal_faktur',
                'verified_at', '-verified_at',
                'vendor_nama', '-vendor_nama',
                'nominal', '-nominal',
                'status', '-status',
                'created_at', '-created_at'
            ]
            if ordering in allowed_ordering:
                qs = qs.order_by(ordering)
            else:
                qs = qs.order_by('-verified_at', '-created_at')
        else:
            qs = qs.order_by('-verified_at', '-created_at')

        return qs

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        # Validasi 0: Faktur yang sudah dibatalkan tidak bisa diedit
        if instance.status == UtangSupplier.STATUS_DIBATALKAN:
            return Response({'error': 'Faktur yang sudah dibatalkan tidak dapat diedit.'}, status=status.HTTP_400_BAD_REQUEST)

        # Validasi 1: Faktur lunas tidak bisa diedit
        if instance.status == UtangSupplier.STATUS_LUNAS:
            return Response({'error': 'Faktur yang sudah lunas tidak dapat diedit.'}, status=status.HTTP_400_BAD_REQUEST)

        # Validasi 2: Faktur yang sudah ada pembayaran / cicilan / retur tidak bisa diedit
        has_realisasi = instance.pembayaran.filter(
            status__in=[PembayaranUtang.STATUS_REALISASI_SEBAGIAN, PembayaranUtang.STATUS_REALISASI_LUNAS, PembayaranUtang.STATUS_RETUR]
        ).exists()
        if has_realisasi or (instance.total_dibayar and instance.total_dibayar > 0):
            return Response({'error': 'Faktur yang sudah memiliki riwayat pembayaran atau retur tidak dapat diedit.'}, status=status.HTTP_400_BAD_REQUEST)

        # Validasi 3: Faktur yang sedang dalam antrean pengajuan pembayaran tidak bisa diedit
        has_pending = instance.pembayaran.filter(status=PembayaranUtang.STATUS_PENDING).exists()
        if has_pending:
            return Response({'error': 'Faktur sedang dalam proses antrean pengajuan pembayaran. Batalkan pengajuan pembayaran terlebih dahulu untuk mengedit data.'}, status=status.HTTP_400_BAD_REQUEST)

        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)

        # Jika vendor_id diubah, sinkronkan nama vendor dari master rekanan
        vendor_id = data.get('vendor_id')
        if vendor_id:
            with connection.cursor() as cursor:
                cursor.execute("SELECT nama FROM rssams.rekanan WHERE id_rekanan = %s AND del = 'N' LIMIT 1", [vendor_id])
                row = cursor.fetchone()
                if row:
                    data['vendor_nama'] = row[0]

        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        instance.refresh_from_db()
        instance.refresh_status()

        return Response({
            'message': 'Catatan utang berhasil diperbarui.',
            'utang': UtangSupplierSerializer(instance, context={'request': request}).data
        }, status=status.HTTP_200_OK)

    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        qs = self.filter_queryset(self.get_queryset())
        active_qs = qs.exclude(status__in=[UtangSupplier.STATUS_LUNAS, UtangSupplier.STATUS_DIBATALKAN])

        total_nominal = active_qs.aggregate(total=models.Sum('nominal'))['total'] or Decimal('0')
        total_dibayar = PembayaranUtang.objects.filter(
            utang__in=active_qs,
            status__in=[PembayaranUtang.STATUS_REALISASI_SEBAGIAN, PembayaranUtang.STATUS_REALISASI_LUNAS, PembayaranUtang.STATUS_RETUR]
        ).aggregate(total=models.Sum('jumlah_bayar'))['total'] or Decimal('0')
        
        total_sisa = max(Decimal('0'), total_nominal - total_dibayar)

        total_faktur = qs.count()
        total_lunas_count = qs.filter(status=UtangSupplier.STATUS_LUNAS).count()
        total_dibatalkan_count = qs.filter(status=UtangSupplier.STATUS_DIBATALKAN).count()
        total_aktif_count = active_qs.count()

        return Response({
            'total_faktur': total_faktur,
            'total_lunas_count': total_lunas_count,
            'total_dibatalkan_count': total_dibatalkan_count,
            'total_aktif_count': total_aktif_count,
            'utang_count': total_aktif_count,
            'total_nominal': total_nominal,
            'total_dibayar': total_dibayar,
            'total_sisa': total_sisa,
        })

    @action(detail=True, methods=['post'], url_path='batalkan')
    def batalkan(self, request, pk=None):
        """Membatalkan catatan utang (agar track audit tetap ada tanpa menghapus data)."""
        instance = self.get_object()

        if instance.status == UtangSupplier.STATUS_DIBATALKAN:
            return Response({'error': 'Faktur ini sudah dibatalkan sebelumnya.'}, status=status.HTTP_400_BAD_REQUEST)

        # Cek apakah sudah ada pembayaran realisasi
        has_realisasi = instance.pembayaran.filter(
            status__in=[PembayaranUtang.STATUS_REALISASI_SEBAGIAN, PembayaranUtang.STATUS_REALISASI_LUNAS, PembayaranUtang.STATUS_RETUR]
        ).exists()
        if has_realisasi or (instance.total_dibayar and instance.total_dibayar > 0):
            return Response({'error': 'Faktur yang sudah memiliki riwayat pembayaran terealisasi tidak dapat dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)

        alasan = (request.data.get('alasan') or request.data.get('alasan_batal') or '').strip()
        if not alasan:
            return Response({'error': 'Alasan pembatalan wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            # Batalkan otomatis jika ada pembayaran yang masih berstatus pending
            instance.pembayaran.filter(status=PembayaranUtang.STATUS_PENDING).update(
                status=PembayaranUtang.STATUS_DITOLAK
            )

            instance.status = UtangSupplier.STATUS_DIBATALKAN
            instance.alasan_batal = alasan
            instance.dibatalkan_by = request.user
            instance.dibatalkan_at = timezone.now()
            instance.save(update_fields=['status', 'alasan_batal', 'dibatalkan_by', 'dibatalkan_at', 'updated_at'])

        return Response({
            'message': f'Faktur {instance.nomor_faktur} berhasil dibatalkan.',
            'utang': UtangSupplierSerializer(instance, context={'request': request}).data
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='create-manual')
    def create_manual(self, request):
        """Membuat catatan utang secara manual (tidak dari database legacy)."""
        data = request.data
        vendor_id = data.get('vendor_id')
        if not vendor_id:
            return Response({'error': 'vendor_id wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        # Ambil nama vendor dari rssams.rekanan
        vendor_nama = ''
        with connection.cursor() as cursor:
            cursor.execute("SELECT nama FROM rssams.rekanan WHERE id_rekanan = %s AND del = 'N'", [vendor_id])
            row = cursor.fetchone()
            if row:
                vendor_nama = row[0]

        if not vendor_nama:
            return Response({'error': 'Vendor tidak ditemukan di master rekanan.'}, status=status.HTTP_404_NOT_FOUND)

        nomor_faktur = (data.get('nomor_faktur') or '').strip()
        if not nomor_faktur:
            return Response({'error': 'nomor_faktur wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        # Cek apakah nomor faktur + vendor sudah pernah diinput (dan belum dibatalkan)
        duplicate = UtangSupplier.objects.filter(
            vendor_id=int(vendor_id),
            nomor_faktur__iexact=nomor_faktur
        ).exclude(status=UtangSupplier.STATUS_DIBATALKAN).first()

        if duplicate:
            nominal_str = f"{float(duplicate.nominal):,.0f}".replace(',', '.')
            return Response({
                'error': f"Faktur dengan nomor '{nomor_faktur}' untuk vendor '{vendor_nama}' sudah tercatat di sistem (ID #{duplicate.id}, Nominal: Rp {nominal_str}, Status: {duplicate.get_status_display()}). Jika faktur ini merupakan koreksi/perbaikan, silakan batalkan faktur sebelumnya terlebih dahulu."
            }, status=status.HTTP_400_BAD_REQUEST)

        tgl_faktur = data.get('tanggal_faktur')
        if not tgl_faktur:
            return Response({'error': 'tanggal_faktur wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        keterangan = (data.get('keterangan') or '').strip()
        if not keterangan:
            return Response({'error': 'keterangan wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        nominal_raw = data.get('nominal')
        try:
            nominal = Decimal(str(nominal_raw))
            if nominal <= 0:
                raise ValueError
        except (TypeError, ValueError, InvalidOperation):
            return Response({'error': 'nominal tidak valid atau wajib lebih dari 0.'}, status=status.HTTP_400_BAD_REQUEST)

        import uuid
        faktur_id = f'MNL-{uuid.uuid4().hex[:12].upper()}'

        utang = UtangSupplier.objects.create(
            app_siaga_faktur_id=faktur_id,
            sumber=UtangSupplier.SUMBER_MANUAL,
            nomor_faktur=nomor_faktur,
            nomor_spb=(data.get('nomor_spb') or '').strip(),
            vendor_id=int(vendor_id),
            vendor_nama=vendor_nama,
            tanggal_faktur=data.get('tanggal_faktur') or None,
            tanggal_jatuh_tempo=data.get('tanggal_jatuh_tempo') or None,
            tanggal_titip=data.get('tanggal_titip') or timezone.localdate(),
            nominal=nominal,
            keterangan_titip=keterangan,
            status=UtangSupplier.STATUS_BELUM_DIBAYAR,
            verified_by=request.user,
            verified_at=timezone.now(),
        )

        return Response({
            'message': 'Catatan utang manual berhasil dibuat.',
            'utang': UtangSupplierSerializer(utang, context={'request': request}).data,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='vendor-deposit')
    def vendor_deposit(self, request):
        vendor_id = request.query_params.get('vendor_id')
        if not vendor_id:
            return Response({'error': 'vendor_id wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)
        
        deposits = DepositVendor.objects.filter(vendor_id=vendor_id).order_by('created_at')
        active_deposits = [d for d in deposits if d.sisa_deposit > 0]
        total_sisa = sum((d.sisa_deposit for d in active_deposits), Decimal('0'))

        return Response({
            'vendor_id': int(vendor_id),
            'total_sisa_deposit': float(total_sisa),
            'deposits': DepositVendorSerializer(active_deposits, many=True).data,
        })

    @action(detail=False, methods=['get'], url_path='list-deposit-vendor')
    def list_deposit_vendor(self, request):
        qs = DepositVendor.objects.all().order_by('-created_at')
        
        vendor_id = request.query_params.get('vendor_id')
        search = (request.query_params.get('search') or '').strip()
        status_filter = request.query_params.get('status') or 'aktif'

        if vendor_id:
            qs = qs.filter(vendor_id=vendor_id)
        if search:
            qs = qs.filter(
                models.Q(vendor_nama__icontains=search) |
                models.Q(keterangan__icontains=search) |
                models.Q(utang_asal__nomor_faktur__icontains=search) |
                models.Q(utang_asal__nomor_spb__icontains=search) |
                models.Q(utang_asal__keterangan_titip__icontains=search)
            )

        deposits_data = DepositVendorSerializer(qs, many=True).data

        vendor_map = {}
        for dep_obj, dep_data in zip(qs, deposits_data):
            vid = dep_obj.vendor_id
            if vid not in vendor_map:
                vendor_map[vid] = {
                    'vendor_id': vid,
                    'vendor_nama': dep_obj.vendor_nama,
                    'total_retur': Decimal('0'),
                    'total_terpakai': Decimal('0'),
                    'total_sisa_deposit': Decimal('0'),
                    'count': 0,
                    'items': [],
                }
            vendor_map[vid]['total_retur'] += dep_obj.nominal_retur
            vendor_map[vid]['total_terpakai'] += dep_obj.terpakai
            vendor_map[vid]['total_sisa_deposit'] += dep_obj.sisa_deposit
            vendor_map[vid]['count'] += 1
            vendor_map[vid]['items'].append(dep_data)

        vendor_list = list(vendor_map.values())
        if status_filter == 'aktif':
            vendor_list = [v for v in vendor_list if v['total_sisa_deposit'] > 0]
        elif status_filter == 'habis':
            vendor_list = [v for v in vendor_list if v['total_sisa_deposit'] <= 0]

        total_retur_all = sum((v['total_retur'] for v in vendor_list), Decimal('0'))
        total_terpakai_all = sum((v['total_terpakai'] for v in vendor_list), Decimal('0'))
        total_sisa_all = sum((v['total_sisa_deposit'] for v in vendor_list), Decimal('0'))

        return Response({
            'summary': {
                'total_vendor': len(vendor_list),
                'total_retur': float(total_retur_all),
                'total_terpakai': float(total_terpakai_all),
                'total_sisa_deposit': float(total_sisa_all),
            },
            'vendors': vendor_list,
        })

    @action(detail=True, methods=['post'], url_path='input-retur')
    def input_retur(self, request, pk=None):
        utang = self.get_object()
        if utang.status in [UtangSupplier.STATUS_DIAJUKAN, UtangSupplier.STATUS_SEBAGIAN_DIAJUKAN]:
            return Response({'error': 'Faktur ini sedang dalam proses pengajuan pembayaran. Batalkan pengajuan terlebih dahulu untuk mencatat retur.'}, status=status.HTTP_400_BAD_REQUEST)

        has_realisasi = utang.pembayaran.filter(status__in=['realisasi_sebagian', 'realisasi_lunas']).exists()
        if utang.status not in [UtangSupplier.STATUS_BELUM_DIBAYAR] or has_realisasi:
            return Response({'error': 'Fitur retur saat ini hanya berlaku untuk faktur yang belum pernah dibayar sama sekali.'}, status=status.HTTP_400_BAD_REQUEST)

        nominal_retur_raw = request.data.get('nominal_retur')
        keterangan = (request.data.get('keterangan') or '').strip()

        try:
            nominal_retur = Decimal(str(nominal_retur_raw))
            if nominal_retur <= 0:
                raise ValueError
        except (TypeError, ValueError, InvalidOperation):
            return Response({'error': 'Nominal retur tidak valid atau harus lebih dari 0.'}, status=status.HTTP_400_BAD_REQUEST)

        if not keterangan:
            return Response({'error': 'Keterangan / Nomor Nota Retur wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        sisa_utang = utang.sisa_utang
        if sisa_utang < nominal_retur:
            return Response({'error': f'Nominal retur (Rp {nominal_retur:,.2f}) tidak boleh melebihi sisa utang saat ini (Rp {sisa_utang:,.2f}).'}, status=status.HTTP_400_BAD_REQUEST)

        today = timezone.localdate()
        pembayaran = PembayaranUtang.objects.create(
            utang=utang,
            tanggal_rencana_bayar=today,
            tanggal_proses=today,
            tanggal_app=today,
            jumlah_bayar=nominal_retur,
            potongan_deposit=Decimal('0'),
            jumlah_kas_keluar=Decimal('0'),
            keterangan=f"Retur Barang: {keterangan}",
            status=PembayaranUtang.STATUS_RETUR,
            created_by=request.user,
        )

        utang.refresh_status()
        utang.refresh_from_db()

        return Response({
            'message': f'Retur sebesar Rp {nominal_retur:,.2f} berhasil dicatat pada faktur {utang.nomor_faktur or utang.nomor_spb}. Sisa utang kini menjadi Rp {utang.sisa_utang:,.2f}.',
            'utang': UtangSupplierSerializer(utang, context={'request': request}).data,
            'pembayaran': PembayaranUtangSerializer(pembayaran, context={'request': request}).data,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='bayar')
    def bayar(self, request, pk=None):
        utang = self.get_object()
        if utang.status == UtangSupplier.STATUS_LUNAS:
            return Response({'error': 'Utang sudah lunas.'}, status=status.HTTP_400_BAD_REQUEST)
        if utang.status in [UtangSupplier.STATUS_DIAJUKAN, UtangSupplier.STATUS_SEBAGIAN_DIAJUKAN]:
            return Response({'error': 'Faktur ini sedang dalam proses pengajuan pembayaran.'}, status=status.HTTP_400_BAD_REQUEST)
        
        tgl_rencana = request.data.get('tanggal_rencana_bayar') or timezone.now().date().isoformat()

        potongan_deposit_raw = request.data.get('potongan_deposit') or 0
        try:
            potongan_deposit = Decimal(str(potongan_deposit_raw))
            if potongan_deposit < 0:
                raise ValueError
        except (TypeError, ValueError, InvalidOperation):
            return Response({'error': 'Nilai potongan deposit tidak valid.'}, status=status.HTTP_400_BAD_REQUEST)

        if potongan_deposit > 0:
            deposits = DepositVendor.objects.filter(vendor_id=utang.vendor_id)
            total_sisa_deposit = sum((d.sisa_deposit for d in deposits), Decimal('0'))
            if potongan_deposit > total_sisa_deposit:
                return Response({'error': f'Potongan deposit (Rp {potongan_deposit:,.2f}) melebihi saldo deposit vendor yang tersedia (Rp {total_sisa_deposit:,.2f}).'}, status=status.HTTP_400_BAD_REQUEST)

        payload = {
            **request.data,
            'utang': utang.id,
            'tanggal_rencana_bayar': tgl_rencana,
            'tanggal_proses': request.data.get('tanggal_proses') or tgl_rencana,
            'tanggal_app': request.data.get('tanggal_app') or tgl_rencana,
            'potongan_deposit': potongan_deposit,
        }
        serializer = PembayaranUtangInputSerializer(data=payload)
        serializer.is_valid(raise_exception=True)
        
        tgl_proses = serializer.validated_data.get('tanggal_proses') or tgl_rencana
        jumlah_bayar = serializer.validated_data['jumlah_bayar']
        jumlah_kas_keluar = max(jumlah_bayar - potongan_deposit, Decimal('0'))

        pembayaran = PembayaranUtang.objects.create(
            utang=utang,
            tanggal_rencana_bayar=serializer.validated_data.get('tanggal_rencana_bayar'),
            tanggal_proses=tgl_proses,
            tanggal_app=serializer.validated_data.get('tanggal_app') or tgl_rencana,
            jumlah_bayar=jumlah_bayar,
            potongan_deposit=potongan_deposit,
            jumlah_kas_keluar=jumlah_kas_keluar,
            keterangan=serializer.validated_data.get('keterangan', ''),
            status=PembayaranUtang.STATUS_PENDING,
            created_by=request.user,
        )
        utang.refresh_status()
        utang.refresh_from_db()
        return Response({
            'utang': UtangSupplierSerializer(utang, context={'request': request}).data,
            'pembayaran': PembayaranUtangSerializer(pembayaran, context={'request': request}).data,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='ots-preview', parser_classes=[MultiPartParser, FormParser])
    def ots_preview(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'error': 'File Excel wajib diunggah.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            if 'LIST FAKTUR' in wb.sheetnames:
                sheet = wb['LIST FAKTUR']
            else:
                sheet = wb.active

            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 3:
                return Response({'error': 'Sheet Excel tidak memiliki data yang cukup.'}, status=status.HTTP_400_BAD_REQUEST)

            green_hex = 'FF92D050'
            staged_items = []
            SUMBER_MANUAL = UtangSupplier.SUMBER_MANUAL
            SUMBER_FARMASI = UtangSupplier.SUMBER_FARMASI

            # Read Cell T1 (authoritative "Sisa Utang Aktif" from Excel subtotal formula)
            excel_t1_raw = sheet.cell(row=1, column=20).value  # Col T, Row 1
            try:
                excel_t1_value = Decimal(str(excel_t1_raw)) if excel_t1_raw else None
            except Exception:
                excel_t1_value = None

            seen_keys = {}
            seen_spb = {}

            # Fetch existing non-OTS SPBs and Fakturs from SIMAK Database with department scoping
            existing_db_spb_map = {}
            existing_db_faktur_map = {}
            for u in UtangSupplier.objects.exclude(app_siaga_faktur_id__startswith='OTS-'):
                st_display = u.get_status_display() or u.status
                s_label = "LOGISTIK" if u.sumber == SUMBER_MANUAL else "FARMASI"
                if u.nomor_spb:
                    existing_db_spb_map[(u.sumber, u.nomor_spb.strip().upper())] = (st_display, u.vendor_nama, s_label)
                if u.nomor_faktur:
                    existing_db_faktur_map[(u.sumber, u.nomor_faktur.strip().upper())] = (st_display, u.vendor_nama, s_label)

            total_rows = 0
            total_nominal = Decimal('0')
            total_bayar = Decimal('0')
            total_sisa_utang = Decimal('0')
            total_anomali = 0
            total_lunas = 0
            total_utang_aktif = 0

            for r_idx in range(2, len(rows)):
                r = rows[r_idx]
                if not r:
                    continue

                status_code = str(r[3] or '').strip()
                kategori = str(r[4] or '').strip()
                no_spb = str(r[7] or '').strip()
                vendor_nama = str(r[8] or '').strip()
                tgl_faktur_raw = r[9]
                tgl_titip_raw = r[10] if len(r) > 10 else None
                nominal_raw = r[11]
                ket_excel = str(r[12] or '').strip() # Col M (KETERANGAN / NO FAKTUR)
                no_faktur = ket_excel or f"INV/OTS/{r_idx + 1}"
                byr_raw = r[18] if len(r) > 18 else 0

                if not vendor_nama or nominal_raw is None:
                    continue

                try:
                    nominal = Decimal(str(nominal_raw))
                    if nominal <= Decimal('0'):
                        continue
                except (InvalidOperation, ValueError, TypeError):
                    continue

                try:
                    byr = Decimal(str(byr_raw)) if (byr_raw is not None and str(byr_raw).strip() != '') else Decimal('0')
                except (InvalidOperation, ValueError, TypeError):
                    byr = Decimal('0')

                sisa = max(Decimal('0'), nominal - byr)
                row_num = r_idx + 1

                # Determine department sumber
                kat_upper = kategori.upper()
                if any(kw in kat_upper for kw in ['ATK', 'RUMAH TANGGA', 'CETAKAN']):
                    sumber = SUMBER_MANUAL
                    sumber_label = "LOGISTIK"
                elif 'BHP' in kat_upper or 'OBAT' in kat_upper:
                    sumber = SUMBER_FARMASI
                    sumber_label = "FARMASI"
                else:
                    sumber = SUMBER_MANUAL
                    sumber_label = "LOGISTIK"

                # Helper date parser supporting datetime.date, datetime.datetime, and multiple string formats
                def parse_date_str(val):
                    d_obj = parse_lenient_date(val)
                    if d_obj:
                        return d_obj.strftime('%Y-%m-%d')
                    if not val:
                        return ''
                    val_s = str(val).strip()
                    return val_s[:10]

                tgl_faktur_str = (
                    parse_date_str(tgl_faktur_raw) or
                    parse_date_str(tgl_titip_raw) or
                    (parse_date_str(r[15]) if len(r) > 15 else '') or
                    (parse_date_str(r[16]) if len(r) > 16 else '') or
                    (parse_date_str(r[17]) if len(r) > 17 else '')
                )
                tgl_titip_str = parse_date_str(tgl_titip_raw) or (parse_date_str(r[15]) if len(r) > 15 else '') or tgl_faktur_str
                tgl_rencana_str = parse_date_str(r[15]) if len(r) > 15 else ''
                tgl_proses_str = parse_date_str(r[16]) if len(r) > 16 else ''
                tgl_app_str = parse_date_str(r[17]) if len(r) > 17 else ''

                # Check cell fill color safely
                is_green = False
                try:
                    cell = sheet.cell(row=row_num, column=1)
                    if cell and cell.fill and getattr(cell.fill, 'start_color', None):
                        st_c = cell.fill.start_color
                        c_val = getattr(st_c, 'rgb', None) or getattr(st_c, 'value', None)
                        if c_val and str(c_val).upper() in ['FF92D050', '92D050']:
                            is_green = True
                except Exception:
                    pass

                # Status determination: Prioritize explicit Excel Status Code 'U' (Utang Aktif) vs 'L' (Lunas)
                if status_code == 'U':
                    if byr >= nominal:
                        byr = Decimal('0')
                    sisa = max(Decimal('0'), nominal - byr)
                    status_ditentukan = 'sebagian' if byr > Decimal('0') else 'belum_dibayar'
                    total_utang_aktif += 1
                elif status_code == 'L':
                    byr = nominal
                    sisa = Decimal('0')
                    status_ditentukan = 'lunas'
                    total_lunas += 1
                else:
                    if sisa <= Decimal('0'):
                        status_ditentukan = 'lunas'
                        total_lunas += 1
                    elif byr > Decimal('0'):
                        status_ditentukan = 'sebagian'
                        total_utang_aktif += 1
                    else:
                        status_ditentukan = 'belum_dibayar'
                        total_utang_aktif += 1

                # Exact duplicate detection
                key_spb = (sumber, no_spb.upper()) if no_spb else None
                # Only use faktur-based key for Non-SPB items (manual entries like remunerasi).
                # Two different SPBs with the same supplier invoice number are legitimate separate transactions.
                key_faktur = (sumber, no_faktur.upper()) if (no_faktur and not no_spb) else None
                is_exact_duplicate = False
                anomali_reasons = []

                if key_spb and key_spb in seen_spb:
                    is_exact_duplicate = True
                    anomali_reasons.append(f"Duplikat persis dengan baris #{seen_spb[key_spb]}")
                elif key_faktur and key_faktur in seen_keys:
                    is_exact_duplicate = True
                    anomali_reasons.append(f"Duplikat persis dengan baris #{seen_keys[key_faktur]}")

                if key_spb and not is_exact_duplicate:
                    seen_spb[key_spb] = row_num
                if key_faktur and not is_exact_duplicate:
                    seen_keys[key_faktur] = row_num

                if key_spb and key_spb in existing_db_spb_map:
                    st_db, v_db, s_label = existing_db_spb_map[key_spb]
                    anomali_reasons.append(f"SPB '{no_spb}' sudah tercatat di DB SIMAK ({s_label} - Status: {st_db}, Vendor: {v_db})")
                elif key_faktur and key_faktur in existing_db_faktur_map:
                    st_db, v_db, s_label = existing_db_faktur_map[key_faktur]
                    anomali_reasons.append(f"Faktur '{no_faktur}' sudah tercatat di DB SIMAK ({s_label} - Status: {st_db}, Vendor: {v_db})")



                is_anomali = bool(anomali_reasons)
                if is_anomali:
                    total_anomali += 1

                total_rows += 1
                total_nominal += nominal
                total_bayar += byr
                total_sisa_utang += sisa

                staged_items.append({
                    'id': f"OTS-{row_num}",
                    'row_idx': row_num,
                    'status_excel': status_code,
                    'kategori': kategori,
                    'no_spb': no_spb,
                    'vendor_nama': vendor_nama,
                    'tgl_faktur': tgl_faktur_str,
                    'tgl_titip': tgl_titip_str,
                    'nominal': float(nominal),
                    'no_faktur': ket_excel or f"INV/OTS/{row_num}",
                    'keterangan_excel': ket_excel,
                    'jumlah_bayar': float(byr),
                    'sisa_utang': float(sisa),
                    'tgl_rencana_bayar': tgl_rencana_str,
                    'tgl_proses': tgl_proses_str,
                    'tgl_app': tgl_app_str,
                    'is_green': is_green,
                    'status_ditentukan': status_ditentukan,
                    'is_anomali': is_anomali,
                    'anomali_reasons': anomali_reasons,
                    'user_action': 'abaikan' if is_exact_duplicate else 'terima',
                })

            # Reconcile with Excel Cell T1: flag rows that cause discrepancy
            if excel_t1_value is not None:
                # Calculate our active debt total (sum of sisa for non-lunas items)
                our_aktif_total = sum(
                    Decimal(str(item['sisa_utang']))
                    for item in staged_items
                    if item['status_ditentukan'] != 'lunas'
                )
                diff = our_aktif_total - excel_t1_value
                if diff > Decimal('0'):
                    # Find rows whose sisa_utang individually or cumulatively match the discrepancy
                    remaining_diff = diff
                    for item in staged_items:
                        if remaining_diff <= Decimal('0'):
                            break
                        if item['status_ditentukan'] == 'lunas':
                            continue
                        item_sisa = Decimal(str(item['sisa_utang']))
                        if item_sisa > Decimal('0') and item_sisa == remaining_diff:
                            item['is_anomali'] = True
                            item['anomali_reasons'].append(
                                f"⚠️ Anomali Subtotal Excel: Faktur '{item['vendor_nama']}' (Rp {float(item_sisa):,.0f}) "
                                f"ber-kode 'U' namun tidak tercakup dalam rumus subtotal T1 Excel "
                                f"(selisih Rp {float(remaining_diff):,.0f}). User bisa pilih TERIMA atau ABAIKAN."
                            )
                            item['user_action'] = 'abaikan'
                            total_sisa_utang -= item_sisa
                            total_utang_aktif -= 1
                            if not any(r != item['anomali_reasons'][-1] for r in item['anomali_reasons'][:-1] if 'Anomali' in r):
                                total_anomali += 1
                            remaining_diff -= item_sisa

            return Response({
                'summary': {
                    'total_rows': total_rows,
                    'total_nominal': float(total_nominal),
                    'total_bayar': float(total_bayar),
                    'total_sisa_utang': float(total_sisa_utang),
                    'total_anomali': total_anomali,
                    'total_lunas': total_lunas,
                    'total_utang_aktif': total_utang_aktif,
                    'excel_t1_value': float(excel_t1_value) if excel_t1_value is not None else None,
                },
                'items': staged_items
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'error': f'Gagal membaca file Excel: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='ots-commit')
    def ots_commit(self, request):
        items = request.data.get('items')
        if not items or not isinstance(items, list):
            return Response({'error': 'Daftar items yang diverifikasi wajib dikirim.'}, status=status.HTTP_400_BAD_REQUEST)

        active_items = [item for item in items if item.get('user_action') != 'abaikan']

        # Group items by SPB + Vendor + Sumber so installments of the SAME SPB are consolidated under 1 Main Record!
        grouped_spb = defaultdict(list)
        standalone_items = []

        for item in active_items:
            no_spb = (item.get('no_spb') or '').strip().upper()
            v_nama = (item.get('vendor_nama') or '').strip().upper()
            kategori = item.get('kategori', '')

            kat_upper = kategori.upper()
            if any(kw in kat_upper for kw in ['ATK', 'RUMAH TANGGA', 'CETAKAN']):
                sumber = UtangSupplier.SUMBER_MANUAL
            elif 'BHP' in kat_upper or 'OBAT' in kat_upper:
                sumber = UtangSupplier.SUMBER_FARMASI
            else:
                sumber = UtangSupplier.SUMBER_MANUAL

            if no_spb:
                grouped_spb[(sumber, v_nama, no_spb)].append(item)
            else:
                standalone_items.append([item])

        all_groups = list(grouped_spb.values()) + standalone_items

        # Pre-fetch all vendor masters into memory for instant O(1) lookup
        with connection.cursor() as cursor:
            cursor.execute("SELECT id_rekanan, nama, kategori FROM rssams.rekanan")
            rekanan_rows = cursor.fetchall()

        rekanan_exact_map = {}
        rekanan_stripped_map = {}
        next_rekanan_id = (max((r[0] for r in rekanan_rows), default=0)) + 1

        for r_id, r_n, r_k in rekanan_rows:
            if r_n:
                r_low = r_n.lower().strip()
                rekanan_exact_map[r_low] = (r_id, r_n, r_k)
                r_strip = re.sub(r'[^a-zA-Z0-9]', '', r_low)
                if r_strip:
                    rekanan_stripped_map[r_strip] = (r_id, r_n, r_k)

        committed_count = 0
        auto_lunas_count = 0
        vendor_lunas_watermarks = {}  # (vendor_id, vendor_nama_upper) -> latest_lunas_date

        with transaction.atomic():
            for group in all_groups:
                if not group:
                    continue

                # Main item selection: Prefer the active debt item (sisa_utang > 0), or latest item in group
                unpaid_items = [i for i in group if Decimal(str(i.get('sisa_utang', 0))) > Decimal('0')]
                main_item = unpaid_items[-1] if unpaid_items else group[-1]

                v_nama = (main_item.get('vendor_nama', 'VENDOR UNKNOWN').strip())[:145]
                kategori = main_item.get('kategori', '')
                no_spb = (main_item.get('no_spb', '').strip())[:45]
                no_faktur = (main_item.get('no_faktur', '').strip() or f"OTS/{main_item.get('row_idx')}")[:95]

                kat_upper = kategori.upper()
                if any(kw in kat_upper for kw in ['ATK', 'RUMAH TANGGA', 'CETAKAN']):
                    sumber = UtangSupplier.SUMBER_MANUAL
                elif 'BHP' in kat_upper or 'OBAT' in kat_upper:
                    sumber = UtangSupplier.SUMBER_FARMASI
                else:
                    sumber = UtangSupplier.SUMBER_MANUAL

                def parse_date_obj(val_str):
                    return parse_lenient_date(val_str)

                # Calculate total contract nominal & total payments across group
                total_nominal = sum(Decimal(str(i.get('nominal', 0))) for i in group)
                total_bayar = sum(Decimal(str(i.get('jumlah_bayar', 0))) for i in group)
                total_sisa = max(Decimal('0'), total_nominal - total_bayar)

                if total_nominal <= Decimal('0'):
                    continue

                status_code = str(main_item.get('status_excel', '')).strip().upper()
                if status_code == 'U':
                    if total_bayar >= total_nominal:
                        total_bayar = Decimal('0')
                    total_sisa = max(Decimal('0'), total_nominal - total_bayar)
                    st = UtangSupplier.STATUS_SEBAGIAN if total_bayar > Decimal('0') else UtangSupplier.STATUS_BELUM_DIBAYAR
                elif status_code == 'L':
                    total_bayar = total_nominal
                    total_sisa = Decimal('0')
                    st = UtangSupplier.STATUS_LUNAS
                else:
                    if total_sisa <= Decimal('0'):
                        st = UtangSupplier.STATUS_LUNAS
                    elif total_bayar > Decimal('0'):
                        st = UtangSupplier.STATUS_SEBAGIAN
                    else:
                        st = UtangSupplier.STATUS_BELUM_DIBAYAR

                tgl_faktur = (
                    parse_date_obj(main_item.get('tgl_faktur')) or
                    parse_date_obj(main_item.get('tgl_titip')) or
                    parse_date_obj(main_item.get('tgl_rencana_bayar')) or
                    parse_date_obj(main_item.get('tgl_proses')) or
                    parse_date_obj(main_item.get('tgl_app')) or
                    timezone.localdate()
                )
                tgl_titip = (
                    parse_date_obj(main_item.get('tgl_titip')) or
                    parse_date_obj(main_item.get('tgl_rencana_bayar')) or
                    parse_date_obj(main_item.get('tgl_proses')) or
                    parse_date_obj(main_item.get('tgl_app')) or
                    tgl_faktur
                )

                vendor_id = main_item.get('vendor_id')
                v_nama_final = v_nama

                # Fast memory lookup for vendor master resolution
                if v_nama:
                    v_low = v_nama.lower().strip()
                    if v_low in rekanan_exact_map:
                        vendor_id, v_nama_final, _ = rekanan_exact_map[v_low]
                    else:
                        v_strip = re.sub(r'[^a-zA-Z0-9]', '', v_low)
                        if v_strip in rekanan_stripped_map:
                            vendor_id, v_nama_final, _ = rekanan_stripped_map[v_strip]
                        elif not vendor_id:
                            vendor_id = next_rekanan_id
                            v_nama_final = v_nama
                            rekanan_exact_map[v_low] = (vendor_id, v_nama, kategori)
                            rekanan_stripped_map[v_strip] = (vendor_id, v_nama, kategori)
                            with connection.cursor() as cursor:
                                cursor.execute("""
                                    INSERT INTO rssams.rekanan (id_rekanan, nama, alamat, telp, kc, del, sumber, kategori)
                                    VALUES (%s, %s, '', '', '', 'N', 'ots_import', %s)
                                """, [vendor_id, v_nama[:100], (kategori or '')[:100]])
                            next_rekanan_id += 1

                vendor_id = vendor_id or 9999

                # Track latest Lunas date per vendor for High-Water Mark Auto-Lunas
                if st == UtangSupplier.STATUS_LUNAS and tgl_faktur:
                    v_clean = _clean_vendor_name(v_nama_final)
                    if vendor_id and vendor_id != 9999:
                        curr = vendor_lunas_watermarks.get(vendor_id)
                        if not curr or tgl_faktur > curr:
                            vendor_lunas_watermarks[vendor_id] = tgl_faktur
                    if v_clean:
                        curr = vendor_lunas_watermarks.get(v_clean)
                        if not curr or tgl_faktur > curr:
                            vendor_lunas_watermarks[v_clean] = tgl_faktur

                ket_detail = (main_item.get('keterangan_excel') or main_item.get('no_faktur') or '').strip()
                full_keterangan = f"[{kategori}] {ket_detail}" if kategori else ket_detail

                # Create 1 Main UtangSupplier Record for this SPB!
                utang = UtangSupplier.objects.create(
                    app_siaga_faktur_id=f"OTS-{main_item.get('row_idx')}",
                    sumber=sumber,
                    vendor_id=vendor_id,
                    vendor_nama=v_nama_final[:145],
                    kategori=kategori[:100] if kategori else "",
                    nomor_faktur=no_faktur,
                    nomor_spb=no_spb if no_spb else "",
                    tanggal_faktur=tgl_faktur,
                    tanggal_titip=tgl_titip,
                    nominal=total_nominal,
                    keterangan_titip=full_keterangan[:250],
                    status=st,
                    verified_by=request.user,
                    verified_at=timezone.now(),
                )

                # Create child PembayaranUtang for EVERY payment item in group!
                payments_to_create = []
                for item in group:
                    byr = Decimal(str(item.get('jumlah_bayar', 0)))
                    if byr > Decimal('0'):
                        tgl_parsed_rencana = parse_date_obj(item.get('tgl_rencana_bayar'))
                        tgl_parsed_proses = parse_date_obj(item.get('tgl_proses'))
                        tgl_parsed_app = parse_date_obj(item.get('tgl_app'))

                        tgl_proses = tgl_parsed_proses or tgl_parsed_app or tgl_parsed_rencana or tgl_faktur
                        tgl_rencana = tgl_parsed_rencana or tgl_proses
                        tgl_app = tgl_parsed_app or tgl_proses
                        item_ket = (item.get('keterangan_excel') or item.get('no_faktur') or f"Row #{item.get('row_idx')}").strip()

                        payments_to_create.append(PembayaranUtang(
                            utang=utang,
                            tanggal_rencana_bayar=tgl_rencana,
                            tanggal_proses=tgl_proses,
                            tanggal_app=tgl_app,
                            jumlah_bayar=byr,
                            potongan_deposit=Decimal('0'),
                            jumlah_kas_keluar=byr,
                            keterangan=f"Realisasi Saldo Awal OTS (Baris #{item.get('row_idx')}) - {item_ket}"[:250],
                            status=PembayaranUtang.STATUS_REALISASI_LUNAS,
                            created_by=request.user,
                        ))

                if payments_to_create:
                    PembayaranUtang.objects.bulk_create(payments_to_create)

                utang.refresh_status()
                committed_count += 1

            # Execute High-Water Mark Auto-Lunas for un-imported legacy purchases <= vendor's latest lunas date
            if vendor_lunas_watermarks:
                with connection.cursor() as cursor:
                    # 1. Fetch pending Farmasi purchases
                    cursor.execute("""
                        SELECT t.id, t.id_rekanan, r.nama, t.tgl_faktur, t.no_faktur, t.no_spb,
                               COALESCE(t.gtotal, t.total, 0) AS total_biaya
                        FROM rssams.tran_beli_brg_farmasi t
                        LEFT JOIN rssams.rekanan r ON r.id_rekanan = t.id_rekanan
                        LEFT JOIN keuangan_utang_supplier u ON u.app_siaga_faktur_id = t.id
                        WHERE u.id IS NULL
                          AND (t.id NOT IN (SELECT nomor_spb FROM keuangan_utang_supplier WHERE nomor_spb != ''))
                          AND (t.no_faktur IS NULL OR t.no_faktur = '' OR t.no_faktur NOT IN (SELECT nomor_faktur FROM keuangan_utang_supplier WHERE nomor_faktur != ''))
                    """)
                    columns = [col[0] for col in cursor.description]
                    all_pending_farm = [dict(zip(columns, row)) for row in cursor.fetchall()]

                    # 2. Fetch pending Logistik purchases
                    cursor.execute("""
                        SELECT t.id, r.id_rekanan, COALESCE(t.rekanan, r.nama, '') AS nama, t.tgl_spk AS tgl_faktur, t.no_spk AS no_faktur, t.id AS no_spb,
                               COALESCE(t.nilai, 0) AS total_biaya
                        FROM rssams.tran_beli_brg_log t
                        LEFT JOIN rssams.rekanan r ON r.nama = t.rekanan
                        LEFT JOIN keuangan_utang_supplier u ON u.app_siaga_faktur_id = t.id
                        WHERE t.done = 'Y'
                          AND COALESCE(t.rekanan, '') != 'STOCK OPNAME'
                          AND COALESCE(t.no_spk, '') NOT LIKE 'OPNAME-%%'
                          AND u.id IS NULL
                          AND (t.id NOT IN (SELECT nomor_spb FROM keuangan_utang_supplier WHERE nomor_spb != ''))
                          AND (t.no_spk IS NULL OR t.no_spk = '' OR t.no_spk NOT IN (SELECT nomor_faktur FROM keuangan_utang_supplier WHERE nomor_faktur != ''))
                    """)
                    columns = [col[0] for col in cursor.description]
                    all_pending_log = [dict(zip(columns, row)) for row in cursor.fetchall()]

                auto_list = []
                for p in all_pending_farm:
                    v_id = p.get('id_rekanan')
                    v_clean = _clean_vendor_name(p.get('nama'))
                    max_w = (vendor_lunas_watermarks.get(v_id) if (v_id and v_id != 9999) else None) or vendor_lunas_watermarks.get(v_clean)

                    if max_w:
                        tgl_fak = p.get('tgl_faktur')
                        tgl_fak_date = None
                        if tgl_fak:
                            if isinstance(tgl_fak, datetime):
                                tgl_fak_date = tgl_fak.date()
                            elif isinstance(tgl_fak, str):
                                try:
                                    tgl_fak_date = datetime.strptime(tgl_fak[:10], '%Y-%m-%d').date()
                                except Exception:
                                    tgl_fak_date = None
                            else:
                                tgl_fak_date = tgl_fak

                        if tgl_fak_date is None or tgl_fak_date <= max_w:
                            auto_list.append((p, UtangSupplier.SUMBER_FARMASI, max_w, tgl_fak_date or max_w))

                for p in all_pending_log:
                    v_id = p.get('id_rekanan')
                    v_clean = _clean_vendor_name(p.get('nama'))
                    max_w = (vendor_lunas_watermarks.get(v_id) if (v_id and v_id != 9999) else None) or vendor_lunas_watermarks.get(v_clean)

                    if max_w:
                        tgl_fak = p.get('tgl_faktur')
                        tgl_fak_date = None
                        if tgl_fak:
                            if isinstance(tgl_fak, datetime):
                                tgl_fak_date = tgl_fak.date()
                            elif isinstance(tgl_fak, str):
                                try:
                                    tgl_fak_date = datetime.strptime(tgl_fak[:10], '%Y-%m-%d').date()
                                except Exception:
                                    tgl_fak_date = None
                            else:
                                tgl_fak_date = tgl_fak

                        if tgl_fak_date is None or tgl_fak_date <= max_w:
                            auto_list.append((p, UtangSupplier.SUMBER_MANUAL, max_w, tgl_fak_date or max_w))

                utang_auto_objs = []
                auto_meta = []
                for p_item, p_sumber, max_w, tgl_fak_date in auto_list:
                    tot_b = Decimal(str(p_item.get('total_biaya') or 0))
                    if tot_b <= Decimal('0'):
                        continue
                    spb_id = str(p_item.get('id'))
                    no_spb_str = str(p_item.get('no_spb') or spb_id or '')[:45]
                    no_fak_str = str(p_item.get('no_faktur') or f"INV/OTS/AUTO/{spb_id}")[:95]
                    v_id = p_item.get('id_rekanan') or 9999
                    v_nama = str(p_item.get('nama') or 'VENDOR UNKNOWN')[:145]
                    app_id = f"OTS-AUTO-{spb_id}"

                    utang_auto_objs.append(UtangSupplier(
                        app_siaga_faktur_id=app_id,
                        sumber=p_sumber,
                        vendor_id=v_id,
                        vendor_nama=v_nama,
                        kategori="OBAT DAN BHP" if p_sumber == UtangSupplier.SUMBER_FARMASI else "BIAYA ATK, CETAKAN, BHP RUMAH TANGGA DLL.",
                        nomor_faktur=no_fak_str,
                        nomor_spb=no_spb_str,
                        tanggal_faktur=tgl_fak_date,
                        tanggal_titip=tgl_fak_date,
                        nominal=tot_b,
                        keterangan_titip=f"[Auto-Lunas OTS Cutoff] Lunas pra-cutoff {max_w}",
                        status=UtangSupplier.STATUS_LUNAS,
                        verified_by=request.user,
                        verified_at=timezone.now(),
                    ))
                    auto_meta.append((app_id, tot_b, tgl_fak_date, max_w))

                if utang_auto_objs:
                    UtangSupplier.objects.bulk_create(utang_auto_objs, batch_size=100, ignore_conflicts=True)
                    auto_lunas_count = len(utang_auto_objs)

                    created_app_ids = [m[0] for m in auto_meta]
                    utang_map = {}
                    chunk_size = 500
                    for i in range(0, len(created_app_ids), chunk_size):
                        chunk = created_app_ids[i:i + chunk_size]
                        for u in UtangSupplier.objects.filter(app_siaga_faktur_id__in=chunk).values_list('app_siaga_faktur_id', 'id'):
                            utang_map[u[0]] = u[1]

                    pembayaran_auto_objs = []
                    for app_id, tot_b, tgl_fak_date, max_w in auto_meta:
                        u_id = utang_map.get(app_id)
                        if u_id:
                            pembayaran_auto_objs.append(PembayaranUtang(
                                utang_id=u_id,
                                tanggal_rencana_bayar=tgl_fak_date,
                                tanggal_proses=tgl_fak_date,
                                tanggal_app=tgl_fak_date,
                                jumlah_bayar=tot_b,
                                potongan_deposit=Decimal('0'),
                                jumlah_kas_keluar=tot_b,
                                keterangan=f"[Auto-Lunas OTS Cutoff] Realisasi lunas otomatis pra-cutoff {max_w}",
                                status=PembayaranUtang.STATUS_REALISASI_LUNAS,
                                created_by=request.user,
                            ))
                    if pembayaran_auto_objs:
                        PembayaranUtang.objects.bulk_create(pembayaran_auto_objs, batch_size=100, ignore_conflicts=True)

        msg = f'Berhasil menyimpan {committed_count} SPB/faktur dari Excel ke SIMAK.'
        if auto_lunas_count > 0:
            msg += f' Juga melunaskan otomatis {auto_lunas_count} faktur lama yang terpotong tanggal lunas terbaru vendor.'

        return Response({
            'message': msg,
            'committed_count': committed_count,
            'auto_lunas_count': auto_lunas_count,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='ots-rollback')
    def ots_rollback(self, request):
        with transaction.atomic():
            ots_utang_qs = UtangSupplier.objects.filter(app_siaga_faktur_id__startswith='OTS-')
            total_utang = ots_utang_qs.count()

            pembayaran_qs = PembayaranUtang.objects.filter(utang__app_siaga_faktur_id__startswith='OTS-')
            total_pembayaran = pembayaran_qs.count()

            pembayaran_qs.delete()
            ots_utang_qs.delete()

        return Response({
            'message': f'Berhasil mengembalikan data (Undo Import). Menghapus {total_utang} data utang dan {total_pembayaran} riwayat pembayaran hasil import Excel OTS.',
            'deleted_utang_count': total_utang,
            'deleted_pembayaran_count': total_pembayaran
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='sync-ots-dates')
    def sync_ots_dates(self, request):
        qs = PembayaranUtang.objects.filter(
            keterangan__startswith='Realisasi Saldo Awal OTS'
        ).exclude(
            tanggal_proses=F('tanggal_app')
        )
        
        count = qs.count()
        if count == 0:
            return Response({'message': 'Semua tanggal pembayaran OTS sudah sinkron dengan data asli di Excel.'}, status=status.HTTP_200_OK)
            
        updated = qs.update(
            tanggal_proses=F('tanggal_app'), 
            tanggal_rencana_bayar=F('tanggal_app')
        )
        
        return Response({
            'message': f'Berhasil memperbaiki {updated} riwayat pembayaran cicilan OTS yang tanggalnya tidak sinkron.'
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='reset-all')
    def reset_all(self, request):
        """Mereset / Menghapus SELURUH data utang supplier, pembayaran utang, dan deposit vendor."""
        with transaction.atomic():
            total_pembayaran = PembayaranUtang.objects.all().count()
            PembayaranUtang.objects.all().delete()

            total_deposit = DepositVendor.objects.all().count()
            DepositVendor.objects.all().delete()

            total_utang = UtangSupplier.objects.all().count()
            UtangSupplier.objects.all().delete()

        return Response({
            'message': f'Berhasil mereset database. Menghapus {total_utang} data utang, {total_pembayaran} data riwayat pembayaran, dan {total_deposit} data deposit vendor. Seluruh faktur kini kembali bersih di Menunggu Verifikasi.',
            'deleted_utang_count': total_utang,
            'deleted_pembayaran_count': total_pembayaran,
            'deleted_deposit_count': total_deposit,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='ots-export-anomali')
    def ots_export_anomali(self, request):
        items = request.data.get('items', [])
        anomali_items = [item for item in items if item.get('is_anomali')]
        
        if not anomali_items:
            return Response({'error': 'Tidak ada data anomali untuk di-export.'}, status=status.HTTP_400_BAD_REQUEST)
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Anomali OTS"

        headers = [
            "Baris Excel", "Vendor / Distributor", "Kategori Transaksi", "No. SPB", 
            "No. Faktur Supplier", "Tipe / Kategori Anomali", "Tgl Faktur", "Tgl Titip", 
            "Nominal (Rp)", "Dibayar (Rp)", "Sisa Utang (Rp)", "Status Excel", "Detail Penyebab Warning"
        ]
        
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        def determine_anomaly_category(reasons):
            reasons_str = " ".join(reasons).lower()
            cats = []
            if 'human error' in reasons_str or 'duplikat persis' in reasons_str:
                cats.append('Human Error (Duplikat Persis)')
            if 'cicilan' in reasons_str or 'split faktur' in reasons_str or 'juga digunakan' in reasons_str:
                cats.append('SPB Cicilan / Split Delivery')
            if 'sudah tercatat' in reasons_str or 'sudah ada' in reasons_str:
                cats.append('Sudah Ada di DB SIMAK')
            if 'tidak memiliki nomor spb' in reasons_str or 'tanpa nomor spb' in reasons_str or 'manual non-spb' in reasons_str:
                cats.append('Catat Manual (Non-SPB)')
            if 'kode status excel' in reasons_str:
                cats.append('Status Excel Mismatch')
            return " | ".join(cats) if cats else 'Perlu Review'

        for item in anomali_items:
            reasons_list = item.get('anomali_reasons', [])
            reasons_text = " | ".join(reasons_list)
            tipe_anomali = determine_anomaly_category(reasons_list)
            
            row_data = [
                item.get('row_idx'),
                item.get('vendor_nama'),
                item.get('kategori'),
                item.get('no_spb'),
                item.get('no_faktur'),
                tipe_anomali,
                item.get('tgl_faktur'),
                item.get('tgl_titip'),
                float(item.get('nominal', 0)),
                float(item.get('jumlah_bayar', 0)),
                float(item.get('sisa_utang', 0)),
                item.get('status_excel'),
                reasons_text
            ]
            ws.append(row_data)

        for row in ws.iter_rows(min_row=2, max_row=len(anomali_items) + 1):
            for col_idx, cell in enumerate(row):
                cell.border = thin_border
                if col_idx in (8, 9, 10):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in (0, 3, 5, 6, 7, 11):
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)

        ws.auto_filter.ref = ws.dimensions

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 65)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Laporan_Anomali_OTS_2026.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        qs = self.filter_queryset(self.get_queryset())
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daftar Utang Supplier"

        ws.merge_cells('A1:O1')
        ws['A1'] = 'REKAP DAFTAR UTANG SUPPLIER (OBAT, BHP & LOGISTIK)'
        ws['A1'].font = Font(bold=True, size=14, color='1E293B')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws['A2'] = f'Tanggal Export: {timezone.now().strftime("%d-%m-%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=10, color='64748B')

        headers = [
            'No', 'Sumber', 'Vendor / Supplier', 'Kategori', 'No. SPB', 'No. Faktur',
            'Tgl SPB', 'Tgl Faktur', 'Tgl Titip', 'Jatuh Tempo', 'Umur Utang',
            'Nominal Faktur (Rp)', 'Sudah Dibayar (Rp)', 'Sisa Utang (Rp)', 'Status'
        ]
        ws.append([])
        ws.append(headers)

        header_row = 4
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        today_date = timezone.localdate()
        tot_nom = 0.0
        tot_byr = 0.0
        tot_sisa = 0.0

        utang_list = list(qs)
        utang_ids = [u.id for u in utang_list]
        pay_map = {}
        if utang_ids:
            pays = (
                PembayaranUtang.objects
                .filter(utang_id__in=utang_ids, status__in=['realisasi_sebagian', 'realisasi_lunas', 'retur'])
                .values('utang_id')
                .annotate(total=Sum('jumlah_bayar'))
            )
            for p in pays:
                pay_map[p['utang_id']] = float(p['total'] or 0)

        for idx, u in enumerate(utang_list, 1):
            nom = float(u.nominal or 0)
            byr = pay_map.get(u.id, 0.0)
            sisa = max(nom - byr, 0.0)
            tot_nom += nom
            tot_byr += byr
            tot_sisa += sisa

            umur_str = f"{max(0, (today_date - u.tanggal_titip).days)} Hari" if u.tanggal_titip else '-'
            status_lbl = dict(UtangSupplier.STATUS_CHOICES).get(u.status, u.status)

            ws.append([
                idx,
                u.get_sumber_display(),
                u.vendor_nama,
                u.kategori or '-',
                u.nomor_spb or '-',
                u.nomor_faktur or '-',
                u.tanggal_spb.strftime('%d-%m-%Y') if u.tanggal_spb else '-',
                u.tanggal_faktur.strftime('%d-%m-%Y') if u.tanggal_faktur else '-',
                u.tanggal_titip.strftime('%d-%m-%Y') if u.tanggal_titip else '-',
                u.tanggal_jatuh_tempo.strftime('%d-%m-%Y') if u.tanggal_jatuh_tempo else '-',
                umur_str,
                nom,
                byr,
                sisa,
                status_lbl,
            ])

        total_row_idx = ws.max_row + 1
        ws.cell(row=total_row_idx, column=1, value='TOTAL')
        ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=11)
        ws.cell(row=total_row_idx, column=12, value=tot_nom)
        ws.cell(row=total_row_idx, column=13, value=tot_byr)
        ws.cell(row=total_row_idx, column=14, value=tot_sisa)

        total_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
        total_font = Font(bold=True, color='0F172A')

        for c_idx in range(1, 16):
            c = ws.cell(row=total_row_idx, column=c_idx)
            c.fill = total_fill
            c.font = total_font
            if c_idx in (12, 13, 14):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')
            else:
                c.alignment = Alignment(horizontal='center', vertical='center')

        ws.auto_filter.ref = f'A4:O{ws.max_row}'

        col_widths = {
            'A': 8, 'B': 12, 'C': 32, 'D': 24, 'E': 18, 'F': 22,
            'G': 14, 'H': 14, 'I': 14, 'J': 14, 'K': 14,
            'L': 20, 'M': 20, 'N': 20, 'O': 18
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Rekap_Utang_Supplier_{today_date}.xlsx"'
def _handle_petty_cash_payment_realisasi(pembayaran, user):
    """
    Jika faktur utang bersumber dari KEUANGAN (Pengisian Kas Kecil / Petty Cash),
    tambahkan SaldoPettyCash dan catat RiwayatSaldoPettyCash saat pembayaran direalisasikan.
    """
    utang = pembayaran.utang
    if utang.sumber == UtangSupplier.SUMBER_KEUANGAN or utang.kategori == 'PENGISIAN PETTY CASH':
        nominal_bayar = pembayaran.jumlah_bayar
        if nominal_bayar and nominal_bayar > 0:
            saldo = get_or_create_saldo()
            saldo_sebelum = saldo.saldo
            saldo.saldo += nominal_bayar
            saldo.updated_by = user
            saldo.save()

            RiwayatSaldoPettyCash.objects.create(
                jenis='penambahan',
                jumlah=nominal_bayar,
                saldo_sebelum=saldo_sebelum,
                saldo_sesudah=saldo.saldo,
                keterangan=f'Pencairan pengisian saldo kas kecil ({utang.nomor_spb})' + (f' - {utang.keterangan_titip}' if utang.keterangan_titip else ''),
                created_by=user,
                nama_pengaju=utang.vendor_nama,
                unit_pengaju='Keuangan / Kasir Kas Kecil',
            )

def _handle_petty_cash_payment_batal_realisasi(pembayaran, user):
    """
    Jika realisasi pembayaran utang pengisian kas kecil dibatalkan,
    tarik kembali penambahan saldo dari SaldoPettyCash dan catat pengurangan di riwayat.
    """
    utang = pembayaran.utang
    if utang.sumber == UtangSupplier.SUMBER_KEUANGAN or utang.kategori == 'PENGISIAN PETTY CASH':
        nominal_bayar = pembayaran.jumlah_bayar
        if nominal_bayar and nominal_bayar > 0:
            saldo = get_or_create_saldo()
            saldo_sebelum = saldo.saldo
            saldo.saldo -= nominal_bayar
            saldo.updated_by = user
            saldo.save()

            RiwayatSaldoPettyCash.objects.create(
                jenis='pengurangan',
                jumlah=nominal_bayar,
                saldo_sebelum=saldo_sebelum,
                saldo_sesudah=saldo.saldo,
                keterangan=f'Koreksi pembatalan realisasi pencairan kas kecil ({utang.nomor_spb})',
                created_by=user,
                nama_pengaju=utang.vendor_nama,
                unit_pengaju='Keuangan / Kasir Kas Kecil',
            )

class PembayaranUtangViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset = PembayaranUtang.objects.select_related('utang', 'created_by').all()
    serializer_class = PembayaranUtangSerializer
    permission_classes = [IsAuthenticated, IsCatatanUtangObatBhpPermission]

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        search = (params.get('search') or '').strip()
        vendor_id = params.get('vendor_id')
        utang_id = params.get('utang') or params.get('utang__id')
        status_param = (params.get('status') or '').strip()
        sumber_filter = (params.get('sumber') or '').strip()
        dari = params.get('dari')
        sampai = params.get('sampai')

        if search:
            qs = qs.filter(
                Q(utang__nomor_faktur__icontains=search)
                | Q(utang__nomor_spb__icontains=search)
                | Q(utang__vendor_nama__icontains=search)
                | Q(utang__keterangan_titip__icontains=search)
                | Q(utang__kategori__icontains=search)
                | Q(keterangan__icontains=search)
            )
        if vendor_id:
            qs = qs.filter(utang__vendor_id=vendor_id)
        if utang_id:
            qs = qs.filter(utang_id=utang_id)
        if status_param:
            if status_param == 'realisasi':
                qs = qs.filter(Q(status__startswith='realisasi') | Q(status='realisasi') | Q(status='retur'))
            else:
                qs = qs.filter(status=status_param)
        if sumber_filter and sumber_filter != 'semua':
            qs = qs.filter(utang__sumber=sumber_filter)
        if dari:
            qs = qs.filter(tanggal_proses__gte=dari)
        if sampai:
            qs = qs.filter(tanggal_proses__lte=sampai)

        order = _utang_order_clause(params.get('ordering'), {
            'vendor': 'utang__vendor_nama',
            '-vendor': '-utang__vendor_nama',
            'nomor_faktur': 'utang__nomor_faktur',
            '-nomor_faktur': '-utang__nomor_faktur',
            'tanggal_proses': 'tanggal_proses',
            '-tanggal_proses': '-tanggal_proses',
            'jumlah_bayar': 'jumlah_bayar',
            '-jumlah_bayar': '-jumlah_bayar',
            'status': 'status',
            '-status': '-status',
        })
        return qs.order_by(order, '-created_at')

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        utang = instance.utang
        if instance.status != PembayaranUtang.STATUS_PENDING:
            return Response({'error': 'Hanya pengajuan pembayaran yang berstatus pending yang dapat dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)
        self.perform_destroy(instance)
        utang.refresh_status()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'], url_path='realisasi')
    def realisasi(self, request, pk=None):
        pembayaran = self.get_object()
        if pembayaran.status != PembayaranUtang.STATUS_PENDING:
            return Response({'error': 'Pengajuan ini sudah direalisasikan atau tidak valid.'}, status=status.HTTP_400_BAD_REQUEST)
        tanggal_realisasi = request.data.get('tanggal_realisasi') or timezone.now().date().isoformat()
        
        jumlah_bayar_raw = request.data.get('jumlah_bayar')
        if jumlah_bayar_raw is not None:
            try:
                jumlah_bayar = Decimal(str(jumlah_bayar_raw))
                if jumlah_bayar <= 0:
                    return Response({'error': 'Jumlah bayar harus lebih dari 0.'}, status=status.HTTP_400_BAD_REQUEST)
                pembayaran.jumlah_bayar = jumlah_bayar
            except (TypeError, ValueError, InvalidOperation):
                return Response({'error': 'Jumlah bayar tidak valid.'}, status=status.HTTP_400_BAD_REQUEST)

        # Hitung sisa utang faktur SEBELUM realisasi ini disimpan
        total_realisasi_lainnya = pembayaran.utang.pembayaran.filter(
            status__in=[PembayaranUtang.STATUS_REALISASI_SEBAGIAN, PembayaranUtang.STATUS_REALISASI_LUNAS, PembayaranUtang.STATUS_RETUR]
        ).exclude(pk=pembayaran.pk).aggregate(total=Sum('jumlah_bayar'))['total'] or Decimal('0')
        
        sisa_sebelumnya = pembayaran.utang.nominal - total_realisasi_lainnya
        
        with transaction.atomic():
            if pembayaran.jumlah_bayar >= sisa_sebelumnya:
                pembayaran.status = PembayaranUtang.STATUS_REALISASI_LUNAS
            else:
                pembayaran.status = PembayaranUtang.STATUS_REALISASI_SEBAGIAN

            pembayaran.tanggal_proses = tanggal_realisasi
            if not pembayaran.tanggal_app:
                pembayaran.tanggal_app = tanggal_realisasi
            pembayaran.save(update_fields=['status', 'tanggal_proses', 'tanggal_app', 'jumlah_bayar'])

            # Potong terpakai pada DepositVendor jika menggunakan potongan_deposit
            if pembayaran.potongan_deposit and pembayaran.potongan_deposit > 0:
                sisa_potong = pembayaran.potongan_deposit
                active_deposits = DepositVendor.objects.filter(vendor_id=pembayaran.utang.vendor_id).order_by('created_at')
                for dep in active_deposits:
                    sisa_dep = dep.sisa_deposit
                    if sisa_dep <= 0:
                        continue
                    potong_dep = min(sisa_potong, sisa_dep)
                    dep.terpakai += potong_dep
                    dep.save(update_fields=['terpakai', 'updated_at'])
                    sisa_potong -= potong_dep
                    if sisa_potong <= 0:
                        break

            # Jika utang adalah pengisian petty cash, tambahkan saldo kas kecil
            _handle_petty_cash_payment_realisasi(pembayaran, request.user)

            utang = pembayaran.utang
            utang.refresh_status()

        return Response({
            'pembayaran': PembayaranUtangSerializer(pembayaran, context={'request': request}).data,
            'utang': UtangSupplierSerializer(utang, context={'request': request}).data,
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='edit-tanggal')
    def edit_tanggal(self, request, pk=None):
        pembayaran = self.get_object()
        tanggal_baru = request.data.get('tanggal_realisasi') or request.data.get('tanggal_proses')
        if not tanggal_baru:
            return Response({'error': 'Tanggal realisasi baru wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)
        
        pembayaran.tanggal_proses = tanggal_baru
        pembayaran.tanggal_app = tanggal_baru
        pembayaran.save(update_fields=['tanggal_proses', 'tanggal_app'])
        
        utang = pembayaran.utang
        utang.refresh_status()
        return Response({
            'message': f'Tanggal realisasi pembayaran berhasil diperbarui.',
            'pembayaran': PembayaranUtangSerializer(pembayaran, context={'request': request}).data,
            'utang': UtangSupplierSerializer(utang, context={'request': request}).data,
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='batal-realisasi')
    def batal_realisasi(self, request, pk=None):
        pembayaran = self.get_object()
        utang = pembayaran.utang
        jumlah_bayar = pembayaran.jumlah_bayar
        
        with transaction.atomic():
            if pembayaran.potongan_deposit and pembayaran.potongan_deposit > 0:
                sisa_kembali = pembayaran.potongan_deposit
                deposits = DepositVendor.objects.filter(vendor_id=utang.vendor_id, terpakai__gt=0).order_by('-created_at')
                for dep in deposits:
                    kembali = min(sisa_kembali, dep.terpakai)
                    dep.terpakai -= kembali
                    dep.save(update_fields=['terpakai', 'updated_at'])
                    sisa_kembali -= kembali
                    if sisa_kembali <= 0:
                        break
            
            pembayaran.status = 'batal'
            pembayaran.keterangan = f"{pembayaran.keterangan or ''} [DIBATALKAN REALISASI]".strip()
            pembayaran.save(update_fields=['status', 'keterangan', 'updated_at'])
            
            # Jika utang adalah pengisian petty cash, tarik kembali penambahan saldo
            _handle_petty_cash_payment_batal_realisasi(pembayaran, request.user)

            utang.refresh_status()
            
        return Response({
            'message': f'Realisasi pembayaran Rp {jumlah_bayar:,.0f} berhasil dibatalkan dan sisa utang telah dipulihkan.'
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='bulk-realisasi')
    def bulk_realisasi(self, request):
        ids = request.data.get('ids', [])
        if not ids or not isinstance(ids, list):
            return Response({'error': 'Daftar ID pengajuan pembayaran (ids) wajib dikirim.'}, status=status.HTTP_400_BAD_REQUEST)

        tanggal_realisasi = request.data.get('tanggal_realisasi') or timezone.localdate().isoformat()

        pembayaran_list = list(
            PembayaranUtang.objects.filter(
                id__in=ids,
                status=PembayaranUtang.STATUS_PENDING
            ).select_related('utang')
        )

        if not pembayaran_list:
            return Response({'error': 'Tidak ada pengajuan pembayaran pending yang valid untuk direalisasikan.'}, status=status.HTTP_400_BAD_REQUEST)

        realized_count = 0
        total_nominal = Decimal('0')

        with transaction.atomic():
            for pembayaran in pembayaran_list:
                total_realisasi_lainnya = pembayaran.utang.pembayaran.filter(
                    status__in=[PembayaranUtang.STATUS_REALISASI_SEBAGIAN, PembayaranUtang.STATUS_REALISASI_LUNAS, PembayaranUtang.STATUS_RETUR]
                ).exclude(pk=pembayaran.pk).aggregate(total=Sum('jumlah_bayar'))['total'] or Decimal('0')
                
                sisa_sebelumnya = pembayaran.utang.nominal - total_realisasi_lainnya
                
                if pembayaran.jumlah_bayar >= sisa_sebelumnya:
                    pembayaran.status = PembayaranUtang.STATUS_REALISASI_LUNAS
                else:
                    pembayaran.status = PembayaranUtang.STATUS_REALISASI_SEBAGIAN

                pembayaran.tanggal_proses = tanggal_realisasi
                if not pembayaran.tanggal_app:
                    pembayaran.tanggal_app = tanggal_realisasi
                pembayaran.save(update_fields=['status', 'tanggal_proses', 'tanggal_app'])

                if pembayaran.potongan_deposit and pembayaran.potongan_deposit > 0:
                    sisa_potong = pembayaran.potongan_deposit
                    active_deposits = DepositVendor.objects.filter(vendor_id=pembayaran.utang.vendor_id).order_by('created_at')
                    for dep in active_deposits:
                        sisa_dep = dep.sisa_deposit
                        if sisa_dep <= 0:
                            continue
                        potong_dep = min(sisa_potong, sisa_dep)
                        dep.terpakai += potong_dep
                        dep.save(update_fields=['terpakai', 'updated_at'])
                        sisa_potong -= potong_dep
                        if sisa_potong <= 0:
                            break

                # Jika utang adalah pengisian petty cash, tambahkan saldo kas kecil
                _handle_petty_cash_payment_realisasi(pembayaran, request.user)

                utang = pembayaran.utang
                utang.refresh_status()
                realized_count += 1
                total_nominal += pembayaran.jumlah_bayar

        return Response({
            'message': f'Berhasil merealisasikan {realized_count} pengajuan pembayaran sekaligus.',
            'realized_count': realized_count,
            'total_nominal': float(total_nominal)
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        from itertools import groupby

        qs = self.get_queryset().filter(status=PembayaranUtang.STATUS_PENDING).select_related('utang', 'created_by')
        ids_param = request.query_params.get('ids')
        if ids_param:
            try:
                ids = [int(x.strip()) for x in str(ids_param).split(',') if x.strip()]
                if ids:
                    qs = qs.filter(id__in=ids)
            except (ValueError, TypeError):
                pass

        qs = list(qs)
        # Urutkan berdasarkan vendor_nama
        qs.sort(key=lambda item: (
            (item.utang.vendor_nama if item.utang and item.utang.vendor_nama else '').upper(),
            item.id
        ))

        wb = Workbook()
        ws = wb.active
        ws.title = "Pengajuan Pembayaran"

        ws.merge_cells('A1:G1')
        ws['A1'] = 'REKAP PENGAJUAN PEMBAYARAN UTANG SUPPLIER'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f'Tanggal Cetak: {timezone.now().strftime("%d-%m-%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=10)

        headers = [
            'No', 'Sumber', 'Vendor / Supplier', 'Umur Utang',
            'Jumlah Bayar (Rp)', 'Keterangan', 'Pengaju (Operator)'
        ]
        ws.append([])
        ws.append(headers)

        header_row = 4
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        subtotal_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        subtotal_font = Font(bold=True, color='0F172A')

        today_date = timezone.localdate()
        global_index = 1
        grand_total = Decimal('0')

        def get_vendor_name(item):
            return (item.utang.vendor_nama if item.utang and item.utang.vendor_nama else 'TANPA VENDOR').strip()

        for vendor_nama, group_items in groupby(qs, key=get_vendor_name):
            items_list = list(group_items)
            vendor_subtotal = Decimal('0')

            for item in items_list:
                jumlah = item.jumlah_bayar or Decimal('0')
                vendor_subtotal += jumlah
                grand_total += jumlah

                utang = item.utang
                if utang and utang.tanggal_titip:
                    days = (today_date - utang.tanggal_titip).days
                    umur_utang_str = f"{max(0, days)} Hari"
                else:
                    umur_utang_str = '-'

                sumber_label = utang.get_sumber_display() if utang else '-'
                operator = item.created_by.username if item.created_by else '-'

                clean_ket = item.keterangan or ''
                if utang and utang.vendor_nama:
                    v_name = utang.vendor_nama.strip()
                    if v_name:
                        clean_ket = clean_ket.replace(f"({v_name})", "").replace(f"({v_name.upper()})", "").replace(f"({v_name.lower()})", "")
                clean_ket = re.sub(r'\s*\([^)]+\)\s*(?=Faktur|\b)', ' ', clean_ket, flags=re.IGNORECASE)
                clean_ket = re.sub(r'\bFaktur\s+', '', clean_ket, flags=re.IGNORECASE)
                clean_ket = re.sub(r'\s+', ' ', clean_ket).strip()

                ws.append([
                    global_index,
                    sumber_label,
                    utang.vendor_nama if utang else '-',
                    umur_utang_str,
                    float(jumlah),
                    clean_ket,
                    operator,
                ])
                global_index += 1

                row_num = ws.max_row
                for col in range(1, 8):
                    c = ws.cell(row=row_num, column=col)
                    c.border = thin_border
                    if col in [1, 2, 4]:
                        c.alignment = Alignment(horizontal='center')
                    elif col == 5:
                        c.number_format = '#,##0.00'
                        c.alignment = Alignment(horizontal='right')

            # Subtotal per vendor tepat setelah kelompok vendor berakhir
            subtotal_row = ws.max_row + 1
            ws.cell(row=subtotal_row, column=1, value=f'SUBTOTAL {vendor_nama.upper()}')
            ws.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=4)
            
            subtotal_cell = ws.cell(row=subtotal_row, column=5, value=float(vendor_subtotal))
            subtotal_cell.number_format = '#,##0.00'

            for col in range(1, 8):
                c = ws.cell(row=subtotal_row, column=col)
                c.fill = subtotal_fill
                c.font = subtotal_font
                c.border = thin_border
                if col == 1:
                    c.alignment = Alignment(horizontal='right', vertical='center')
                elif col == 5:
                    c.alignment = Alignment(horizontal='right', vertical='center')

        # Baris Grand Total Pengajuan
        grand_row = ws.max_row + 1
        ws.cell(row=grand_row, column=1, value='GRAND TOTAL PENGAJUAN')
        ws.merge_cells(start_row=grand_row, start_column=1, end_row=grand_row, end_column=4)
        
        grand_cell = ws.cell(row=grand_row, column=5, value=float(grand_total))
        grand_cell.number_format = '#,##0.00'

        grand_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        grand_font = Font(bold=True, color='FFFFFF')

        for col in range(1, 8):
            c = ws.cell(row=grand_row, column=col)
            c.fill = grand_fill
            c.font = grand_font
            c.border = thin_border
            if col == 1:
                c.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 5:
                c.alignment = Alignment(horizontal='right', vertical='center')

        col_widths = {
            'A': 8,
            'B': 16,
            'C': 35,
            'D': 16,
            'E': 22,
            'F': 38,
            'G': 20,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Rekap_Pengajuan_Pembayaran_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response
        return qs.order_by(order, '-created_at')

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        rows = response.data.get('results') if isinstance(response.data, dict) else response.data
        if isinstance(rows, list):
            running = defaultdict(Decimal)
            page_rows = sorted(rows, key=lambda item: (item.get('utang') or 0, item.get('tanggal_proses') or '', item.get('id') or 0))
            prior_cache = {}
            for item in page_rows:
                utang_id = item.get('utang')
                if utang_id not in prior_cache:
                    pay = PembayaranUtang.objects.filter(
                        utang_id=utang_id,
                        tanggal_proses__lt=item.get('tanggal_proses'),
                    ).aggregate(total=Sum('jumlah_bayar'))['total'] or Decimal('0')
                    prior_cache[utang_id] = pay
                running[utang_id] += Decimal(str(item.get('jumlah_bayar') or 0))
                item['running_total_dibayar'] = prior_cache[utang_id] + running[utang_id]
                item['running_sisa_utang'] = max(Decimal(str(item.get('nominal') or 0)) - item['running_total_dibayar'], Decimal('0'))
        return response

class UtangMenungguVerifikasiView(APIView):
    permission_classes = [IsAuthenticated, IsCatatanUtangObatBhpPermission]
    pagination_class = OptionalPageNumberPagination

    def get(self, request):
        params = request.query_params
        sumber_filter = (params.get('sumber') or 'semua').strip()
        try:
            page = int(params.get('page', 1))
            page_size = min(int(params.get('page_size', 10)), 100)
        except ValueError:
            page, page_size = 1, 10
        offset = max(page - 1, 0) * page_size

        # SELECT clause farmasi
        farmasi_select = """
            SELECT
                CONVERT(t.id USING utf8mb4)             AS app_siaga_faktur_id,
                CONVERT(t.id USING utf8mb4)             AS nomor_spb,
                t.tgl_spb                               AS tanggal_spb,
                CONVERT(t.no_faktur USING utf8mb4)       AS nomor_faktur,
                t.id_rekanan                            AS vendor_id,
                t.id_rekanan                            AS vendor_id_hint,
                CONVERT(COALESCE(r.nama, '') USING utf8mb4) AS vendor_nama,
                t.tgl_faktur                            AS tanggal_faktur,
                t.tgl_jtempo                            AS tanggal_jatuh_tempo,
                t.total                                 AS total_sebelum_diskon,
                COALESCE(t.disc1, 0)                    AS disc1,
                COALESCE(t.disc2, 0)                    AS disc2,
                COALESCE(t.disc3, 0)                    AS disc3,
                (t.total - COALESCE(t.disc1, 0) - COALESCE(t.disc2, 0) - COALESCE(t.disc3, 0)) AS total_setelah_diskon,
                COALESCE(t.ppn, 0)                      AS ppn,
                COALESCE(t.materai, 0)                  AS materai,
                t.gtotal                                AS nominal,
                ''                                      AS keterangan,
                'farmasi'                               AS sumber
        """

        # SELECT clause logistik
        logistik_select = """
            SELECT
                CONVERT(t.id USING utf8mb4)             AS app_siaga_faktur_id,
                CONVERT(COALESCE(NULLIF(s.no_spb, ''), NULLIF(t.id_spb, ''), t.id) USING utf8mb4) AS nomor_spb,
                t.tgl_spk                               AS tanggal_spb,
                CONVERT(COALESCE(NULLIF(NULLIF(t.no_spk, ''), '-'), '-') USING utf8mb4) AS nomor_faktur,
                r.id_rekanan                            AS vendor_id,
                r.id_rekanan                            AS vendor_id_hint,
                CONVERT(COALESCE(r.nama, t.rekanan) USING utf8mb4) AS vendor_nama,
                t.tgl_spk                               AS tanggal_faktur,
                NULL                                    AS tanggal_jatuh_tempo,
                t.nilai                                 AS total_sebelum_diskon,
                0.00                                    AS disc1,
                0.00                                    AS disc2,
                0.00                                    AS disc3,
                t.nilai                                 AS total_setelah_diskon,
                0.00                                    AS ppn,
                0.00                                    AS materai,
                t.nilai                                 AS nominal,
                ''                                      AS keterangan,
                'logistik'                              AS sumber
        """

        # SELECT clause keuangan (pengisian kembali saldo)
        keuangan_select = """
            SELECT
                CONVERT(CONCAT('KEU-', t.id) USING utf8mb4) COLLATE utf8mb4_general_ci AS app_siaga_faktur_id,
                CONVERT(t.no_pengajuan USING utf8mb4) COLLATE utf8mb4_general_ci       AS nomor_spb,
                t.tanggal                                                              AS tanggal_spb,
                CONVERT(COALESCE(NULLIF(t.alasan, ''), t.no_pengajuan) USING utf8mb4) COLLATE utf8mb4_general_ci AS nomor_faktur,
                0                                                                      AS vendor_id,
                0                                                                      AS vendor_id_hint,
                CONVERT(COALESCE(NULLIF(TRIM(CONCAT(usr.first_name, ' ', usr.last_name)), ''), usr.username, 'Kasir Petty Cash') USING utf8mb4) COLLATE utf8mb4_general_ci AS vendor_nama,
                t.tanggal                                                              AS tanggal_faktur,
                NULL                                                                   AS tanggal_jatuh_tempo,
                t.nominal_diajukan                                                     AS total_sebelum_diskon,
                0.00                                                                   AS disc1,
                0.00                                                                   AS disc2,
                0.00                                                                   AS disc3,
                t.nominal_diajukan                                                     AS total_setelah_diskon,
                0.00                                                                   AS ppn,
                0.00                                                                   AS materai,
                t.nominal_diajukan                                                     AS nominal,
                ''                                                                     AS keterangan,
                'keuangan'                                                             AS sumber
        """

        where_f, vals_f = _build_pending_where(params)
        where_l, vals_l = _build_pending_where_logistik(params)
        where_k, vals_k = _build_pending_where_keuangan(params)

        # Order mapping berlaku di wrapper query (alias kolom output)
        order = _utang_order_clause(params.get('ordering'), {
            '-tanggal_spb': 'tanggal_spb DESC',
            'tanggal_spb': 'tanggal_spb',
            '-tanggal_faktur': 'tanggal_faktur DESC',
            'tanggal_faktur': 'tanggal_faktur',
            'vendor': 'vendor_nama',
            '-vendor': 'vendor_nama DESC',
            'nomor_spb': 'nomor_spb',
            '-nomor_spb': 'nomor_spb DESC',
            'nomor_faktur': 'nomor_faktur',
            '-nomor_faktur': 'nomor_faktur DESC',
            'created_at': 'tanggal_spb DESC',
            '-created_at': 'tanggal_spb DESC',
            'verified_at': 'tanggal_spb DESC',
            '-verified_at': 'tanggal_spb DESC',
            'tanggal_jatuh_tempo': 'tanggal_jatuh_tempo',
            '-tanggal_jatuh_tempo': 'tanggal_jatuh_tempo DESC',
            'nominal': 'nominal',
            '-nominal': 'nominal DESC',
        })

        with connection.cursor() as cursor:
            if sumber_filter == 'farmasi':
                count_sql = f'SELECT COUNT(*) {_pending_base_sql()} WHERE {where_f}'
                sum_sql = f'SELECT SUM(t.gtotal) {_pending_base_sql()} WHERE {where_f}'
                count_vals = vals_f
                sum_vals = vals_f
                data_sql = f"""
                    SELECT * FROM (
                        {farmasi_select}
                        {_pending_base_sql()}
                        WHERE {where_f}
                    ) AS combined
                    ORDER BY {order}, app_siaga_faktur_id DESC
                    LIMIT %s OFFSET %s
                """
                data_vals = vals_f + [page_size, offset]

            elif sumber_filter == 'logistik':
                count_sql = f'SELECT COUNT(*) {_pending_base_sql_logistik()} WHERE {where_l}'
                sum_sql = f'SELECT SUM(t.nilai) {_pending_base_sql_logistik()} WHERE {where_l}'
                count_vals = vals_l
                sum_vals = vals_l
                data_sql = f"""
                    SELECT * FROM (
                        {logistik_select}
                        {_pending_base_sql_logistik()}
                        WHERE {where_l}
                    ) AS combined
                    ORDER BY {order}, app_siaga_faktur_id DESC
                    LIMIT %s OFFSET %s
                """
                data_vals = vals_l + [page_size, offset]

            elif sumber_filter == 'keuangan':
                count_sql = f'SELECT COUNT(*) {_pending_base_sql_keuangan()} WHERE {where_k}'
                sum_sql = f'SELECT SUM(t.nominal_diajukan) {_pending_base_sql_keuangan()} WHERE {where_k}'
                count_vals = vals_k
                sum_vals = vals_k
                data_sql = f"""
                    SELECT * FROM (
                        {keuangan_select}
                        {_pending_base_sql_keuangan()}
                        WHERE {where_k}
                    ) AS combined
                    ORDER BY {order}, app_siaga_faktur_id DESC
                    LIMIT %s OFFSET %s
                """
                data_vals = vals_k + [page_size, offset]

            else:  # semua
                count_sql = f"""
                    SELECT COUNT(*) FROM (
                        SELECT 1 {_pending_base_sql()} WHERE {where_f}
                        UNION ALL
                        SELECT 1 {_pending_base_sql_logistik()} WHERE {where_l}
                        UNION ALL
                        SELECT 1 {_pending_base_sql_keuangan()} WHERE {where_k}
                    ) AS combined
                """
                sum_sql = f"""
                    SELECT SUM(nominal) FROM (
                        SELECT t.gtotal AS nominal {_pending_base_sql()} WHERE {where_f}
                        UNION ALL
                        SELECT t.nilai AS nominal {_pending_base_sql_logistik()} WHERE {where_l}
                        UNION ALL
                        SELECT t.nominal_diajukan AS nominal {_pending_base_sql_keuangan()} WHERE {where_k}
                    ) AS combined
                """
                count_vals = vals_f + vals_l + vals_k
                sum_vals = vals_f + vals_l + vals_k
                data_sql = f"""
                    SELECT * FROM (
                        {farmasi_select}
                        {_pending_base_sql()}
                        WHERE {where_f}
                        UNION ALL
                        {logistik_select}
                        {_pending_base_sql_logistik()}
                        WHERE {where_l}
                        UNION ALL
                        {keuangan_select}
                        {_pending_base_sql_keuangan()}
                        WHERE {where_k}
                    ) AS combined
                    ORDER BY {order}, app_siaga_faktur_id DESC
                    LIMIT %s OFFSET %s
                """
                data_vals = vals_f + vals_l + vals_k + [page_size, offset]

            cursor.execute(count_sql, count_vals)
            total = cursor.fetchone()[0]
            cursor.execute(sum_sql, sum_vals)
            total_nominal = cursor.fetchone()[0] or 0
            cursor.execute(data_sql, data_vals)
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return Response({
            'count': total,
            'total_nominal': float(total_nominal),
            'next': None,
            'previous': None,
            'results': rows,
        })

    def post(self, request):
        app_siaga_faktur_id = request.data.get('app_siaga_faktur_id')
        sumber = (request.data.get('sumber') or 'farmasi').strip()

        if not app_siaga_faktur_id:
            return Response({'app_siaga_faktur_id': 'Faktur APP_SIAGA wajib dipilih.'}, status=status.HTTP_400_BAD_REQUEST)

        if sumber not in (UtangSupplier.SUMBER_FARMASI, UtangSupplier.SUMBER_LOGISTIK, UtangSupplier.SUMBER_KEUANGAN):
            return Response({'sumber': f'Nilai sumber tidak valid: {sumber}'}, status=status.HTTP_400_BAD_REQUEST)

        # Cek duplikat berdasarkan kombinasi (faktur_id, sumber)
        if UtangSupplier.objects.filter(app_siaga_faktur_id=str(app_siaga_faktur_id), sumber=sumber).exists():
            return Response({'error': 'Faktur ini sudah diverifikasi.'}, status=status.HTTP_400_BAD_REQUEST)

        if sumber == UtangSupplier.SUMBER_FARMASI:
            faktur = _fetch_app_siaga_faktur(app_siaga_faktur_id)
            if not faktur:
                return Response({'error': 'Faktur farmasi tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
            utang = UtangSupplier.objects.create(
                app_siaga_faktur_id=str(faktur['app_siaga_faktur_id']),
                sumber=UtangSupplier.SUMBER_FARMASI,
                nomor_spb=faktur.get('nomor_spb') or '',
                tanggal_spb=faktur.get('tanggal_spb'),
                nomor_faktur=faktur.get('nomor_faktur') or '',
                vendor_id=faktur.get('vendor_id') or 0,
                vendor_nama=faktur.get('vendor_nama') or '',
                tanggal_faktur=faktur.get('tanggal_faktur'),
                tanggal_jatuh_tempo=faktur.get('tanggal_jatuh_tempo'),
                nominal=faktur.get('nominal') or 0,
                tanggal_titip=request.data.get('tanggal_titip') or timezone.localdate(),
                keterangan_titip=request.data.get('keterangan_titip') or '',
                verified_by=request.user,
                verified_at=timezone.now(),
            )

        elif sumber == UtangSupplier.SUMBER_KEUANGAN:
            raw_id = str(app_siaga_faktur_id).replace('KEU-', '')
            try:
                ps = PengajuanPenambahanSaldo.objects.get(pk=raw_id, status='disetujui')
            except (PengajuanPenambahanSaldo.DoesNotExist, ValueError):
                return Response({'error': 'Pengajuan pengisian kembali saldo tidak ditemukan atau belum disetujui.'}, status=status.HTTP_404_NOT_FOUND)
            
            with transaction.atomic():
                utang = UtangSupplier.objects.create(
                    app_siaga_faktur_id=f"KEU-{ps.id}",
                    sumber=UtangSupplier.SUMBER_KEUANGAN,
                    nomor_spb=ps.no_pengajuan,
                    tanggal_spb=ps.tanggal,
                    nomor_faktur=ps.alasan or ps.no_pengajuan,
                    vendor_id=0,
                    vendor_nama=f"Petty Cash - {user_display_name(ps.created_by)}",
                    kategori='PENGISIAN PETTY CASH',
                    tanggal_faktur=ps.tanggal,
                    tanggal_jatuh_tempo=ps.tanggal,
                    nominal=ps.nominal_diajukan or 0,
                    tanggal_titip=request.data.get('tanggal_titip') or timezone.localdate(),
                    keterangan_titip=request.data.get('keterangan_titip') or ps.keterangan or ps.alasan,
                    status=UtangSupplier.STATUS_BELUM_DIBAYAR,
                    verified_by=request.user,
                    verified_at=timezone.now(),
                )
            return Response(UtangSupplierSerializer(utang, context={'request': request}).data, status=status.HTTP_201_CREATED)

        else:  # logistik
            vendor_id = request.data.get('vendor_id')
            if not vendor_id:
                return Response(
                    {'vendor_id': 'Vendor wajib dipilih untuk pembelian logistik.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Validasi vendor_id ke rssams.rekanan
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT id_rekanan, nama FROM rssams.rekanan WHERE id_rekanan = %s AND del = 'N' LIMIT 1",
                    [vendor_id],
                )
                rek = cursor.fetchone()
            if not rek:
                return Response(
                    {'vendor_id': 'Vendor tidak ditemukan di master rekanan.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            vendor_id_int, vendor_nama_rek = rek

            faktur = _fetch_logistik_pembelian(app_siaga_faktur_id)
            if not faktur:
                return Response({'error': 'Data logistik tidak ditemukan atau belum selesai (done != Y).'}, status=status.HTTP_404_NOT_FOUND)

            is_cash = str(faktur.get('metode_pembayaran') or '').upper() == 'CASH'
            initial_status = UtangSupplier.STATUS_LUNAS if is_cash else UtangSupplier.STATUS_BELUM_DIBAYAR

            utang = UtangSupplier.objects.create(
                app_siaga_faktur_id=str(faktur['app_siaga_faktur_id']),
                sumber=UtangSupplier.SUMBER_LOGISTIK,
                nomor_spb=faktur.get('nomor_spb') or '',
                tanggal_spb=faktur.get('tanggal_spb'),
                nomor_faktur=faktur.get('nomor_faktur') or '',
                vendor_id=vendor_id_int,
                vendor_nama=vendor_nama_rek,
                tanggal_faktur=faktur.get('tanggal_faktur'),
                tanggal_jatuh_tempo=None,  # tran_beli_brg_log tidak punya jatuh tempo
                nominal=faktur.get('nominal') or 0,
                tanggal_titip=request.data.get('tanggal_titip') or timezone.localdate(),
                keterangan_titip=request.data.get('keterangan_titip') or '',
                status=initial_status,
                verified_by=request.user,
                verified_at=timezone.now(),
            )

        return Response(UtangSupplierSerializer(utang, context={'request': request}).data, status=status.HTTP_201_CREATED)

def _normalize_vendor_name_key(name):
    if not name:
        return ""
    s = str(name).upper().strip()
    s = re.sub(r'\b(PT|CV|UD|PD|NV|TBK)\b', '', s)
    return re.sub(r'[^A-Z0-9]', '', s)

class UtangPelunasanDataLamaView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        selected_items = request.data.get('items')
        selected_map = {}
        if selected_items and isinstance(selected_items, list):
            for item in selected_items:
                if isinstance(item, dict) and item.get('app_siaga_faktur_id'):
                    key = (str(item.get('app_siaga_faktur_id')), (item.get('sumber') or 'farmasi').strip())
                    selected_map[key] = True

        with connection.cursor() as cursor:
            # Build max tanggal_faktur LUNAS mapping by both vendor_id AND normalized vendor_nama
            cursor.execute("""
                SELECT vendor_id, vendor_nama, MAX(tanggal_faktur) as max_faktur
                FROM keuangan_utang_supplier
                WHERE tanggal_faktur IS NOT NULL AND status = 'lunas'
                GROUP BY vendor_id, vendor_nama
            """)
            vendor_max_faktur_by_id = {}
            vendor_max_faktur_by_name = {}

            for r in cursor.fetchall():
                v_id, v_name, max_faktur = r
                if v_id:
                    if v_id not in vendor_max_faktur_by_id or max_faktur > vendor_max_faktur_by_id[v_id]:
                        vendor_max_faktur_by_id[v_id] = max_faktur
                norm_name = _normalize_vendor_name_key(v_name)
                if norm_name:
                    if norm_name not in vendor_max_faktur_by_name or max_faktur > vendor_max_faktur_by_name[norm_name]:
                        vendor_max_faktur_by_name[norm_name] = max_faktur

            # 1. Fetch unverified farmasi purchases
            cursor.execute("""
                SELECT 
                    CONVERT(t.id USING utf8mb4) AS app_siaga_faktur_id,
                    CONVERT(COALESCE(NULLIF(t.no_spb, ''), t.id) USING utf8mb4) AS nomor_spb,
                    t.tgl_faktur AS tanggal_spb,
                    CONVERT(COALESCE(NULLIF(t.no_faktur, ''), '-') USING utf8mb4) AS nomor_faktur,
                    r.id_rekanan AS vendor_id,
                    CONVERT(COALESCE(r.nama, '') USING utf8mb4) AS vendor_nama,
                    t.tgl_faktur AS tanggal_faktur,
                    t.gtotal AS nominal
                FROM rssams.tran_beli_brg_farmasi t
                LEFT JOIN rssams.rekanan r ON r.id_rekanan = t.id_rekanan
                LEFT JOIN keuangan_utang_supplier u ON u.app_siaga_faktur_id = t.id
                WHERE u.id IS NULL 
                  AND (t.id NOT IN (SELECT nomor_spb FROM keuangan_utang_supplier WHERE nomor_spb != ''))
                  AND (t.no_faktur IS NULL OR t.no_faktur = '' OR t.no_faktur NOT IN (SELECT nomor_faktur FROM keuangan_utang_supplier WHERE nomor_faktur != ''))
            """)
            cols_f = [col[0] for col in cursor.description]
            farmasi_rows = [dict(zip(cols_f, row)) for row in cursor.fetchall()]

            # 2. Fetch unverified logistik purchases
            cursor.execute("""
                SELECT 
                    CONVERT(t.id USING utf8mb4) AS app_siaga_faktur_id,
                    CONVERT(COALESCE(NULLIF(s.no_spb, ''), NULLIF(t.id_spb, ''), t.id) USING utf8mb4) AS nomor_spb,
                    t.tgl_spk AS tanggal_spb,
                    CONVERT(COALESCE(NULLIF(NULLIF(t.no_spk, ''), '-'), '-') USING utf8mb4) AS nomor_faktur,
                    r.id_rekanan AS vendor_id,
                    CONVERT(COALESCE(r.nama, t.rekanan) USING utf8mb4) AS vendor_nama,
                    t.tgl_spk AS tanggal_faktur,
                    t.nilai AS nominal
                FROM rssams.tran_beli_brg_log t
                LEFT JOIN rssams.logistik_spb s ON s.id = t.id_spb
                LEFT JOIN rssams.rekanan r ON (r.nama = t.rekanan OR r.nama = s.no_spb)
                LEFT JOIN keuangan_utang_supplier u ON u.app_siaga_faktur_id = CONCAT('LOG-', t.id)
                WHERE t.done = 'Y' 
                  AND u.id IS NULL 
                  AND COALESCE(t.rekanan, '') != 'STOCK OPNAME'
                  AND COALESCE(t.no_spk, '') NOT LIKE 'OPNAME-%%'
                  AND (t.id NOT IN (SELECT nomor_spb FROM keuangan_utang_supplier WHERE nomor_spb != ''))
                  AND (s.no_spb IS NULL OR s.no_spb = '' OR s.no_spb NOT IN (SELECT nomor_spb FROM keuangan_utang_supplier WHERE nomor_spb != ''))
                  AND (t.no_spk IS NULL OR t.no_spk = '' OR t.no_spk NOT IN (SELECT nomor_faktur FROM keuangan_utang_supplier WHERE nomor_faktur != ''))
            """)
            cols_l = [col[0] for col in cursor.description]
            logistik_rows = [dict(zip(cols_l, row)) for row in cursor.fetchall()]

        now = timezone.now()
        today = timezone.localdate()
        total_nominal = Decimal('0')

        to_create = []
        for r in farmasi_rows:
            f_id = str(r['app_siaga_faktur_id'])
            v_id = r.get('vendor_id')
            v_nama = r.get('vendor_nama') or ''
            norm_v_nama = _normalize_vendor_name_key(v_nama)
            tgl_f = r.get('tanggal_faktur')
            
            is_old = False
            max_faktur_ref = None
            if selected_map:
                if (f_id, 'farmasi') in selected_map:
                    is_old = True
            else:
                max_faktur = vendor_max_faktur_by_id.get(v_id)
                norm_max_faktur = vendor_max_faktur_by_name.get(norm_v_nama)
                if norm_max_faktur and (not max_faktur or norm_max_faktur > max_faktur):
                    max_faktur = norm_max_faktur

                max_faktur_ref = max_faktur
                if max_faktur:
                    if tgl_f is None or tgl_f <= max_faktur:
                        is_old = True
                else:
                    if tgl_f is None or tgl_f.year < 2026:
                        is_old = True

            if is_old:
                nom = Decimal(str(r.get('nominal') or 0))
                ket = 'Dipelunaskan otomatis (Pilihan manual massal)' if selected_map else f'Dipelunaskan otomatis (Sisa data lama sebelum tanggal faktur OTS vendor: {max_faktur_ref or "< 2026"})'
                to_create.append(UtangSupplier(
                    app_siaga_faktur_id=f_id,
                    sumber=UtangSupplier.SUMBER_FARMASI,
                    nomor_spb=r.get('nomor_spb') or '',
                    tanggal_spb=r.get('tanggal_spb'),
                    nomor_faktur=r.get('nomor_faktur') or '',
                    vendor_id=v_id or 0,
                    vendor_nama=v_nama,
                    tanggal_faktur=tgl_f,
                    nominal=nom,
                    status=UtangSupplier.STATUS_LUNAS,
                    tanggal_titip=tgl_f or today,
                    keterangan_titip=ket,
                    verified_by=request.user,
                    verified_at=now
                ))
                total_nominal += nom

        for r in logistik_rows:
            f_id = str(r['app_siaga_faktur_id'])
            v_id = r.get('vendor_id')
            v_nama = r.get('vendor_nama') or ''
            norm_v_nama = _normalize_vendor_name_key(v_nama)
            tgl_f = r.get('tanggal_faktur')
            
            is_old = False
            max_faktur_ref = None
            if selected_map:
                if (f_id, 'logistik') in selected_map or (f"LOG-{f_id}", 'logistik') in selected_map:
                    is_old = True
            else:
                max_faktur = vendor_max_faktur_by_id.get(v_id)
                norm_max_faktur = vendor_max_faktur_by_name.get(norm_v_nama)
                if norm_max_faktur and (not max_faktur or norm_max_faktur > max_faktur):
                    max_faktur = norm_max_faktur

                max_faktur_ref = max_faktur
                if max_faktur:
                    if tgl_f is None or tgl_f <= max_faktur:
                        is_old = True
                else:
                    if tgl_f is None or tgl_f.year < 2026:
                        is_old = True

            if is_old:
                nom = Decimal(str(r.get('nominal') or 0))
                app_id = f"LOG-{f_id}"
                ket = 'Dipelunaskan otomatis (Pilihan manual massal)' if selected_map else f'Dipelunaskan otomatis (Sisa data lama sebelum tanggal faktur OTS vendor: {max_faktur_ref or "< 2026"})'
                to_create.append(UtangSupplier(
                    app_siaga_faktur_id=app_id,
                    sumber=UtangSupplier.SUMBER_LOGISTIK,
                    nomor_spb=r.get('nomor_spb') or '',
                    tanggal_spb=r.get('tanggal_spb'),
                    nomor_faktur=r.get('nomor_faktur') or '',
                    vendor_id=v_id or 0,
                    vendor_nama=v_nama,
                    tanggal_faktur=tgl_f,
                    nominal=nom,
                    status=UtangSupplier.STATUS_LUNAS,
                    tanggal_titip=tgl_f or today,
                    keterangan_titip=ket,
                    verified_by=request.user,
                    verified_at=now
                ))
                total_nominal += nom

        UtangSupplier.objects.bulk_create(to_create, batch_size=100, ignore_conflicts=True)
        return Response({
            'success': True,
            'count': len(to_create),
            'total_nominal': float(total_nominal),
            'message': f'Berhasil melunaskan {len(to_create)} faktur sisa lama (berdasarkan tanggal faktur OTS masing-masing vendor) dengan total nominal Rp {total_nominal:,.0f}.'
        })

    def delete(self, request):
        qs = UtangSupplier.objects.filter(
            models.Q(keterangan_titip__icontains='Dipelunaskan otomatis') |
            models.Q(keterangan_titip__icontains='Dipelutaskan otomatis')
        )
        count = qs.count()
        total_nominal = sum((u.nominal or Decimal('0')) for u in qs)
        qs.delete()
        return Response({
            'success': True,
            'count': count,
            'total_nominal': float(total_nominal),
            'message': f'Berhasil membatalkan (Undo) pelunasan {count} faktur sisa data lama. Data dikembalikan ke status Menunggu Verifikasi.'
        })

class UtangVendorOptionsView(APIView):
    permission_classes = [IsAuthenticated, IsCatatanUtangObatBhpPermission]

    def get(self, request):
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id_rekanan AS id, nama
                FROM rssams.rekanan
                WHERE del = 'N'
                ORDER BY nama
                LIMIT 500
                """
            )
            columns = [col[0] for col in cursor.description]
            return Response([dict(zip(columns, row)) for row in cursor.fetchall()])

class IndukPembiayaanViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset = IndukPembiayaan.objects.prefetch_related('anggota').all()
    serializer_class = IndukPembiayaanSerializer
    permission_classes = [IsKeuanganOrManajerPermission]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def get_queryset(self):
        qs = super().get_queryset()
        search = (self.request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(Q(nama__icontains=search) | Q(kode__icontains=search))
        return qs.order_by('nama')

    @action(detail=True, methods=['get'], url_path='anggota')
    def list_anggota(self, request, pk=None):
        induk = self.get_object()
        anggota = induk.anggota.all()
        return Response(PembiayaanIndukMappingSerializer(anggota, many=True).data)

    @action(detail=True, methods=['post'], url_path='tambah-anggota')
    def tambah_anggota(self, request, pk=None):
        induk = self.get_object()
        id_pembiayaan_list = request.data.get('id_pembiayaan_list') or []
        if not id_pembiayaan_list and request.data.get('id_pembiayaan'):
            id_pembiayaan_list = [request.data.get('id_pembiayaan')]

        if not id_pembiayaan_list:
            return Response({'error': 'Pilih setidaknya satu pembiayaan untuk ditambahkan.'}, status=status.HTTP_400_BAD_REQUEST)

        # Ambil nama pembiayaan dari rssams.pbiaya
        str_ids = [str(x) for x in id_pembiayaan_list]
        with connection.cursor() as cursor:
            placeholders = ','.join(['%s'] * len(str_ids))
            cursor.execute(f"SELECT id_pembiayaan, pembiayaan FROM rssams.pbiaya WHERE id_pembiayaan IN ({placeholders})", str_ids)
            pbiaya_dict = {str(row[0]): row[1] for row in cursor.fetchall()}

        created_count = 0
        updated_count = 0
        with transaction.atomic():
            for raw_id in str_ids:
                nama = pbiaya_dict.get(raw_id) or f"Pembiayaan ID {raw_id}"
                mapping, created = PembiayaanIndukMapping.objects.update_or_create(
                    id_pembiayaan=raw_id,
                    defaults={
                        'induk': induk,
                        'nama_pembiayaan': nama,
                        'created_by': request.user,
                    }
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1

        return Response({
            'message': f'Berhasil menambahkan {created_count + updated_count} pembiayaan ke {induk.nama}.',
            'created_count': created_count,
            'updated_count': updated_count,
            'induk': IndukPembiayaanSerializer(induk).data,
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='keluarkan-anggota')
    def keluarkan_anggota(self, request, pk=None):
        induk = self.get_object()
        id_pembiayaan = str(request.data.get('id_pembiayaan') or '').strip()
        if not id_pembiayaan:
            return Response({'error': 'id_pembiayaan wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        deleted, _ = PembiayaanIndukMapping.objects.filter(induk=induk, id_pembiayaan=id_pembiayaan).delete()
        if deleted:
            return Response({
                'message': f'Pembiayaan ID {id_pembiayaan} berhasil dikeluarkan dari {induk.nama}.',
                'induk': IndukPembiayaanSerializer(induk).data,
            }, status=status.HTTP_200_OK)
        return Response({'error': 'Pembiayaan tidak ditemukan di induk ini.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'], url_path='auto-group')
    def auto_group(self, request):
        """
        Saran pengelompokan otomatis berdasarkan nama awalan/pola umum:
        ADMEDIKA, ISOMEDIK, FHI (FULLERTON), OWLEXA, TPA, HALODOC, TMC, MAG, S.O-ADMEDIKA, BPJS, dll.
        """
        apply_now = str(request.data.get('apply') or '').lower() in ('1', 'true', 'yes')

        AUTO_PATTERNS = [
            ('ADMEDIKA', ['ADMEDIKA', 'S.O-ADMEDIKA', 'SO-ADMEDIKA']),
            ('ISOMEDIK', ['ISOMEDIK']),
            ('FULLERTON HEALTH (FHI)', ['FHI', 'FULLERTON', 'FULLERTHON']),
            ('OWLEXA HEALTHCARE', ['OWLEXA']),
            ('TPA (THIRD PARTY)', ['TPA']),
            ('HALODOC', ['HALODOC']),
            ('TMC INDONESIA', ['TMC']),
            ('MAG (MULTI ARTHA GUNA)', ['MAG']),
            ('MANDIRI INHEALTH', ['INHEALTH', 'MANDIRI INHEALTH']),
            ('BPJS KESEHATAN', ['BPJS KESEHATAN']),
            ('BPJS KETENAGAKERJAAN', ['BPJS KETENAGAKERJAAN', 'BPJS NAKER', 'BPJS-TK']),
        ]

        with connection.cursor() as cursor:
            cursor.execute("SELECT id_pembiayaan, pembiayaan FROM rssams.pbiaya WHERE status = 1 ORDER BY pembiayaan")
            all_pbiaya = cursor.fetchall()

        suggestions = []
        for induk_name, patterns in AUTO_PATTERNS:
            matching_items = []
            for id_p, name in all_pbiaya:
                clean_name = str(name or '').strip().upper()
                for pat in patterns:
                    if clean_name.startswith(pat) or f" {pat} " in f" {clean_name} " or f"-{pat}" in clean_name or f"{pat}-" in clean_name:
                        matching_items.append({'id_pembiayaan': str(id_p), 'nama_pembiayaan': name})
                        break
            if matching_items:
                suggestions.append({
                    'induk_nama': induk_name,
                    'total_match': len(matching_items),
                    'items': matching_items,
                })

        if not apply_now:
            return Response({
                'suggestions': suggestions,
                'total_groups': len(suggestions),
                'total_pembiayaan_matched': sum(s['total_match'] for s in suggestions),
            })

        # Apply changes
        total_assigned = 0
        with transaction.atomic():
            for group in suggestions:
                induk, _ = IndukPembiayaan.objects.get_or_create(
                    nama=group['induk_nama'],
                    defaults={'created_by': request.user}
                )
                for item in group['items']:
                    PembiayaanIndukMapping.objects.update_or_create(
                        id_pembiayaan=item['id_pembiayaan'],
                        defaults={
                            'induk': induk,
                            'nama_pembiayaan': item['nama_pembiayaan'],
                            'created_by': request.user,
                        }
                    )
                    total_assigned += 1

        return Response({
            'message': f'Berhasil mengelompokkan {total_assigned} pembiayaan ke dalam {len(suggestions)} Induk Pembiayaan.',
            'total_assigned': total_assigned,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='set-anggota-induk')
    def set_anggota_induk(self, request):
        """Menetapkan atau melepaskan induk untuk satu pembiayaan secara cepat"""
        id_pembiayaan = str(request.data.get('id_pembiayaan') or '').strip()
        induk_id = request.data.get('induk_id')

        if not id_pembiayaan:
            return Response({'error': 'id_pembiayaan wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        with connection.cursor() as cursor:
            cursor.execute("SELECT pembiayaan FROM rssams.pbiaya WHERE id_pembiayaan = %s LIMIT 1", [id_pembiayaan])
            row = cursor.fetchone()
            nama = row[0] if row else f"Pembiayaan ID {id_pembiayaan}"

        if not induk_id:
            # Unassign
            PembiayaanIndukMapping.objects.filter(id_pembiayaan=id_pembiayaan).delete()
            return Response({'message': f'Pembiayaan {nama} kini berdiri mandiri (tanpa induk).', 'induk_id': None, 'induk_nama': None})

        try:
            induk = IndukPembiayaan.objects.get(pk=induk_id)
        except IndukPembiayaan.DoesNotExist:
            return Response({'error': 'Induk Pembiayaan tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)

        mapping, _ = PembiayaanIndukMapping.objects.update_or_create(
            id_pembiayaan=id_pembiayaan,
            defaults={
                'induk': induk,
                'nama_pembiayaan': nama,
                'created_by': request.user,
            }
        )
        return Response({
            'message': f'Pembiayaan {nama} berhasil dimasukkan ke {induk.nama}.',
            'induk_id': induk.id,
            'induk_nama': induk.nama,
        })

class AlokasiDanaViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset           = AlokasiDana.objects.select_related('created_by', 'induk_pembiayaan').prefetch_related('pembayaran__faktur', 'pembayaran__created_by').all()
    serializer_class  = AlokasiDanaSerializer
    permission_classes = [IsKeuanganPermission]

    def perform_create(self, serializer):
        is_induk = serializer.validated_data.get('is_induk', False)
        induk = serializer.validated_data.get('induk_pembiayaan')
        id_pembiayaan = serializer.validated_data.get('id_pembiayaan') or ''
        nama_pembiayaan = serializer.validated_data.get('nama_pembiayaan') or ''
        if is_induk and induk:
            nama_pembiayaan = induk.nama
            if not id_pembiayaan:
                id_pembiayaan = f"INDUK-{induk.id}"
        serializer.save(
            created_by=self.request.user,
            id_pembiayaan=id_pembiayaan,
            nama_pembiayaan=nama_pembiayaan,
        )

    def get_queryset(self):
        qs          = super().get_queryset()
        id_pbiaya   = self.request.query_params.get('id_pembiayaan')
        induk_id    = self.request.query_params.get('induk_pembiayaan')
        is_induk    = self.request.query_params.get('is_induk')
        dari        = self.request.query_params.get('dari')
        sampai      = self.request.query_params.get('sampai')
        bank        = self.request.query_params.get('bank')
        search      = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(nama_pembiayaan__icontains=search) | Q(keterangan__icontains=search))
        if id_pbiaya:   qs = qs.filter(id_pembiayaan=id_pbiaya)
        if induk_id:    qs = qs.filter(induk_pembiayaan_id=induk_id)
        if is_induk is not None:
            qs = qs.filter(is_induk=(str(is_induk).lower() in ('1', 'true', 'yes')))
        if dari:        qs = qs.filter(tanggal_penerimaan__gte=dari)
        if sampai:      qs = qs.filter(tanggal_penerimaan__lte=sampai)
        if bank:        qs = qs.filter(bank=bank)
        return qs.order_by('-tanggal_penerimaan')

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.digunakan > 0:
            return Response({'error': 'Alokasi yang sudah dipakai tidak bisa dihapus.'}, status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)

class TagihanViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset           = Tagihan.objects.select_related('pemasok', 'created_by').prefetch_related('items', 'pembayaran__akun').all()
    permission_classes = [IsManajerOrAbovePermission]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return TagihanInputSerializer
        return TagihanSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def get_queryset(self):
        qs      = super().get_queryset()
        pemasok = self.request.query_params.get('pemasok')
        st      = self.request.query_params.get('status')
        dari    = self.request.query_params.get('dari')
        sampai  = self.request.query_params.get('sampai')
        if pemasok: qs = qs.filter(pemasok_id=pemasok)
        if st:      qs = qs.filter(status=st)
        if dari:    qs = qs.filter(tanggal__gte=dari)
        if sampai:  qs = qs.filter(tanggal__lte=sampai)
        return qs

    @action(detail=True, methods=['post'], url_path='terima')
    def terima(self, request, pk=None):
        tagihan = self.get_object()
        if tagihan.status != 'draft':
            return Response({'error': 'Hanya tagihan draft yang bisa diterima.'}, status=status.HTTP_400_BAD_REQUEST)
        tagihan.status = 'diterima'
        tagihan.save()
        return Response({'message': 'Tagihan berhasil diterima.'})

    @action(detail=True, methods=['post'], url_path='bayar')
    def bayar(self, request, pk=None):
        tagihan = self.get_object()
        if tagihan.status in ['lunas', 'batal']:
            return Response({'error': 'Tagihan sudah lunas atau dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = PembayaranTagihanInputSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        jumlah = serializer.validated_data['jumlah']
        if jumlah > tagihan.sisa_tagihan:
            return Response({'error': f'Jumlah bayar melebihi sisa tagihan ({tagihan.sisa_tagihan}).'}, status=status.HTTP_400_BAD_REQUEST)
        pembayaran = PembayaranTagihan.objects.create(
            tagihan=tagihan, created_by=request.user,
            tanggal=serializer.validated_data['tanggal'],
            jumlah=jumlah, metode=serializer.validated_data['metode'],
            keterangan=serializer.validated_data.get('keterangan', ''),
            akun=serializer.validated_data['akun'],
        )
        tagihan.total_dibayar += jumlah
        tagihan.status = 'lunas' if tagihan.total_dibayar >= tagihan.total_tagihan else 'sebagian'
        tagihan.save()
        return Response(PembayaranTagihanSerializer(pembayaran).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='batal')
    def batal(self, request, pk=None):
        tagihan = self.get_object()
        if tagihan.status == 'lunas':
            return Response({'error': 'Tagihan yang sudah lunas tidak bisa dibatalkan.'}, status=status.HTTP_400_BAD_REQUEST)
        tagihan.status = 'batal'
        tagihan.save()
        return Response({'message': 'Tagihan berhasil dibatalkan.'})

class RekeningBankViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    queryset           = RekeningBank.objects.prefetch_related('riwayat__updated_by').select_related('updated_by').all()
    permission_classes = [IsManajerOrAbovePermission]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return RekeningBankInputSerializer
        return RekeningBankSerializer

    def perform_create(self, serializer):
        serializer.save(updated_by=self.request.user)

    def get_queryset(self):
        qs   = super().get_queryset()
        bank = self.request.query_params.get('bank')
        if bank: qs = qs.filter(bank=bank)
        return qs

    @action(detail=True, methods=['post'], url_path='update-saldo')
    def update_saldo(self, request, pk=None):
        rekening   = self.get_object()
        serializer = UpdateSaldoSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        saldo_baru = serializer.validated_data['saldo_baru']
        keterangan = serializer.validated_data.get('keterangan', '')
        with transaction.atomic():
            saldo_sebelum      = rekening.saldo
            selisih            = saldo_baru - saldo_sebelum
            rekening.saldo     = saldo_baru
            rekening.updated_by = request.user
            rekening.save()
            RiwayatSaldoRekening.objects.create(
                rekening=rekening, saldo_sebelum=saldo_sebelum,
                saldo_sesudah=saldo_baru, selisih=selisih,
                keterangan=keterangan, updated_by=request.user,
            )
        return Response(RekeningBankSerializer(rekening, context={'request': request}).data)

class PettyCashViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return PettyCashInputSerializer
        return PettyCashSerializer

    def get_queryset(self):
        qs = PettyCash.objects.select_related('created_by', 'disetujui_oleh', 'dicairkan_oleh', 'laporan_disetujui_oleh').prefetch_related('laporan').all()
        
        # Manajer ke atas dan petugas kas petty cash bisa lihat semua. User biasa hanya milik sendiri.
        if not (is_manajer_or_above(self.request.user) or is_petty_cash_cashier(self.request.user)):
            qs = qs.filter(created_by=self.request.user)
        
        st     = self.request.query_params.get('status')
        dari   = self.request.query_params.get('dari')
        sampai = self.request.query_params.get('sampai')
        if st:     qs = qs.filter(status=st)
        if dari:   qs = qs.filter(tanggal__gte=dari)
        if sampai: qs = qs.filter(tanggal__lte=sampai)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def get_serializer_context(self):
        return {'request': self.request}

    # POST /{id}/approval/ — manajer/direktur setujui atau tolak
    @action(detail=True, methods=['post'], url_path='approval')
    def approval(self, request, pk=None):
        if not is_direktur_or_wadir(request.user):
            return Response({'error': 'Hanya direktur atau wakil direktur yang dapat memproses approval.'}, status=403)
        instance = self.get_object()
        if instance.status != 'pending':
            return Response({'error': 'Hanya pengajuan berstatus pending yang dapat diproses.'}, status=400)
        aksi   = request.data.get('aksi')
        catatan = request.data.get('catatan_tolak', '')
        if aksi not in ('setujui', 'tolak'):
            return Response({'error': 'aksi harus setujui atau tolak.'}, status=400)
        if aksi == 'setujui':
            saldo = get_or_create_saldo()
            if saldo.saldo < instance.nominal:
                return Response({
                    'error': f'Saldo tidak mencukupi. Saldo saat ini: Rp {saldo.saldo:,.0f}, nominal pengajuan: Rp {instance.nominal:,.0f}.'
                }, status=400)
            instance.status         = 'disetujui'
            instance.disetujui_oleh = request.user
            instance.catatan_tolak  = ''
        else:
            if not catatan:
                return Response({'error': 'Catatan tolak wajib diisi.'}, status=400)
            instance.status       = 'ditolak'
            instance.catatan_tolak = catatan
        instance.save()
        return Response({'message': f'Pengajuan berhasil {"disetujui" if aksi == "setujui" else "ditolak"}.', 'status': instance.status})

    # POST /{id}/cairkan/ — petugas kas petty cash mencairkan dana
    @action(detail=True, methods=['post'], url_path='cairkan')
    def cairkan(self, request, pk=None):
        if not is_petty_cash_cashier(request.user):
            return Response({'error': 'Hanya petugas kas petty cash yang dapat mencairkan dana.'}, status=403)
        instance = self.get_object()
        if instance.status != 'disetujui':
            return Response({'error': 'Hanya pengajuan berstatus disetujui yang dapat dicairkan.'}, status=400)
        instance.status        = 'dicairkan'
        instance.dicairkan_oleh = request.user
        instance.save()
        return Response({'message': 'Dana berhasil dicairkan.', 'status': instance.status})

    # POST /{id}/laporan/ — karyawan upload nota + rincian
    @action(detail=True, methods=['post'], url_path='laporan', parser_classes=[MultiPartParser, FormParser, JSONParser])
    def laporan(self, request, pk=None):
        instance = self.get_object()
        if not is_direktur_or_wadir(request.user) and instance.created_by != request.user:
            return Response({'error': 'Laporan penggunaan hanya dapat disubmit oleh pemohon.'}, status=403)
        if instance.status != 'dicairkan':
            return Response({'error': 'Laporan hanya bisa disubmit setelah dana dicairkan.'}, status=400)
        if hasattr(instance, 'laporan'):
            return Response({'error': 'Laporan sudah pernah disubmit.'}, status=400)

        serializer = LaporanPenggunaanInputSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        nominal_dicairkan = instance.nominal
        nominal_digunakan = serializer.validated_data['nominal_digunakan']
        if nominal_digunakan > nominal_dicairkan:
            return Response({
                'error': f'Nominal digunakan tidak boleh melebihi dana yang dicairkan. Maksimal Rp {nominal_dicairkan:,.0f}.'
            }, status=400)
        selisih           = nominal_dicairkan - nominal_digunakan

        laporan = serializer.save(petty_cash=instance, selisih=selisih)

        instance.status = 'menunggu_approval_laporan'
        instance.catatan_tolak = ''
        instance.laporan_disetujui_oleh = None
        instance.laporan_disetujui_at = None
        instance.save()

        return Response(LaporanPenggunaanSerializer(laporan, context={'request': request}).data, status=status.HTTP_201_CREATED)

    # POST /{id}/approval-laporan/ — wadir/direktur approve laporan penggunaan
    @action(detail=True, methods=['post'], url_path='approval-laporan')
    def approval_laporan(self, request, pk=None):
        if not is_direktur_or_wadir(request.user):
            return Response({'error': 'Hanya direktur atau wakil direktur yang dapat approve laporan penggunaan.'}, status=403)
        instance = self.get_object()
        if instance.status != 'menunggu_approval_laporan':
            return Response({'error': 'Hanya laporan berstatus menunggu approval yang dapat diproses.'}, status=400)
        if not hasattr(instance, 'laporan'):
            return Response({'error': 'Laporan penggunaan belum tersedia.'}, status=400)

        aksi = request.data.get('aksi', 'setujui')
        catatan = request.data.get('catatan_tolak', '')
        if aksi not in ('setujui', 'tolak'):
            return Response({'error': 'aksi harus setujui atau tolak.'}, status=400)

        laporan = instance.laporan
        if aksi == 'tolak':
            if not catatan:
                return Response({'error': 'Catatan tolak wajib diisi.'}, status=400)
            with transaction.atomic():
                laporan.delete()
                instance.status = 'dicairkan'
                instance.catatan_tolak = catatan
                instance.laporan_disetujui_oleh = None
                instance.laporan_disetujui_at = None
                instance.save()
            return Response({
                'message': 'Laporan penggunaan ditolak. User dapat upload laporan ulang.',
                'status': instance.status,
            }, status=status.HTTP_200_OK)

        with transaction.atomic():
            instance.status = 'menunggu_pengembalian' if laporan.selisih > 0 else 'dilaporkan'
            instance.catatan_tolak = ''
            instance.laporan_disetujui_oleh = request.user
            instance.laporan_disetujui_at = timezone.now()
            instance.save()

        return Response({
            'message': 'Laporan penggunaan berhasil disetujui.',
            'status': instance.status,
        }, status=status.HTTP_200_OK)

    # POST /{id}/konfirmasi-pengembalian/ — petugas kas petty cash konfirmasi uang kembali & selesaikan
    @action(detail=True, methods=['post'], url_path='konfirmasi-pengembalian')
    def konfirmasi_pengembalian(self, request, pk=None):
        if not is_petty_cash_cashier(request.user):
            return Response({'error': 'Hanya petugas kas petty cash yang dapat mengkonfirmasi pengembalian.'}, status=403)
        instance = self.get_object()
        if instance.status not in ('menunggu_pengembalian', 'dilaporkan'):
            return Response({'error': 'Status tidak valid untuk dikonfirmasi.'}, status=400)
 
        laporan = instance.laporan
        nominal_pakai = laporan.nominal_digunakan
 
        with transaction.atomic():
            saldo         = get_or_create_saldo()
            saldo_sebelum = saldo.saldo
            saldo.saldo  -= nominal_pakai
            saldo.updated_by = request.user
            saldo.save()
 
            RiwayatSaldoPettyCash.objects.create(
                jenis='pengurangan',
                jumlah=nominal_pakai,
                saldo_sebelum=saldo_sebelum,
                saldo_sesudah=saldo.saldo,
                keterangan=f'Realisasi petty cash {instance.no_pengajuan} - {instance.keperluan[:50]}',
                created_by=request.user,
                nama_pengaju=user_display_name(instance.created_by),
                unit_pengaju=laporan_unit_label(instance.created_by),
            )
 
            laporan.pengembalian_selesai = True
            laporan.dikonfirmasi_oleh    = request.user
            laporan.save()
 
            instance.status = 'selesai'
            instance.save()
 
        return Response({'message': 'Pengembalian dikonfirmasi. Petty cash selesai.', 'status': instance.status})

    # POST /{id}/revisi/ — karyawan revisi yang ditolak
    @action(detail=True, methods=['post'], url_path='revisi', parser_classes=[MultiPartParser, FormParser, JSONParser])
    def revisi(self, request, pk=None):
        instance = self.get_object()
        if not is_direktur_or_wadir(request.user) and instance.created_by != request.user:
            return Response({'error': 'Revisi hanya dapat dilakukan oleh pemohon.'}, status=403)
        if instance.status != 'ditolak':
            return Response({'error': 'Hanya pengajuan ditolak yang bisa direvisi.'}, status=400)
        serializer = PettyCashInputSerializer(instance, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save(status='pending', catatan_tolak='')
        return Response(PettyCashSerializer(instance, context={'request': request}).data)

    # POST /{id}/batal/ — pemohon / kasir / pimpinan membatalkan pengajuan di setiap tahapan
    @action(detail=True, methods=['post'], url_path='batal')
    def batal(self, request, pk=None):
        instance = self.get_object()
        can_cancel_all = is_direktur_or_wadir(request.user) or is_petty_cash_cashier(request.user) or request.user.is_superuser
        if not can_cancel_all:
            if instance.created_by != request.user:
                return Response({'error': 'Hanya pemohon, petugas kasir, atau pimpinan yang dapat membatalkan pengajuan ini.'}, status=403)
        
        if instance.status == 'dibatalkan':
            return Response({'error': 'Pengajuan ini sudah dibatalkan sebelumnya.'}, status=400)
            
        alasan = request.data.get('alasan_batal') or request.data.get('alasan') or request.data.get('catatan_tolak')
        if not alasan or not str(alasan).strip():
            return Response({'error': 'Alasan pembatalan wajib diisi.'}, status=400)
            
        with transaction.atomic():
            # Jika pengajuan dibatalkan setelah status 'selesai' (di mana saldo sudah dipotong saat konfirmasi), kembalikan saldo ke kas
            if instance.status == 'selesai' and hasattr(instance, 'laporan'):
                nominal_pakai = instance.laporan.nominal_digunakan
                if nominal_pakai > 0:
                    saldo = get_or_create_saldo()
                    saldo_sebelum = saldo.saldo
                    saldo.saldo += nominal_pakai
                    saldo.updated_by = request.user
                    saldo.save()
                    
                    RiwayatSaldoPettyCash.objects.create(
                        jenis='penambahan',
                        jumlah=nominal_pakai,
                        saldo_sebelum=saldo_sebelum,
                        saldo_sesudah=saldo.saldo,
                        keterangan=f'Koreksi pembatalan petty cash {instance.no_pengajuan} - {str(alasan).strip()[:50]}',
                        created_by=request.user,
                        nama_pengaju=user_display_name(instance.created_by),
                        unit_pengaju=laporan_unit_label(instance.created_by),
                    )

            instance.status = 'dibatalkan'
            instance.catatan_tolak = f"Dibatalkan: {str(alasan).strip()}"
            instance.save()

        return Response({
            'message': 'Pengajuan berhasil dibatalkan.',
            'status': instance.status,
            'data': PettyCashSerializer(instance, context={'request': request}).data
        })

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        can_delete_all = is_direktur_or_wadir(request.user)
        if not can_delete_all:
            if instance.created_by != request.user:
                return Response({'error': 'Hanya pemohon yang dapat menghapus pengajuan sendiri.'}, status=403)
            if instance.status != 'pending':
                return Response({'error': 'Hanya pengajuan pending yang dapat dihapus.'}, status=400)
        return super().destroy(request, *args, **kwargs)

class ReimbursementViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ReimbursementInputSerializer
        return ReimbursementSerializer

    def get_queryset(self):
        qs = Reimbursement.objects.select_related('created_by', 'disetujui_oleh', 'dicairkan_oleh').all()
        
        # User biasa hanya lihat milik sendiri
        if not is_manajer_or_above(self.request.user):
            qs = qs.filter(created_by=self.request.user)
        
        st     = self.request.query_params.get('status')
        dari   = self.request.query_params.get('dari')
        sampai = self.request.query_params.get('sampai')
        if st:     qs = qs.filter(status=st)
        if dari:   qs = qs.filter(tanggal__gte=dari)
        if sampai: qs = qs.filter(tanggal__lte=sampai)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def get_serializer_context(self):
        return {'request': self.request}

    # POST /{id}/approval/ — manajer/direktur setujui atau tolak
    @action(detail=True, methods=['post'], url_path='approval')
    def approval(self, request, pk=None):
        if not is_direktur_or_wadir(request.user):
            return Response({'error': 'Hanya direktur atau wakil direktur yang dapat memproses approval.'}, status=403)
        instance = self.get_object()
        if instance.status != 'pending':
            return Response({'error': 'Hanya pengajuan berstatus pending yang dapat diproses.'}, status=400)
        aksi    = request.data.get('aksi')
        catatan = request.data.get('catatan_tolak', '')
        if aksi not in ('setujui', 'tolak'):
            return Response({'error': 'aksi harus setujui atau tolak.'}, status=400)
        if aksi == 'setujui':
            instance.status         = 'disetujui'
            instance.disetujui_oleh = request.user
            instance.catatan_tolak  = ''
        else:
            if not catatan:
                return Response({'error': 'Catatan tolak wajib diisi.'}, status=400)
            instance.status        = 'ditolak'
            instance.catatan_tolak = catatan
        instance.save()
        return Response({'message': f'Reimbursement berhasil {"disetujui" if aksi == "setujui" else "ditolak"}.', 'status': instance.status})

    # POST /{id}/cairkan/ — manajer mencairkan
    @action(detail=True, methods=['post'], url_path='cairkan')
    def cairkan(self, request, pk=None):
        if not is_manajer_or_above(self.request.user):        
            return Response({'error': 'Hanya manajer, direktur, atau wakil direktur yang dapat mencairkan.'}, status=403)
        instance = self.get_object()
        if instance.status != 'disetujui':
            return Response({'error': 'Hanya reimbursement berstatus disetujui yang dapat dicairkan.'}, status=400)

        # Cek saldo mencukupi
        saldo = get_or_create_saldo()
        if saldo.saldo < instance.nominal:
            return Response({
                'error': f'Saldo tidak mencukupi. Saldo saat ini: Rp {saldo.saldo:,.0f}, dibutuhkan: Rp {instance.nominal:,.0f}.'
            }, status=400)

        with transaction.atomic():
            saldo_sebelum        = saldo.saldo
            saldo.saldo         -= instance.nominal
            saldo.updated_by     = request.user
            saldo.save()

            RiwayatSaldoPettyCash.objects.create(
                jenis='pengurangan',
                jumlah=instance.nominal,
                saldo_sebelum=saldo_sebelum,
                saldo_sesudah=saldo.saldo,
                keterangan=f'Reimbursement {instance.no_reimbursement} - {instance.keperluan[:50]}',
                created_by=request.user,
                nama_pengaju=user_display_name(instance.created_by),
                unit_pengaju=laporan_unit_label(instance.created_by),
            )
 
            instance.status         = 'dicairkan'
            instance.dicairkan_oleh = request.user
            instance.save()
 
        return Response({'message': 'Reimbursement berhasil dicairkan.', 'status': instance.status})

    # POST /{id}/revisi/ — karyawan revisi yang ditolak
    @action(detail=True, methods=['post'], url_path='revisi', parser_classes=[MultiPartParser, FormParser, JSONParser])
    def revisi(self, request, pk=None):
        instance = self.get_object()
        if not is_direktur_or_wadir(request.user) and instance.created_by != request.user:
            return Response({'error': 'Revisi hanya dapat dilakukan oleh pemohon.'}, status=403)
        if instance.status != 'ditolak':
            return Response({'error': 'Hanya reimbursement ditolak yang bisa direvisi.'}, status=400)
        serializer = ReimbursementInputSerializer(instance, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save(status='pending', catatan_tolak='')
        return Response(ReimbursementSerializer(instance, context={'request': request}).data)

    # POST /{id}/batal/ — pemohon / kasir / pimpinan membatalkan reimbursement di setiap tahapan
    @action(detail=True, methods=['post'], url_path='batal')
    def batal(self, request, pk=None):
        instance = self.get_object()
        can_cancel_all = is_direktur_or_wadir(request.user) or is_petty_cash_cashier(request.user) or request.user.is_superuser
        if not can_cancel_all:
            if instance.created_by != request.user:
                return Response({'error': 'Hanya pemohon, petugas kasir, atau pimpinan yang dapat membatalkan reimbursement ini.'}, status=403)
        
        if instance.status == 'dibatalkan':
            return Response({'error': 'Reimbursement ini sudah dibatalkan sebelumnya.'}, status=400)
            
        alasan = request.data.get('alasan_batal') or request.data.get('alasan') or request.data.get('catatan_tolak')
        if not alasan or not str(alasan).strip():
            return Response({'error': 'Alasan pembatalan wajib diisi.'}, status=400)
            
        with transaction.atomic():
            # Jika reimbursement sudah dicairkan (saldo sudah dipotong), kembalikan saldo ke kasir
            if instance.status == 'dicairkan':
                nominal = instance.nominal
                if nominal > 0:
                    saldo = get_or_create_saldo()
                    saldo_sebelum = saldo.saldo
                    saldo.saldo += nominal
                    saldo.updated_by = request.user
                    saldo.save()
                    
                    RiwayatSaldoPettyCash.objects.create(
                        jenis='penambahan',
                        jumlah=nominal,
                        saldo_sebelum=saldo_sebelum,
                        saldo_sesudah=saldo.saldo,
                        keterangan=f'Koreksi pembatalan reimbursement {instance.no_reimbursement} - {str(alasan).strip()[:50]}',
                        created_by=request.user,
                        nama_pengaju=user_display_name(instance.created_by),
                        unit_pengaju=laporan_unit_label(instance.created_by),
                    )

            instance.status = 'dibatalkan'
            instance.catatan_tolak = f"Dibatalkan: {str(alasan).strip()}"
            instance.save()

        return Response({
            'message': 'Reimbursement berhasil dibatalkan.',
            'status': instance.status,
            'data': ReimbursementSerializer(instance, context={'request': request}).data
        })

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        can_delete_all = is_direktur_or_wadir(request.user)
        if not can_delete_all:
            if instance.created_by != request.user:
                return Response({'error': 'Hanya pemohon yang dapat menghapus reimbursement sendiri.'}, status=403)
            if instance.status != 'pending':
                return Response({'error': 'Hanya reimbursement pending yang dapat dihapus.'}, status=400)
        return super().destroy(request, *args, **kwargs)

def get_or_create_saldo():
    saldo, _ = SaldoPettyCash.objects.get_or_create(pk=1, defaults={'saldo': 0})
    return saldo

class SaldoPettyCashViewSet(viewsets.ViewSet):
    permission_classes = [IsPettyCashSaldoPermission]
 
    # GET /api/keuangan/saldo-petty-cash/
    def list(self, request):
        saldo = get_or_create_saldo()
        riwayat = RiwayatSaldoPettyCash.objects.select_related(
            'created_by', 'created_by__unit'
        ).all()[:100]
        return Response({
            'saldo': SaldoPettyCashSerializer(saldo).data,
            'riwayat': RiwayatSaldoPettyCashSerializer(riwayat, many=True).data,
        })

class PengajuanPenambahanSaldoViewSet(OptionalPaginationMixin, viewsets.ModelViewSet):
    permission_classes = [IsManajerOrAbovePermission]
 
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return PengajuanPenambahanSaldoInputSerializer
        return PengajuanPenambahanSaldoSerializer
 
    def get_queryset(self):
        return PengajuanPenambahanSaldo.objects.select_related(
            'created_by', 'diproses_oleh'
        ).all()
 
    def perform_create(self, serializer):
        # Manajer keuangan ke atas bisa mengajukan penambahan saldo.
        if not is_manajer_or_above(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Hanya manajer, direktur, atau wakil direktur yang dapat mengajukan penambahan saldo.')
        serializer.save(created_by=self.request.user)
 
    # POST /{id}/approval/ — direktur/wadir approve atau tolak
    @action(detail=True, methods=['post'], url_path='approval')
    def approval(self, request, pk=None):
        if not is_direktur_or_wadir(request.user):
            return Response({'error': 'Hanya direktur atau wakil direktur yang dapat memproses.'}, status=403)
 
        instance = self.get_object()
        if instance.status != 'pending':
            return Response({'error': 'Hanya pengajuan pending yang dapat diproses.'}, status=400)
 
        aksi    = request.data.get('aksi')
        catatan = request.data.get('catatan_tolak', '')
 
        if aksi not in ('setujui', 'tolak'):
            return Response({'error': 'aksi harus setujui atau tolak.'}, status=400)
 
        if aksi == 'tolak':
            if not catatan:
                return Response({'error': 'Catatan tolak wajib diisi.'}, status=400)
            instance.status       = 'ditolak'
            instance.catatan_tolak = catatan
            instance.diproses_oleh = request.user
            instance.save()
            return Response({'message': 'Pengajuan ditolak.', 'status': instance.status})
 
        # Setujui — nominal diambil dari yang diajukan atau disesuaikan pimpinan
        nominal = request.data.get('nominal_diajukan') or instance.nominal_diajukan
        if not nominal:
            return Response({'error': 'Nominal pengisian kembali wajib diisi.'}, status=400)
 
        try:
            nominal = Decimal(str(nominal))
            if nominal <= 0:
                raise ValueError
        except Exception:
            return Response({'error': 'Nominal tidak valid.'}, status=400)
 
        with transaction.atomic():
            instance.status          = 'disetujui'
            instance.nominal_diajukan = nominal
            instance.diproses_oleh   = request.user
            instance.catatan_tolak   = ''
            instance.save()
 
        return Response({
            'message': f'Pengisian kembali saldo berhasil disetujui ({instance.no_pengajuan}). Otomatis masuk ke antrean Menunggu Verifikasi di Keuangan.',
            'status': instance.status,
        })

class LaporanPettyCashView(APIView):
    permission_classes = [IsAuthenticated]
 
    def get(self, request):
        if not is_manajer_or_above(request.user):
            return Response({'error': 'Akses ditolak.'}, status=403)
 
        dari   = request.query_params.get('dari')
        sampai = request.query_params.get('sampai')
 
        if not dari or not sampai:
            return Response({'error': 'Parameter dari dan sampai wajib diisi.'}, status=400)
 
        # ── Saldo awal — ambil snapshot saldo terakhir sebelum periode ──
        from datetime import date as date_type
        dari_date   = date_type.fromisoformat(dari)
        sampai_date = date_type.fromisoformat(sampai)

        dari_datetime = timezone.make_aware(
            datetime.combine(dari_date, time.min),
            timezone.get_current_timezone()
        )

        riwayat_terakhir_sebelum = RiwayatSaldoPettyCash.objects.filter(
            created_at__lt=dari_datetime
        ).order_by('-created_at').first()

        saldo_awal = riwayat_terakhir_sebelum.saldo_sesudah if riwayat_terakhir_sebelum else 0

        # ── Riwayat dalam periode ──

        sampai_datetime = timezone.make_aware(
            datetime.combine(sampai_date, time.max),
            timezone.get_current_timezone()
        )

        riwayat_periode = RiwayatSaldoPettyCash.objects.filter(
            created_at__gte=dari_datetime,
            created_at__lte=sampai_datetime,
        ).order_by('created_at')

        # print("dari_date:", dari_date)
        # print("riwayat_terakhir_sebelum:", riwayat_terakhir_sebelum)
        # print("total riwayat semua:", RiwayatSaldoPettyCash.objects.count())
        # if riwayat_terakhir_sebelum:
        #     print("saldo_sesudah:", riwayat_terakhir_sebelum.saldo_sesudah)
        #     print("created_at:", riwayat_terakhir_sebelum.created_at)

        # if riwayat_terakhir_sebelum:
        #     saldo_awal = riwayat_terakhir_sebelum.saldo_sesudah
        # else:
        #     # Hitung mundur dari saldo sekarang
        #     saldo_sekarang = get_or_create_saldo().saldo
        #     total_masuk_periode  = riwayat_periode.filter(jenis='penambahan').aggregate(t=Sum('jumlah'))['t'] or 0
        #     total_keluar_periode = riwayat_periode.filter(jenis='pengurangan').aggregate(t=Sum('jumlah'))['t'] or 0
        #     saldo_awal = saldo_sekarang - total_masuk_periode + total_keluar_periode
        
        # print("riwayat dalam periode:", list(riwayat_periode.values('jenis', 'jumlah', 'created_at')))
        # print("Semua riwayat:", list(RiwayatSaldoPettyCash.objects.values('jenis', 'jumlah', 'created_at').order_by('created_at')))

        total_penambahan  = riwayat_periode.filter(jenis='penambahan').aggregate(t=Sum('jumlah'))['t'] or 0
        total_pengurangan = riwayat_periode.filter(jenis='pengurangan').aggregate(t=Sum('jumlah'))['t'] or 0
        saldo_akhir       = saldo_awal + total_penambahan - total_pengurangan

        # ── Petty Cash (advance) dalam periode ──
        pc_qs = PettyCash.objects.filter(
            tanggal__gte=dari,
            tanggal__lte=sampai,
        ).select_related('created_by__unit', 'laporan')
 
        # ── Reimbursement dalam periode ──
        rb_qs = Reimbursement.objects.filter(
            tanggal__gte=dari,
            tanggal__lte=sampai,
        ).select_related('created_by__unit')
 
        # ── Daftar pengajuan (gabungan PC + Reimbursement) ──
        daftar_pengajuan = []
        for pc in pc_qs:
            # Ambil realisasi kalau sudah ada laporan
            nominal_realisasi = float(pc.laporan.nominal_digunakan) if hasattr(pc, 'laporan') and pc.laporan else None
            daftar_pengajuan.append({
                'no':       pc.no_pengajuan,
                'tanggal':  str(pc.tanggal),
                'jenis':    'Petty Cash',
                'pemohon':  pc.created_by.get_full_name() or pc.created_by.username if pc.created_by else '—',
                'unit':     laporan_unit_label(pc.created_by),
                'keperluan': pc.keperluan,
                'nominal':  float(pc.nominal),
                'nominal_realisasi': nominal_realisasi,
                # Tampilkan realisasi sebagai nominal efektif kalau sudah ada
                'nominal_efektif': nominal_realisasi if nominal_realisasi is not None else float(pc.nominal),
                'status':   pc.status,
            })
        for rb in rb_qs:
            daftar_pengajuan.append({
                'no':       rb.no_reimbursement,
                'tanggal':  str(rb.tanggal),
                'jenis':    'Reimbursement',
                'pemohon':  rb.created_by.get_full_name() or rb.created_by.username if rb.created_by else '—',
                'unit':     laporan_unit_label(rb.created_by),
                'keperluan': rb.keperluan,
                'nominal':  float(rb.nominal),
                'nominal_realisasi': float(rb.nominal) if rb.status == 'dicairkan' else None,
                'status':   rb.status,
            })
        daftar_pengajuan.sort(key=lambda x: x['tanggal'])
 
        # ── Total pencairan per unit ──
        per_unit = defaultdict(lambda: {'pc': 0, 'reimburse': 0, 'total': 0})
        for pc in pc_qs.filter(status__in=['dicairkan', 'dilaporkan', 'menunggu_pengembalian', 'selesai']):
            unit = laporan_unit_label(pc.created_by)
            nominal = float(pc.laporan.nominal_digunakan) if hasattr(pc, 'laporan') and pc.laporan else float(pc.nominal)
            per_unit[unit]['pc']    += nominal
            per_unit[unit]['total'] += nominal
        for rb in rb_qs.filter(status='dicairkan'):
            unit = laporan_unit_label(rb.created_by)
            per_unit[unit]['reimburse'] += float(rb.nominal)
            per_unit[unit]['total']     += float(rb.nominal)
 
        # ── Grafik per bulan (6 bulan terakhir) ──
        from datetime import date, timedelta
        today = date.today()
        grafik = []
        for i in range(5, -1, -1):
            # bulan ke-i sebelum bulan ini
            month = (today.month - i - 1) % 12 + 1
            year  = today.year + ((today.month - i - 1) // 12)
            last_day = calendar.monthrange(year, month)[1]
            bln_dari   = date(year, month, 1)
            bln_sampai = date(year, month, last_day)
            label = bln_dari.strftime('%b %y')
 
            pc_total = PettyCash.objects.filter(
                tanggal__gte=bln_dari, tanggal__lte=bln_sampai,
                status__in=['dicairkan','dilaporkan','menunggu_pengembalian','selesai']
            ).aggregate(t=Sum('nominal'))['t'] or 0
 
            rb_total = Reimbursement.objects.filter(
                tanggal__gte=bln_dari, tanggal__lte=bln_sampai,
                status='dicairkan'
            ).aggregate(t=Sum('nominal'))['t'] or 0
 
            grafik.append({
                'bulan':        label,
                'petty_cash':   float(pc_total),
                'reimbursement': float(rb_total),
                'total':        float(pc_total + rb_total),
            })
 
        # ── Rekap mutasi saldo ──
        rekap_mutasi = [{
            'waktu':      r.created_at.strftime('%d %b %Y %H:%M'),
            'jenis':      r.jenis,
            'jumlah':     float(r.jumlah),
            'saldo_sesudah': float(r.saldo_sesudah),
            'keterangan': r.keterangan,
        } for r in riwayat_periode]
 
        return Response({
            'periode':          {'dari': dari, 'sampai': sampai},
            'saldo_awal':       float(saldo_awal),
            'saldo_akhir':      float(saldo_akhir),
            'total_penambahan': float(total_penambahan),
            'total_pengurangan': float(total_pengurangan),
            'daftar_pengajuan': daftar_pengajuan,
            'per_unit':         [{'unit': k, **v} for k, v in per_unit.items()],
            'grafik':           grafik,
            'rekap_mutasi':     rekap_mutasi,
        })

def is_driver(user):
    return user.is_authenticated and getattr(user, 'is_driver', False)

def is_admin_driver(user):
    """Yang bisa lihat semua data driver dan kelola kendaraan"""
    return user.role in ('direktur', 'wakil_direktur', 'manajer') or user.is_superuser

def has_driver_access(user):
    return is_driver(user) or is_admin_driver(user)

def legacy_fetchall(sql, params=None):
    with connection.cursor() as cursor:
        cursor.execute(sql, params or [])
        cols = [col[0] for col in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]

def legacy_fetchone(sql, params=None):
    rows = legacy_fetchall(sql, params)
    return rows[0] if rows else None

def legacy_paginated(request, base_sql, count_sql, params=None):
    page = int(request.query_params.get('page') or 1)
    page_size = int(request.query_params.get('page_size') or 10)
    offset = (page - 1) * page_size
    count = legacy_fetchone(count_sql, params or [])['total']
    rows = legacy_fetchall(f'{base_sql} LIMIT %s OFFSET %s', [*(params or []), page_size, offset])
    return Response({'count': count, 'results': rows})

_rekanan_columns_cache = None

def _get_rekanan_columns():
    global _rekanan_columns_cache
    if _rekanan_columns_cache is not None:
        return _rekanan_columns_cache

    cols = set()
    table_name = "rssams.rekanan"
    with connection.cursor() as cursor:
        try:
            cursor.execute("SHOW COLUMNS FROM rssams.rekanan")
            cols = {row[0].lower() for row in cursor.fetchall()}
        except Exception:
            try:
                cursor.execute("SHOW COLUMNS FROM rekanan")
                cols = {row[0].lower() for row in cursor.fetchall()}
                table_name = "rekanan"
            except Exception:
                pass

        if cols:
            if 'sumber' not in cols:
                try:
                    cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN sumber VARCHAR(50) DEFAULT 'farmasi'")
                    cols.add('sumber')
                except Exception:
                    pass
            if 'kategori' not in cols:
                try:
                    cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN kategori VARCHAR(100) DEFAULT ''")
                    cols.add('kategori')
                except Exception:
                    pass

    _rekanan_columns_cache = cols
    return _rekanan_columns_cache

_dafbrg_log_columns_cache = None

def _refresh_pembelian_total(pembelian_id, no_invoice=None):
    from django.db import connection
    with connection.cursor() as cursor:
        if no_invoice is not None:
            cursor.execute(
                "UPDATE rssams.tran_beli_brg_log SET no_spk = %s WHERE id = %s",
                [no_invoice or '', pembelian_id],
            )
        cursor.execute(
            """
            UPDATE rssams.tran_beli_brg_log
            SET nilai = COALESCE((SELECT SUM(qty * harga) FROM rssams.item_logistik WHERE id = %s), 0)
            WHERE id = %s
            """,
            [pembelian_id, pembelian_id],
        )

def _refresh_spb_total(spb_id):
    from django.db import connection
    with connection.cursor() as cursor:
        cursor.execute(
            """
            UPDATE rssams.logistik_spb s
            SET nilai = COALESCE((
                SELECT SUM(qty * harga) FROM rssams.logistik_spb_item WHERE spb_id = s.id
            ), 0)
            WHERE id = %s
            """,
            [spb_id]
        )
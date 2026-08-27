from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Logbook, Task, SesiKerja
from .serializers import LogbookSerializer, LogbookInputSerializer, TaskSerializer, TaskCreateSerializer
from .utils import hitung_durasi_sesi


def get_monitoring_level(user):
    """Return: 'all' (direktur), 'unit' (kepala_seksi/manajer), atau None"""
    if not user or not user.is_authenticated:
        return None
    if user.is_superuser or user.role in ('direktur', 'wakil_direktur'):
        return 'all'
    if user.role in ('kepala_seksi', 'manajer'):
        return 'unit'
    return None


class LogbookViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    pagination_class = None

    def get_queryset(self):
        user = self.request.user
        qs = Logbook.objects.select_related('user', 'user__unit')
        monitoring_level = get_monitoring_level(user)

        # Filter akses data
        if monitoring_level is None:
            # Karyawan biasa hanya melihat logbook miliknya sendiri
            qs = qs.filter(user=user)
        elif monitoring_level == 'unit':
            # Manajer/Kepala Seksi hanya melihat unitnya sendiri
            qs = qs.filter(user__unit=user.unit)
        else:
            # Pimpinan (Direktur & Wadir) dapat memfilter berdasarkan user dan unit
            user_id = self.request.query_params.get('user_id')
            if user_id and user_id.isdigit():
                qs = qs.filter(user_id=int(user_id))

            unit_id = self.request.query_params.get('unit_id')
            if unit_id and unit_id.isdigit():
                qs = qs.filter(user__unit_id=int(unit_id))

        # Filter tanggal
        tanggal = self.request.query_params.get('tanggal')
        if tanggal:
            qs = qs.filter(tanggal=tanggal)

        start_date = self.request.query_params.get('start_date')
        if start_date:
            qs = qs.filter(tanggal__gte=start_date)

        end_date = self.request.query_params.get('end_date')
        if end_date:
            qs = qs.filter(tanggal__lte=end_date)

        # Pencarian keyword
        search = self.request.query_params.get('search') or self.request.query_params.get('q')
        if search:
            q = search.strip()
            qs = qs.filter(
                Q(deskripsi__icontains=q) |
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q) |
                Q(user__username__icontains=q) |
                Q(user__unit__nama__icontains=q)
            )

        return qs.order_by('-tanggal', '-jam_mulai', '-created_at')

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return LogbookInputSerializer
        return LogbookSerializer

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def perform_update(self, serializer):
        instance = self.get_object()
        # Cegah user lain mengedit logbook orang lain (kecuali superuser)
        if instance.user != self.request.user and not self.request.user.is_superuser:
            raise PermissionDenied('Anda hanya dapat mengedit logbook milik Anda sendiri.')
            
        # Kunci retroaktif
        if not self.request.user.is_superuser:
            selisih = (timezone.localdate() - instance.tanggal).days
            if selisih > 3:
                raise PermissionDenied('Logbook lebih dari 3 hari yang lalu tidak dapat diubah.')
        serializer.save()

    def perform_destroy(self, instance):
        if instance.user != self.request.user and not self.request.user.is_superuser:
            raise PermissionDenied('Anda hanya dapat menghapus logbook milik Anda sendiri.')
            
        if not self.request.user.is_superuser:
            selisih = (timezone.localdate() - instance.tanggal).days
            if selisih > 3:
                raise PermissionDenied('Logbook lebih dari 3 hari yang lalu tidak dapat dihapus.')
        instance.delete()

    @action(detail=False, methods=['get'])
    def monitoring_summary(self, request):
        """Ringkasan eksekutif khusus Wadir & Direktur"""
        if get_monitoring_level(request.user) is None:
            raise PermissionDenied('Anda tidak memiliki akses ke monitoring.')

        today = timezone.localdate()
        first_day_of_month = today.replace(day=1)

        today_qs = Logbook.objects.filter(tanggal=today)
        month_qs = Logbook.objects.filter(tanggal__gte=first_day_of_month, tanggal__lte=today)

        # Hitung total menit hari ini
        total_menit_today = sum(item.durasi_menit for item in today_qs)
        jam_today = total_menit_today // 60
        sisa_menit_today = total_menit_today % 60
        durasi_today_str = f"{jam_today}j {sisa_menit_today}m" if jam_today > 0 else f"{sisa_menit_today}m"

        data = {
            'today_date': today.strftime('%Y-%m-%d'),
            'today_total_entries': today_qs.count(),
            'today_active_users': today_qs.values('user_id').distinct().count(),
            'today_total_minutes': total_menit_today,
            'today_durasi_format': durasi_today_str,
            'month_total_entries': month_qs.count(),
            'month_active_users': month_qs.values('user_id').distinct().count(),
        }
        return Response(data)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export rekap logbook ke Excel (.xlsx) dengan filter aktif"""
        qs = self.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rekap Logbook"

        # Styling
        font_title = Font(name='Arial', size=14, bold=True, color='0F172A')
        font_meta = Font(name='Arial', size=9, italic=True, color='475569')
        font_th = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        font_td = Font(name='Arial', size=9, color='0F172A')
        font_td_bold = Font(name='Arial', size=9, bold=True, color='0F172A')

        fill_th = PatternFill(start_color='0284C7', end_color='0284C7', fill_type='solid')
        fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1'),
        )

        # Header Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "REKAP LOGBOOK PEKERJAAN HARIAN PEGAWAI"
        ws['A1'].font = font_title
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:H2')
        cetak_str = f"Dicetak pada: {timezone.localtime().strftime('%d/%m/%Y %H:%M')} | Oleh: {request.user.get_full_name() or request.user.username}"
        ws['A2'] = cetak_str
        ws['A2'].font = font_meta
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.append([]) # Empty row

        # Table Header
        headers = [
            "No",
            "Tanggal",
            "Jam Kerja",
            "Durasi",
            "Nama Pegawai",
            "Unit / Bagian",
            "Role",
            "Uraian / Deskripsi Pekerjaan"
        ]
        ws.append(headers)

        header_row = 4
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = font_th
            cell.fill = fill_th
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Populate Data
        current_row = 5
        for idx, item in enumerate(qs, 1):
            user_nama = f"{item.user.first_name or ''} {item.user.last_name or ''}".strip() or item.user.username
            unit_nama = item.user.unit.nama if item.user.unit else '-'
            jam_str = f"{item.jam_mulai.strftime('%H:%M')} - {item.jam_selesai.strftime('%H:%M')}"
            durasi_str = item.durasi_format

            row_data = [
                idx,
                item.tanggal.strftime('%d/%m/%Y'),
                jam_str,
                durasi_str,
                user_nama,
                unit_nama,
                item.user.get_role_display(),
                item.deskripsi
            ]
            ws.append(row_data)

            is_even = (idx % 2 == 0)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = font_td
                cell.border = thin_border
                if is_even:
                    cell.fill = fill_zebra

                if col_idx in (1, 2, 3, 4):
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                elif col_idx in (5, 6, 7):
                    cell.alignment = Alignment(horizontal='left', vertical='top')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            current_row += 1

        # Auto-adjust column widths
        col_widths = {
            1: 6,   # No
            2: 13,  # Tanggal
            3: 16,  # Jam Kerja
            4: 14,  # Durasi
            5: 24,  # Nama Pegawai
            6: 22,  # Unit
            7: 16,  # Role
            8: 55,  # Deskripsi
        }
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Rekap_Logbook_{timezone.localdate().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


def _hitung_total_task(task):
    sesi_selesai = task.sesi_list.filter(selesai__isnull=False)
    total_kerja = sum(s.durasi_kerja for s in sesi_selesai)
    total_lembur = sum(s.durasi_lembur for s in sesi_selesai)
    task.total_menit_kerja = total_kerja
    task.total_menit_lembur = total_lembur
    task.save()


class TaskViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        user = self.request.user
        qs = Task.objects.select_related('user', 'user__unit').prefetch_related('sesi_list')
        monitoring_level = get_monitoring_level(user)

        # Filter akses data
        if monitoring_level is None:
            qs = qs.filter(user=user)
        elif monitoring_level == 'unit':
            qs = qs.filter(user__unit=user.unit)
        else:
            user_id = self.request.query_params.get('user_id')
            if user_id and user_id.isdigit():
                qs = qs.filter(user_id=int(user_id))

            unit_id = self.request.query_params.get('unit_id')
            if unit_id and unit_id.isdigit():
                qs = qs.filter(user__unit_id=int(unit_id))

        status_param = self.request.query_params.get('status')
        if status_param:
            statuses = status_param.split(',')
            qs = qs.filter(status__in=statuses)
            
        search = self.request.query_params.get('search') or self.request.query_params.get('q')
        if search:
            q = search.strip()
            qs = qs.filter(
                Q(judul__icontains=q) |
                Q(no_task__icontains=q) |
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q) |
                Q(user__username__icontains=q)
            )

        return qs.order_by('-updated_at')

    def get_serializer_class(self):
        if self.action == 'create':
            return TaskCreateSerializer
        return TaskSerializer

    def perform_create(self, serializer):
        task = serializer.save(user=self.request.user, started_at=timezone.now(), status='on_progress')
        # Buat sesi kerja pertama
        SesiKerja.objects.create(task=task, mulai=timezone.now())

    def perform_destroy(self, instance):
        if instance.user != self.request.user and not self.request.user.is_superuser:
            raise PermissionDenied('Anda hanya dapat menghapus task milik Anda sendiri.')
        instance.delete()

    @action(detail=True, methods=['post'])
    def pause(self, request, pk=None):
        task = self.get_object()
        if task.status == 'done':
            return Response({'error': 'Task sudah selesai.'}, status=400)
            
        sesi_aktif = task.sesi_list.filter(selesai__isnull=True).first()
        if not sesi_aktif:
            return Response({'error': 'Tidak ada sesi yang sedang berjalan'}, status=400)
        
        now = timezone.now()
        sesi_aktif.selesai = now
        kerja, lembur = hitung_durasi_sesi(sesi_aktif.mulai, sesi_aktif.selesai)
        sesi_aktif.durasi_kerja = kerja
        sesi_aktif.durasi_lembur = lembur
        sesi_aktif.save()
        
        task.status = 'on_hold'
        _hitung_total_task(task)
        
        return Response(TaskSerializer(task).data)

    @action(detail=True, methods=['post'])
    def resume(self, request, pk=None):
        task = self.get_object()
        if task.status == 'done':
            return Response({'error': 'Task sudah selesai.'}, status=400)
            
        if task.sesi_list.filter(selesai__isnull=True).exists():
            return Response({'error': 'Task sudah memiliki sesi aktif.'}, status=400)
            
        SesiKerja.objects.create(task=task, mulai=timezone.now())
        task.status = 'on_progress'
        task.save()
        
        return Response(TaskSerializer(task).data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        task = self.get_object()
        if task.status == 'done':
            return Response({'error': 'Task sudah selesai.'}, status=400)
            
        now = timezone.now()
        sesi_aktif = task.sesi_list.filter(selesai__isnull=True).first()
        
        if sesi_aktif:
            sesi_aktif.selesai = now
            kerja, lembur = hitung_durasi_sesi(sesi_aktif.mulai, sesi_aktif.selesai)
            sesi_aktif.durasi_kerja = kerja
            sesi_aktif.durasi_lembur = lembur
            sesi_aktif.save()
            
        task.status = 'done'
        task.completed_at = now
        _hitung_total_task(task)
        
        return Response(TaskSerializer(task).data)

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Mendapatkan daftar task yang sedang berjalan (untuk timer global)"""
        qs = self.get_queryset().filter(status__in=['on_progress', 'on_hold'])
        return Response(TaskSerializer(qs, many=True).data)
